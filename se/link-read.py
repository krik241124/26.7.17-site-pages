# semrush_backlink_matrix.py
"""
Semrush Backlink Intelligence Engine

Input: Semrush Backlinks CSV/TSV exports with columns such as:
    Page ascore, Source title, Source url, Target url, Anchor,
    External links, Internal links, Nofollow, Sponsored, Ugc,
    Text, Frame, Form, Image, Sitewide,
    First seen, Last seen, New link, Lost link

Design principles:
1. One exported backlink = one fact row. Never collapse rows by referring domain at import time.
2. Referring-domain intelligence is a separate aggregation layer.
3. Page AS is a PAGE-level Semrush metric, so it is never mislabeled as domain DR.
4. "Live" means Lost link == false in the current Semrush snapshot.
5. Sitewide / Sponsored / UGC are explicit Semrush attributes, not guessed spam signals.

Outputs:
    reports/Semrush_Backlink_Intelligence_YYYY-MM-DD.xlsx
    reports/Semrush_Backlink_Opportunity_Payload_YYYY-MM-DD.json
    semrush_backlink_intelligence.db
"""

import os
import re
import csv
import json
import glob
import hashlib
import sqlite3
import datetime
from collections import Counter, defaultdict
from statistics import mean
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import tldextract
    HAS_TLDEXTRACT = True
except ImportError:
    HAS_TLDEXTRACT = False


# =============================
# Configuration
# =============================
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "semrush_backlink_intelligence.db")
DATA_DIR = os.path.join(WORKSPACE_DIR, "data")
REPORT_DIR = os.path.join(WORKSPACE_DIR, "reports")

OVERLAP_THRESHOLD = 1
HIGH_PAGE_AS_THRESHOLD = 40

AS_BANDS = [
    ("0-20", 0, 20),
    ("21-40", 21, 40),
    ("41-60", 41, 60),
    ("61-80", 61, 80),
    ("81-100", 81, 100),
]

SEM_RUSH_REQUIRED = {"source url", "target url"}
SEM_RUSH_RECOMMENDED = {
    "page ascore", "source title", "anchor", "external links", "internal links",
    "nofollow", "sponsored", "ugc", "text", "frame", "form", "image",
    "sitewide", "first seen", "last seen", "new link", "lost link",
}

COLOR_DARK = "1F4E78"
COLOR_BLUE = "2F5597"
COLOR_LIGHT_BLUE = "D9EAF7"
COLOR_RED = "C00000"
COLOR_GRAY = "F2F2F2"
COLOR_GREEN = "70AD47"
COLOR_ORANGE = "F4B183"
PALETTE = ["294266", "F7941D", "20B799", "FDB913", "3FBBDF", "8C67AB", "E5625E", "8391A5"]


def utc_now_iso():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def safe_float(value, default=0.0):
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def safe_int(value, default=0):
    try:
        if value is None or str(value).strip() == "":
            return default
        return int(float(str(value).replace(",", "").strip()))
    except (ValueError, TypeError):
        return default


def parse_bool(value):
    if isinstance(value, bool):
        return 1 if value else 0
    text = str(value or "").strip().lower()
    if text in {"true", "1", "yes", "y", "t"}:
        return 1
    if text in {"false", "0", "no", "n", "f", "", "none", "null", "nan"}:
        return 0
    return 0


def normalize_date(value):
    """Return YYYY-MM-DD where possible, otherwise preserve clean source text."""
    text = str(value or "").strip()
    if not text:
        return ""
    # Fast path: YYYY-MM-DD / YYYY/MM/DD / ISO datetime
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return text
    # Common export variants
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y", "%b %d, %Y", "%d %b %Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def year_from_date(value):
    text = str(value or "")
    return text[:4] if len(text) >= 4 and text[:4].isdigit() else None


def clean_url(url):
    return str(url or "").strip()


def extract_hostname(url):
    try:
        text = clean_url(url)
        if not text:
            return ""
        parsed = urlparse(text if "://" in text else f"https://{text}")
        host = (parsed.hostname or "").lower().strip(".")
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""


def extract_registrable_domain(url):
    """
    Prefer registrable/root domain when tldextract is installed.
    Fallback: normalized hostname. No network call is required by the fallback.
    """
    host = extract_hostname(url)
    if not host:
        return ""
    if HAS_TLDEXTRACT:
        try:
            ext = tldextract.extract(host)
            if ext.domain and ext.suffix:
                return f"{ext.domain}.{ext.suffix}".lower()
        except Exception:
            pass
    return host


def canonical_url(url):
    """Lightweight normalization for duplicate protection; query string is preserved."""
    text = clean_url(url)
    if not text:
        return ""
    try:
        parsed = urlparse(text)
        scheme = (parsed.scheme or "https").lower()
        host = (parsed.hostname or "").lower()
        if host.startswith("www."):
            host = host[4:]
        port = f":{parsed.port}" if parsed.port else ""
        path = parsed.path or "/"
        if path != "/" and path.endswith("/"):
            path = path[:-1]
        query = f"?{parsed.query}" if parsed.query else ""
        return f"{scheme}://{host}{port}{path}{query}"
    except Exception:
        return text.lower()


def primary_link_type(row):
    """Create one mutually exclusive placement type for profile charts."""
    if row.get("is_image"):
        return "Image"
    if row.get("is_form"):
        return "Form"
    if row.get("is_frame"):
        return "Frame"
    if row.get("is_text"):
        return "Text"
    return "Unknown"


def is_follow_link(row):
    """Treat a clean link with none of nofollow/sponsored/UGC attributes as Follow."""
    return int(not (row.get("is_nofollow") or row.get("is_sponsored") or row.get("is_ugc")))


def as_band(score):
    score = safe_float(score, 0)
    for label, lo, hi in AS_BANDS:
        if lo <= score <= hi:
            return label
    return "81-100" if score > 100 else "0-20"


def read_delimited_file(file_path):
    """Read CSV/TSV defensively and return (rows, cleaned_headers)."""
    attempts = [
        ("utf-8-sig", "\t" if file_path.lower().endswith(".tsv") else ","),
        ("utf-16", "\t"),
        ("utf-8", "\t" if file_path.lower().endswith(".tsv") else ","),
        ("latin1", ","),
    ]
    last_error = None
    for encoding, delimiter in attempts:
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                if not reader.fieldnames or len(reader.fieldnames) <= 1:
                    continue
                rows = []
                headers = [str(h).replace("\ufeff", "").strip() for h in reader.fieldnames if h]
                for raw in reader:
                    rows.append({str(k).replace("\ufeff", "").strip(): v for k, v in raw.items() if k})
                return rows, headers
        except Exception as exc:
            last_error = exc
    raise ValueError(f"无法解析文件 {file_path}: {last_error}")


def infer_competitor_domain(file_path, rows):
    """Prefer the dominant Target URL domain; filename is only a fallback."""
    target_domains = []
    for raw in rows:
        lower = {str(k).strip().lower(): v for k, v in raw.items()}
        d = extract_registrable_domain(lower.get("target url", ""))
        if d:
            target_domains.append(d)
    if target_domains:
        domain, count = Counter(target_domains).most_common(1)[0]
        if count / len(target_domains) >= 0.50:
            return domain

    name = os.path.splitext(os.path.basename(file_path))[0].lower()
    name = re.sub(r"(?i)(^backlinks?[-_ ]*|[-_ ]*backlinks?.*$)", "", name)
    name = re.sub(r"(?i)([-_ ]*export.*$|[-_ ]*semrush.*$)", "", name)
    name = name.strip("-_ .")
    return name or "unknown-competitor"


def backlink_fingerprint(row):
    identity = "|".join([
        row.get("source_url", ""),
        row.get("target_url", ""),
        row.get("anchor", ""),
        str(row.get("is_nofollow", 0)),
        str(row.get("is_sponsored", 0)),
        str(row.get("is_ugc", 0)),
        str(row.get("is_text", 0)),
        str(row.get("is_frame", 0)),
        str(row.get("is_form", 0)),
        str(row.get("is_image", 0)),
    ]).lower()
    return hashlib.sha1(identity.encode("utf-8", errors="ignore")).hexdigest()


# =============================
# Database
# =============================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS backlinks")
    cur.execute("""
        CREATE TABLE backlinks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competitor_domain TEXT NOT NULL,
            ref_domain TEXT NOT NULL,
            source_title TEXT,
            source_url TEXT NOT NULL,
            target_url TEXT,
            anchor TEXT,
            page_ascore REAL DEFAULT 0,
            external_links INTEGER DEFAULT 0,
            internal_links INTEGER DEFAULT 0,
            is_nofollow INTEGER DEFAULT 0,
            is_sponsored INTEGER DEFAULT 0,
            is_ugc INTEGER DEFAULT 0,
            is_text INTEGER DEFAULT 0,
            is_frame INTEGER DEFAULT 0,
            is_form INTEGER DEFAULT 0,
            is_image INTEGER DEFAULT 0,
            is_sitewide INTEGER DEFAULT 0,
            is_new INTEGER DEFAULT 0,
            is_lost INTEGER DEFAULT 0,
            first_seen TEXT,
            last_seen TEXT,
            link_fingerprint TEXT NOT NULL,
            raw_data TEXT,
            imported_at TEXT,
            UNIQUE(competitor_domain, link_fingerprint)
        )
    """)
    cur.execute("CREATE INDEX idx_backlinks_competitor ON backlinks(competitor_domain)")
    cur.execute("CREATE INDEX idx_backlinks_ref_domain ON backlinks(ref_domain)")
    cur.execute("CREATE INDEX idx_backlinks_lost ON backlinks(is_lost)")
    cur.execute("CREATE INDEX idx_backlinks_first_seen ON backlinks(first_seen)")
    conn.commit()
    return conn


def normalize_semrush_row(raw, competitor):
    lower = {str(k).strip().lower(): v for k, v in raw.items()}
    source_url = canonical_url(lower.get("source url", ""))
    target_url = canonical_url(lower.get("target url", ""))
    ref_domain = extract_registrable_domain(source_url)
    if not source_url or not ref_domain:
        return None

    row = {
        "competitor_domain": competitor,
        "ref_domain": ref_domain,
        "source_title": str(lower.get("source title", "") or "").strip(),
        "source_url": source_url,
        "target_url": target_url,
        "anchor": str(lower.get("anchor", "") or "").strip(),
        "page_ascore": safe_float(lower.get("page ascore", lower.get("page as", 0))),
        "external_links": safe_int(lower.get("external links", 0)),
        "internal_links": safe_int(lower.get("internal links", 0)),
        "is_nofollow": parse_bool(lower.get("nofollow")),
        "is_sponsored": parse_bool(lower.get("sponsored")),
        "is_ugc": parse_bool(lower.get("ugc")),
        "is_text": parse_bool(lower.get("text")),
        "is_frame": parse_bool(lower.get("frame")),
        "is_form": parse_bool(lower.get("form")),
        "is_image": parse_bool(lower.get("image")),
        "is_sitewide": parse_bool(lower.get("sitewide")),
        "is_new": parse_bool(lower.get("new link", lower.get("newlink"))),
        "is_lost": parse_bool(lower.get("lost link", lower.get("lostlink"))),
        "first_seen": normalize_date(lower.get("first seen", "")),
        "last_seen": normalize_date(lower.get("last seen", "")),
        "raw_data": json.dumps(raw, ensure_ascii=False),
        "imported_at": utc_now_iso(),
    }
    row["link_fingerprint"] = backlink_fingerprint(row)
    return row


def process_semrush_exports(conn):
    os.makedirs(DATA_DIR, exist_ok=True)
    files = sorted(
        glob.glob(os.path.join(DATA_DIR, "*.csv"))
        + glob.glob(os.path.join(DATA_DIR, "*.tsv"))
    )
    if not files:
        print(f"[提示] 请将 Semrush Backlinks CSV/TSV 放入：{DATA_DIR}")
        return 0, [], []

    cur = conn.cursor()
    global_headers = []
    import_log = []
    processed_files = 0

    insert_sql = """
        INSERT OR IGNORE INTO backlinks (
            competitor_domain, ref_domain, source_title, source_url, target_url, anchor,
            page_ascore, external_links, internal_links,
            is_nofollow, is_sponsored, is_ugc,
            is_text, is_frame, is_form, is_image, is_sitewide,
            is_new, is_lost, first_seen, last_seen,
            link_fingerprint, raw_data, imported_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """

    for file_path in files:
        try:
            raw_rows, headers = read_delimited_file(file_path)
        except Exception as exc:
            print(f"[警告] 跳过无法读取的文件 {os.path.basename(file_path)}: {exc}")
            import_log.append((os.path.basename(file_path), "ERROR", 0, 0, str(exc)))
            continue

        if not raw_rows:
            import_log.append((os.path.basename(file_path), "EMPTY", 0, 0, "无数据行"))
            continue

        lower_headers = {h.lower() for h in headers}
        missing_required = sorted(SEM_RUSH_REQUIRED - lower_headers)
        if missing_required:
            msg = f"缺少必要列: {', '.join(missing_required)}"
            print(f"[警告] {os.path.basename(file_path)} {msg}，已跳过。")
            import_log.append((os.path.basename(file_path), "SKIPPED", len(raw_rows), 0, msg))
            continue

        competitor = infer_competitor_domain(file_path, raw_rows)
        print(f"正在导入 Semrush 外链：[{competitor}] <- {os.path.basename(file_path)}")

        for h in headers:
            if h not in global_headers:
                global_headers.append(h)

        normalized = []
        rejected = 0
        for raw in raw_rows:
            row = normalize_semrush_row(raw, competitor)
            if not row:
                rejected += 1
                continue
            normalized.append((
                row["competitor_domain"], row["ref_domain"], row["source_title"], row["source_url"],
                row["target_url"], row["anchor"], row["page_ascore"], row["external_links"],
                row["internal_links"], row["is_nofollow"], row["is_sponsored"], row["is_ugc"],
                row["is_text"], row["is_frame"], row["is_form"], row["is_image"], row["is_sitewide"],
                row["is_new"], row["is_lost"], row["first_seen"], row["last_seen"],
                row["link_fingerprint"], row["raw_data"], row["imported_at"],
            ))

        before = conn.total_changes
        cur.executemany(insert_sql, normalized)
        conn.commit()
        inserted = conn.total_changes - before
        duplicates = max(0, len(normalized) - inserted)
        processed_files += 1

        missing_recommended = sorted(SEM_RUSH_RECOMMENDED - lower_headers)
        note_parts = []
        if rejected:
            note_parts.append(f"无效URL {rejected} 行")
        if duplicates:
            note_parts.append(f"去重 {duplicates} 行")
        if missing_recommended:
            note_parts.append("缺少可选列: " + ", ".join(missing_recommended))
        note = "；".join(note_parts) if note_parts else "OK"
        import_log.append((os.path.basename(file_path), competitor, len(raw_rows), inserted, note))

    return processed_files, global_headers, import_log


# =============================
# Aggregation
# =============================
def fetch_rows(conn):
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM backlinks ORDER BY competitor_domain, ref_domain, source_url")
    rows = [dict(r) for r in cur.fetchall()]
    conn.row_factory = None
    return rows


def summarize_domain(rows):
    live = [r for r in rows if not r["is_lost"]]
    page_as = [safe_float(r["page_ascore"]) for r in rows]
    live_as = [safe_float(r["page_ascore"]) for r in live]
    first_dates = sorted([r["first_seen"] for r in rows if r["first_seen"]])
    last_dates = sorted([r["last_seen"] for r in rows if r["last_seen"]])
    types = Counter(primary_link_type(r) for r in rows)

    return {
        "backlinks": len(rows),
        "live_backlinks": len(live),
        "source_pages": len({r["source_url"] for r in rows}),
        "target_pages": len({r["target_url"] for r in rows if r["target_url"]}),
        "max_page_as": max(page_as) if page_as else 0,
        "avg_page_as": mean(page_as) if page_as else 0,
        "live_avg_page_as": mean(live_as) if live_as else 0,
        "follow_rate": sum(is_follow_link(r) for r in rows) / len(rows) if rows else 0,
        "live_follow_rate": sum(is_follow_link(r) for r in live) / len(live) if live else 0,
        "sitewide_rate": sum(r["is_sitewide"] for r in rows) / len(rows) if rows else 0,
        "sponsored_rate": sum(r["is_sponsored"] for r in rows) / len(rows) if rows else 0,
        "ugc_rate": sum(r["is_ugc"] for r in rows) / len(rows) if rows else 0,
        "avg_external_links": mean([r["external_links"] for r in rows]) if rows else 0,
        "earliest_first_seen": first_dates[0] if first_dates else "",
        "latest_last_seen": last_dates[-1] if last_dates else "",
        "types": types,
    }


def build_intelligence(rows):
    by_comp = defaultdict(list)
    by_global_domain = defaultdict(list)
    by_comp_domain = defaultdict(lambda: defaultdict(list))

    for r in rows:
        by_comp[r["competitor_domain"]].append(r)
        by_global_domain[r["ref_domain"]].append(r)
        by_comp_domain[r["competitor_domain"]][r["ref_domain"]].append(r)

    comp_stats = {}
    comp_time = {}

    for comp, comp_rows in by_comp.items():
        live = [r for r in comp_rows if not r["is_lost"]]
        domain_groups = by_comp_domain[comp]
        live_domains = {d for d, rs in domain_groups.items() if any(not r["is_lost"] for r in rs)}
        dead_domains = set(domain_groups) - live_domains

        hist_as_dist = Counter(as_band(r["page_ascore"]) for r in comp_rows)
        live_as_dist = Counter(as_band(r["page_ascore"]) for r in live)
        live_type_dist = Counter(primary_link_type(r) for r in live)

        comp_stats[comp] = {
            "total_domains": len(domain_groups),
            "total_backlinks": len(comp_rows),
            "live_domains": len(live_domains),
            "lost_domains": len(dead_domains),
            "live_backlinks": len(live),
            "lost_backlinks": len(comp_rows) - len(live),
            "new_flagged_backlinks": sum(r["is_new"] for r in comp_rows),
            "unique_target_pages": len({r["target_url"] for r in comp_rows if r["target_url"]}),
            "follow_rate_live": sum(is_follow_link(r) for r in live) / len(live) if live else 0,
            "nofollow_rate_live": sum(r["is_nofollow"] for r in live) / len(live) if live else 0,
            "sitewide_rate_live": sum(r["is_sitewide"] for r in live) / len(live) if live else 0,
            "sponsored_rate_live": sum(r["is_sponsored"] for r in live) / len(live) if live else 0,
            "ugc_rate_live": sum(r["is_ugc"] for r in live) / len(live) if live else 0,
            "high_as_rate_live": sum(safe_float(r["page_ascore"]) >= HIGH_PAGE_AS_THRESHOLD for r in live) / len(live) if live else 0,
            "avg_page_as_live": mean([safe_float(r["page_ascore"]) for r in live]) if live else 0,
            "link_concentration": len(comp_rows) / len(domain_groups) if domain_groups else 0,
            "live_link_concentration": len(live) / len(live_domains) if live_domains else 0,
            "hist_as_dist": hist_as_dist,
            "live_as_dist": live_as_dist,
            "live_type_dist": live_type_dist,
        }

        backlink_first = Counter()
        live_backlink_first = Counter()
        domain_first = Counter()
        live_domain_first = Counter()

        for r in comp_rows:
            y = year_from_date(r["first_seen"])
            if y:
                backlink_first[y] += 1
                if not r["is_lost"]:
                    live_backlink_first[y] += 1

        for domain, drs in domain_groups.items():
            valid_dates = sorted(r["first_seen"] for r in drs if r["first_seen"])
            if not valid_dates:
                continue
            y = year_from_date(valid_dates[0])
            if y:
                domain_first[y] += 1
                if domain in live_domains:
                    live_domain_first[y] += 1

        comp_time[comp] = {
            "backlink_first": backlink_first,
            "live_backlink_first": live_backlink_first,
            "domain_first": domain_first,
            "live_domain_first": live_domain_first,
        }

    domain_intel = []
    for ref_domain, domain_rows in by_global_domain.items():
        summary = summarize_domain(domain_rows)
        competitors = sorted({r["competitor_domain"] for r in domain_rows})
        if len(competitors) < OVERLAP_THRESHOLD:
            continue

        examples = []
        for r in sorted(domain_rows, key=lambda x: (x["is_lost"], -safe_float(x["page_ascore"]), x["source_url"])):
            status = "LOST" if r["is_lost"] else "LIVE"
            attr = []
            if is_follow_link(r): attr.append("Follow")
            if r["is_nofollow"]: attr.append("Nofollow")
            if r["is_sponsored"]: attr.append("Sponsored")
            if r["is_ugc"]: attr.append("UGC")
            if r["is_sitewide"]: attr.append("Sitewide")
            examples.append(
                f"[{r['competitor_domain']}] [{status}] AS {safe_float(r['page_ascore']):.0f} | "
                f"{r['source_url']} --[{r['anchor'] or '无锚文本'}]--> {r['target_url']} "
                f"({'/'.join(attr) or 'No special attribute'})"
            )
        # Avoid cells becoming unusably huge while retaining representative evidence.
        examples = list(dict.fromkeys(examples))[:20]

        domain_intel.append({
            "ref_domain": ref_domain,
            "alpha": len(competitors),
            "competitors": competitors,
            **summary,
            "examples": examples,
            "rows": domain_rows,
        })

    # Priority is descriptive: broad competitor overlap + active links + source-page authority,
    # while heavy sitewide exposure is pushed lower. No unsupported "spam" label is invented.
    domain_intel.sort(
        key=lambda d: (
            d["alpha"],
            d["live_backlinks"],
            d["max_page_as"],
            d["live_follow_rate"],
            -d["sitewide_rate"],
        ),
        reverse=True,
    )

    return by_comp, by_comp_domain, comp_stats, comp_time, domain_intel


# =============================
# Excel helpers
# =============================
def style_header(ws, row, fill=COLOR_DARK, font_color="FFFFFF"):
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True, color=font_color)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_title(ws, text, width_cols=12, fill=COLOR_BLUE):
    ws.append([text])
    r = ws.max_row
    ws.cell(r, 1).font = Font(bold=True, size=15, color="FFFFFF")
    for c in range(1, width_cols + 1):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
    return r


def percent_fmt(ws, row, cols):
    for c in cols:
        ws.cell(row=row, column=c).number_format = "0.0%"


def cumulative(values):
    out, total = [], 0
    for v in values:
        total += v
        out.append(total)
    return out


def create_line_chart(ws, title, data_min_row, data_max_row, min_col, max_col, cats_row, anchor, width=18, height=9):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height
    data = Reference(ws, min_col=min_col, min_row=data_min_row, max_col=max_col, max_row=data_max_row)
    cats = Reference(ws, min_col=min_col + 1, min_row=cats_row, max_col=max_col, max_row=cats_row)
    chart.add_data(data, from_rows=True, titles_from_data=True)
    chart.set_categories(cats)
    for i, s in enumerate(chart.series):
        try:
            s.graphicalProperties.line.solidFill = PALETTE[i % len(PALETTE)]
            s.graphicalProperties.line.width = 26000
        except Exception:
            pass
    ws.add_chart(chart, anchor)


def create_bar_chart(ws, title, data_range, cats_range, anchor, stacked=False, width=18, height=9):
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.grouping = "stacked" if stacked else "clustered"
    if stacked:
        chart.overlap = 100
    chart.title = title
    chart.width = width
    chart.height = height
    chart.add_data(data_range, titles_from_data=True)
    chart.set_categories(cats_range)
    for i, s in enumerate(chart.series):
        try:
            s.graphicalProperties.solidFill = PALETTE[i % len(PALETTE)]
        except Exception:
            pass
    ws.add_chart(chart, anchor)


# =============================
# Reports
# =============================
def generate_reports(conn, global_headers, import_log):
    os.makedirs(REPORT_DIR, exist_ok=True)
    rows = fetch_rows(conn)
    if not rows:
        print("[提示] 数据库中没有有效 Semrush backlink 行，未生成报告。")
        return None, None

    by_comp, by_comp_domain, comp_stats, comp_time, domain_intel = build_intelligence(rows)
    comp_list = sorted(comp_stats)
    all_years = sorted({
        y
        for comp in comp_time.values()
        for counter in comp.values()
        for y in counter.keys()
    })

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    excel_filename = os.path.join(REPORT_DIR, f"Semrush_Backlink_Intelligence_{date_str}.xlsx")
    json_filename = os.path.join(REPORT_DIR, f"Semrush_Backlink_Opportunity_Payload_{date_str}.json")

    payload = []
    for d in domain_intel:
        payload.append({
            "ref_domain": d["ref_domain"],
            "competitor_overlap_alpha": d["alpha"],
            "competitors": d["competitors"],
            "backlinks": d["backlinks"],
            "live_backlinks": d["live_backlinks"],
            "max_source_page_as": round(d["max_page_as"], 1),
            "avg_source_page_as": round(d["avg_page_as"], 1),
            "follow_rate": round(d["follow_rate"], 4),
            "sitewide_rate": round(d["sitewide_rate"], 4),
            "sponsored_rate": round(d["sponsored_rate"], 4),
            "ugc_rate": round(d["ugc_rate"], 4),
            "earliest_first_seen": d["earliest_first_seen"],
            "latest_last_seen": d["latest_last_seen"],
            "examples": d["examples"][:5],
        })

    with open(json_filename, "w", encoding="utf-8") as f:
        json.dump({"generated_at": utc_now_iso(), "data": payload}, f, ensure_ascii=False, indent=2)

    if not HAS_OPENPYXL:
        print("[警告] 未安装 openpyxl，只生成 JSON，不生成 Excel。")
        return None, json_filename

    wb = Workbook()

    # --------------------------------------------------
    # Sheet 1: Referring-domain Alpha Matrix
    # --------------------------------------------------
    ws1 = wb.active
    ws1.title = "Alpha来源域名矩阵"
    ws1.sheet_view.showGridLines = False
    headers1 = [
        "来源域名 (Ref Domain)", "竞品重合度 Alpha", "被哪些竞品获取",
        "Backlink 行数", "Live Backlinks", "来源页面数", "指向目标页数",
        "最高来源页 Page AS", "平均来源页 Page AS", "Follow 占比",
        "Sitewide 占比", "Sponsored 占比", "UGC 占比",
        "来源页平均 External Links", "最早 First Seen", "最近 Last Seen",
        "Link Type Mix", "具体外链范例（最多20条）",
    ]
    ws1.append(headers1)
    style_header(ws1, 1, fill=COLOR_RED)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:R{max(1, len(domain_intel)+1)}"
    set_widths(ws1, {
        "A": 30, "B": 14, "C": 36, "D": 14, "E": 14, "F": 14, "G": 14,
        "H": 18, "I": 18, "J": 13, "K": 13, "L": 13, "M": 13, "N": 19,
        "O": 14, "P": 14, "Q": 25, "R": 105,
    })

    for d in domain_intel:
        type_mix = ", ".join(f"{k}:{v}" for k, v in d["types"].most_common())
        ws1.append([
            d["ref_domain"], d["alpha"], ", ".join(d["competitors"]),
            d["backlinks"], d["live_backlinks"], d["source_pages"], d["target_pages"],
            round(d["max_page_as"], 1), round(d["avg_page_as"], 1), d["follow_rate"],
            d["sitewide_rate"], d["sponsored_rate"], d["ugc_rate"],
            round(d["avg_external_links"], 1), d["earliest_first_seen"], d["latest_last_seen"],
            type_mix, "\n".join(d["examples"]),
        ])
        r = ws1.max_row
        percent_fmt(ws1, r, [10, 11, 12, 13])
        ws1.cell(r, 18).alignment = Alignment(wrap_text=True, vertical="top")
        ws1.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")

    # --------------------------------------------------
    # Sheet 2: Full backlink detail + derived fields
    # --------------------------------------------------
    ws2 = wb.create_sheet("外链全量明细")
    ws2.sheet_view.showGridLines = False
    derived_headers = [
        "竞品", "来源域名", "Alpha", "Live?", "Follow?", "Primary Link Type", "Page AS Band"
    ]
    ws2.append(derived_headers + global_headers)
    style_header(ws2, 1, fill="4F81BD")
    ws2.freeze_panes = "A2"
    domain_alpha = {d["ref_domain"]: d["alpha"] for d in domain_intel}

    for r in rows:
        raw = json.loads(r["raw_data"] or "{}")
        row_out = [
            r["competitor_domain"], r["ref_domain"], domain_alpha.get(r["ref_domain"], 1),
            "LIVE" if not r["is_lost"] else "LOST",
            "Follow" if is_follow_link(r) else "Attributed",
            primary_link_type(r), as_band(r["page_ascore"]),
        ]
        for h in global_headers:
            row_out.append(raw.get(h, ""))
        ws2.append(row_out)

    if ws2.max_row >= 1:
        ws2.auto_filter.ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"
    set_widths(ws2, {"A": 25, "B": 28, "C": 10, "D": 10, "E": 12, "F": 16, "G": 14})

    # --------------------------------------------------
    # Sheet 3: Single competitor dashboards
    # --------------------------------------------------
    ws3 = wb.create_sheet("单竞品深度分析")
    ws3.sheet_view.showGridLines = False
    set_widths(ws3, {"A": 25, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18, "K": 18, "L": 18})

    for comp in comp_list:
        st = comp_stats[comp]
        add_title(ws3, f"📊 {comp.upper()} — Semrush Backlink Profile", 12)
        ws3.append([])

        # Core metrics
        ws3.append([
            "核心指标", "历史 Ref Domains", "历史 Backlinks", "Live Ref Domains", "Live Backlinks",
            "Lost Backlink %", "Lost Domain %", "Live Follow %", "Live Sitewide %",
            "Live High Page AS %", "Live Sponsored %", "Live UGC %",
        ])
        metric_header = ws3.max_row
        style_header(ws3, metric_header, fill=COLOR_LIGHT_BLUE, font_color="000000")
        lost_bkl_pct = st["lost_backlinks"] / st["total_backlinks"] if st["total_backlinks"] else 0
        lost_dom_pct = st["lost_domains"] / st["total_domains"] if st["total_domains"] else 0
        ws3.append([
            "", st["total_domains"], st["total_backlinks"], st["live_domains"], st["live_backlinks"],
            lost_bkl_pct, lost_dom_pct, st["follow_rate_live"], st["sitewide_rate_live"],
            st["high_as_rate_live"], st["sponsored_rate_live"], st["ugc_rate_live"],
        ])
        metric_data = ws3.max_row
        percent_fmt(ws3, metric_data, list(range(6, 13)))
        ws3.append([])

        # Additional diagnostics
        ws3.append(["结构诊断", "Live 平均 Page AS", "Backlinks / Ref Domain", "Live Backlinks / Live Ref Domain", "Target Pages", "Semrush New Link 标记数"])
        diag_header = ws3.max_row
        style_header(ws3, diag_header, fill=COLOR_GRAY, font_color="000000")
        ws3.append(["", round(st["avg_page_as_live"], 1), round(st["link_concentration"], 2), round(st["live_link_concentration"], 2), st["unique_target_pages"], st["new_flagged_backlinks"]])
        ws3.append([])



        # ==================================================
        # 🧭 单竞品核心指标中文化 + 自动诊断解释
        # ==================================================

        # ① 把上方两个核心表头进一步中文化
        metric_headers_cn = [
            "核心指标",
            "历史引荐域名数（Ref Domains）",
            "历史外链总数（Backlinks）",
            "现存引荐域名数（Live Ref Domains）",
            "现存外链总数（Live Backlinks）",
            "外链流失率（Lost Backlink %）",
            "引荐域名流失率（Lost Domain %）",
            "现存 Follow 链接占比",
            "现存全站重复链接占比（Sitewide）",
            "现存高权威来源页占比（High Page AS）",
            "现存赞助链接占比（Sponsored）",
            "现存用户生成链接占比（UGC）",
        ]

        for col, text in enumerate(metric_headers_cn, 1):
            ws3.cell(row=metric_header, column=col, value=text)

        diag_headers_cn = [
            "结构诊断",
            "现存来源页平均权威分（Avg Page AS）",
            "历史平均每个域名产生外链数",
            "现存平均每个域名产生外链数",
            "被外链覆盖的目标页面数（Target Pages）",
            "Semrush 新增链接标记数（New Link Flags）",
        ]

        for col, text in enumerate(diag_headers_cn, 1):
            ws3.cell(row=diag_header, column=col, value=text)


        # --------------------------------------------------
        # ② 自动生成“这个数字意味着什么”
        # --------------------------------------------------

        def pct(v):
            return f"{v:.1%}"

        def backlink_loss_judgement(v):
            if v < 0.10:
                return "较低", "外链整体保持稳定，历史获得的链接大部分仍然存在。"
            elif v < 0.25:
                return "正常", "存在一定自然流失，但暂未表现出明显的大规模外链衰退。"
            elif v < 0.40:
                return "偏高", "外链流失已经比较明显，需要检查高价值来源是否正在持续消失。"
            else:
                return "很高", "大量历史外链已经失效，可能存在旧内容删除、合作结束、页面迁移或低稳定性渠道较多的问题。"

        def domain_loss_judgement(v):
            if v < 0.10:
                return "较低", "来源网站留存很好，说明外链来源关系整体比较稳定。"
            elif v < 0.25:
                return "正常", "有部分来源域名已经完全失去链接，但整体来源网络仍较稳定。"
            elif v < 0.40:
                return "偏高", "失去的不只是单条链接，而是部分来源网站已经完全退出，需要重点分析 Lost Domains。"
            else:
                return "很高", "来源域名层面的流失明显，说明外链渠道稳定性不足，比单纯 Lost Backlink 更值得警惕。"

        def follow_judgement(v):
            if v >= 0.75:
                return "较高", "外链结构以 Follow 为主，说明大部分现存链接没有 nofollow / sponsored / UGC 等限制属性。"
            elif v >= 0.50:
                return "中等", "Follow 与带属性链接并存，链接来源结构相对混合。"
            else:
                return "较低", "较多链接带有 nofollow、sponsored 或 UGC 等属性，需要进一步检查具体来源类型。"

        def sitewide_judgement(v):
            if v < 0.05:
                return "很低", "外链主要来自独立页面，而不是 Footer、Sidebar 等全站重复位置，结构通常更自然。"
            elif v < 0.20:
                return "中等", "存在一定比例的全站重复链接，但尚未主导整个外链结构。"
            else:
                return "较高", "大量链接可能来自 Footer、Sidebar、合作伙伴模板等重复位置，因此总 Backlinks 可能被明显放大。"

        def high_as_judgement(v):
            if v >= 0.20:
                return "较强", "相当一部分现存链接来自较高 Page AS 页面，来源页质量结构相对突出。"
            elif v >= 0.05:
                return "一般", "拥有一部分较高权威来源页面，但主体仍来自普通或低 Page AS 页面。"
            else:
                return "偏低", "高 Page AS 来源页面比例很少。注意 Page AS 是页面级指标，不等同于来源域名本身没有价值。"

        def concentration_judgement(v):
            if v < 2:
                return "非常分散", "平均一个来源域名只有少量链接，外链广度较高、重复度较低。"
            elif v < 5:
                return "较分散", "一个来源域名通常贡献少量外链，整体结构比较健康。"
            elif v < 15:
                return "中等集中", "部分来源域名会重复提供多条链接，需要结合 Sitewide 占比一起判断。"
            else:
                return "高度集中", "大量 Backlinks 集中在较少 Ref Domains 上，总外链数可能明显高于真实渠道广度。"


        lost_bkl_level, lost_bkl_text = backlink_loss_judgement(lost_bkl_pct)
        lost_dom_level, lost_dom_text = domain_loss_judgement(lost_dom_pct)
        follow_level, follow_text = follow_judgement(st["follow_rate_live"])
        sitewide_level, sitewide_text = sitewide_judgement(st["sitewide_rate_live"])
        high_as_level, high_as_text = high_as_judgement(st["high_as_rate_live"])
        conc_level, conc_text = concentration_judgement(st["live_link_concentration"])


        explain_start_row = ws3.max_row + 1
        # --------------------------------------------------
        # ③ 指标解释表
        # --------------------------------------------------

        ws3.merge_cells(
            start_row=ws3.max_row + 1,
            start_column=1,
            end_row=ws3.max_row + 1,
            end_column=12
        )
        explain_title_row = ws3.max_row

        ws3.cell(
            row=explain_title_row,
            column=1,
            value="🧭 核心指标解释与自动诊断｜这些数字通常意味着什么"
        )
        ws3.cell(row=explain_title_row, column=1).font = Font(
            bold=True,
            color="FFFFFF",
            size=12
        )
        ws3.cell(row=explain_title_row, column=1).fill = PatternFill(
            "solid",
            fgColor="4472C4"
        )
        ws3.cell(row=explain_title_row, column=1).alignment = Alignment(
            vertical="center"
        )

        explain_header_row = ws3.max_row + 1

        ws3.cell(explain_header_row, 1, "指标")
        ws3.cell(explain_header_row, 3, "当前数据")
        ws3.cell(explain_header_row, 4, "指标本身代表什么")
        ws3.cell(explain_header_row, 8, "结合当前数据通常意味着什么")

        ws3.merge_cells(
            start_row=explain_header_row,
            start_column=1,
            end_row=explain_header_row,
            end_column=2
        )
        ws3.merge_cells(
            start_row=explain_header_row,
            start_column=4,
            end_row=explain_header_row,
            end_column=7
        )
        ws3.merge_cells(
            start_row=explain_header_row,
            start_column=8,
            end_row=explain_header_row,
            end_column=12
        )

        style_header(
            ws3,
            explain_header_row,
            fill=COLOR_LIGHT_BLUE,
            font_color="000000"
        )


        explanations = [
            (
                "历史引荐域名",
                f'{st["total_domains"]:,}',
                "历史上至少曾有一条外链指向该竞品的不同来源域名数量。它衡量的是外链渠道的“广度”，而不是链接条数。",
                f'历史累计获得 {st["total_domains"]:,} 个来源域名。与 Backlinks 一起看，可以判断竞品是依赖大量不同网站，还是少数网站重复产生大量链接。'
            ),

            (
                "历史外链总数",
                f'{st["total_backlinks"]:,}',
                "Semrush 导出中记录到的历史 backlink 行数。一个来源域名可能贡献很多条外链，因此不能单独用它衡量外链渠道规模。",
                f'历史共记录 {st["total_backlinks"]:,} 条外链，平均每个来源域名约 {st["link_concentration"]:.2f} 条。'
            ),

            (
                "现存引荐域名",
                f'{st["live_domains"]:,}',
                "当前至少仍有一条未被 Semrush 标记为 Lost 的来源域名数量，是当前真实外链渠道广度的重要指标。",
                f'当前仍有 {st["live_domains"]:,} 个来源域名存活，占历史来源域名约 {(st["live_domains"] / st["total_domains"] if st["total_domains"] else 0):.1%}。'
            ),

            (
                "现存外链",
                f'{st["live_backlinks"]:,}',
                "当前未被标记为 Lost 的 backlink 数量，代表 Semrush 当前仍观察到的外链资产规模。",
                f'当前仍存在 {st["live_backlinks"]:,} 条外链，占历史外链约 {(st["live_backlinks"] / st["total_backlinks"] if st["total_backlinks"] else 0):.1%}。'
            ),

            (
                "外链流失率",
                pct(lost_bkl_pct),
                "Lost Backlinks ÷ 历史 Backlinks。衡量历史获得的单条链接有多少已经消失。",
                f'当前判断：{lost_bkl_level}。{lost_bkl_text}'
            ),

            (
                "引荐域名流失率",
                pct(lost_dom_pct),
                "已经完全没有任何现存 backlink 的来源域名，占历史 Ref Domains 的比例。它比单条 backlink 流失更能体现渠道是否真正消失。",
                f'当前判断：{lost_dom_level}。{lost_dom_text}'
            ),

            (
                "现存 Follow 占比",
                pct(st["follow_rate_live"]),
                "当前链接中未标记为 Nofollow / Sponsored / UGC 的 Follow 链接占比。它描述链接属性结构，不应被简单理解为 Follow 一定好、Nofollow 一定没价值。",
                f'当前判断：{follow_level}。{follow_text}'
            ),

            (
                "现存 Sitewide 占比",
                pct(st["sitewide_rate_live"]),
                "当前外链中被 Semrush 标记为 Sitewide 的比例。Sitewide 常见于 Footer、Sidebar、合作伙伴模板等同站大量页面重复出现的位置。",
                f'当前判断：{sitewide_level}。{sitewide_text}'
            ),

            (
                "高 Page AS 占比",
                pct(st["high_as_rate_live"]),
                "来自高 Page AS 来源页面的现存外链比例。Page AS 是“来源页面”权威分，并不是来源域名 DR 或 Domain AS。",
                f'当前判断：{high_as_level}。{high_as_text}'
            ),

            (
                "现存平均 Page AS",
                f'{st["avg_page_as_live"]:.1f}',
                "所有现存 backlink 来源页面 Page AS 的平均值，用于观察整体来源页权威水平。平均值容易受到大量低分长尾页面影响，因此必须结合 Page AS 分布图一起看。",
                f'当前平均 Page AS 为 {st["avg_page_as_live"]:.1f}。不要只看平均值，下面的 0–20 / 21–40 / 41–60 / 61–80 / 81–100 分布更重要。'
            ),

            (
                "现存外链 / 来源域名",
                f'{st["live_link_concentration"]:.2f}',
                "平均每个现存 Ref Domain 贡献多少条现存 backlink。这个指标衡量的是链接集中度，而不是链接质量。",
                f'当前判断：{conc_level}。{conc_text}'
            ),

            (
                "目标页面数",
                f'{st["unique_target_pages"]:,}',
                "至少获得过一条 backlink 的竞品 Target URL 数量。它可以观察外链是否只集中于首页，还是广泛支持产品页、类目页、博客等页面。",
                f'共有 {st["unique_target_pages"]:,} 个目标页面获得外链。数字越大通常说明外链资产在站内分布越广；但仍需结合各 Target Page 的链接数量判断是否高度集中。'
            ),

            (
                "Sponsored 占比",
                pct(st["sponsored_rate_live"]),
                "现存 backlink 中带有 Sponsored 属性的比例，通常与广告、赞助内容或商业合作有关。",
                f'当前 Sponsored 占比为 {st["sponsored_rate_live"]:.1%}。较高时说明商业合作型链接在整个 profile 中占有明显位置；较低则说明这种属性不是主要来源。'
            ),

            (
                "UGC 占比",
                pct(st["ugc_rate_live"]),
                "现存 backlink 中带 UGC 属性的比例，常见于论坛、评论、社区和其他用户生成内容。",
                f'当前 UGC 占比为 {st["ugc_rate_live"]:.1%}。较高通常意味着社区/论坛型来源更多；较低说明 UGC 并不是主要外链来源。'
            ),

            (
                "Semrush 新增链接标记",
                f'{st["new_flagged_backlinks"]:,}',
                "当前这次 Semrush 数据中 New Link=true 的 backlink 数量。它是 Semrush 的新增标记，不等同于“今年新建了多少链接”，也不应该直接替代 First Seen 趋势。",
                f'本次数据中共有 {st["new_flagged_backlinks"]:,} 条 New Link 标记。真正判断增长速度，应继续结合下面 First Seen 年份趋势。'
            ),
        ]


        for metric, value, meaning, diagnosis in explanations:

            r = ws3.max_row + 1

            ws3.merge_cells(
                start_row=r,
                start_column=1,
                end_row=r,
                end_column=2
            )

            ws3.merge_cells(
                start_row=r,
                start_column=4,
                end_row=r,
                end_column=7
            )

            ws3.merge_cells(
                start_row=r,
                start_column=8,
                end_row=r,
                end_column=12
            )

            ws3.cell(r, 1, metric)
            ws3.cell(r, 3, value)
            ws3.cell(r, 4, meaning)
            ws3.cell(r, 8, diagnosis)

            ws3.cell(r, 1).font = Font(bold=True)
            ws3.cell(r, 3).font = Font(bold=True)

            for c in [1, 3, 4, 8]:
                ws3.cell(r, c).alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )



        # --------------------------------------------------
        # ④ 综合诊断
        # --------------------------------------------------

        overall_row = ws3.max_row + 2

        ws3.merge_cells(
            start_row=overall_row,
            start_column=1,
            end_row=overall_row,
            end_column=12
        )

        overall_text = (
            f"🔎 综合判断：{comp} 当前拥有 {st['live_domains']:,} 个现存引荐域名、"
            f"{st['live_backlinks']:,} 条现存外链；"
            f"外链流失率 {lost_bkl_pct:.1%}，域名流失率 {lost_dom_pct:.1%}；"
            f"Follow 占比 {st['follow_rate_live']:.1%}，Sitewide 占比 {st['sitewide_rate_live']:.1%}；"
            f"平均每个现存来源域名贡献 {st['live_link_concentration']:.2f} 条链接。"
            f"因此判断外链质量时，不应单独看 Backlinks 总量，而应同时看："
            f"【Ref Domain 广度 → Lost Domain 稳定性 → Sitewide 集中度 → Page AS 来源结构 → Target Page 分布】。"
        )

        ws3.cell(
            overall_row,
            1,
            overall_text
        )

        ws3.cell(overall_row, 1).font = Font(
            bold=True,
            color="1F1F1F"
        )

        ws3.cell(overall_row, 1).fill = PatternFill(
            "solid",
            fgColor="FFF2CC"
        )

        ws3.cell(overall_row, 1).alignment = Alignment(
            wrap_text=True,
            vertical="center"
        )


        ws3.append([])

        

        # 默认折叠解释正文；标题保持显示
        explain_end_row = ws3.max_row
        ws3.sheet_view.showOutlineSymbols = True
        ws3.sheet_properties.outlinePr.summaryBelow = False

        ws3.row_dimensions.group(
            explain_start_row + 1,
            explain_end_row,
            outline_level=1,
            hidden=True
        )

        # 强制显示折叠标记
        ws3.row_dimensions[explain_start_row].collapsed = True
        ws3.row_dimensions[explain_start_row].hidden = False

        # 标题直接告诉使用者怎么展开
        ws3.cell(
            explain_start_row,
            1
        ).value = "🧭 核心指标解释与自动诊断｜点击左侧「+」展开详情"

        for r in range(explain_start_row + 1, explain_end_row + 1):
            ws3.row_dimensions[r].height = 24

        # Page AS distribution
        ws3.append(["来源页面 Page AS 分布（按 backlink 行）"] + [b[0] for b in AS_BANDS])
        as_header = ws3.max_row
        style_header(ws3, as_header, fill=COLOR_GRAY, font_color="000000")
        ws3.append(["历史 Backlinks"] + [st["hist_as_dist"].get(b[0], 0) for b in AS_BANDS])
        as_hist_row = ws3.max_row
        ws3.append(["Live Backlinks"] + [st["live_as_dist"].get(b[0], 0) for b in AS_BANDS])
        as_live_row = ws3.max_row
        ws3.append([])

        # Live link type profile (mutually exclusive)
        ws3.append(["Live Link Type", "Text", "Image", "Form", "Frame", "Unknown"])
        type_header = ws3.max_row
        style_header(ws3, type_header, fill=COLOR_GRAY, font_color="000000")
        ws3.append(["Backlink Count"] + [st["live_type_dist"].get(x, 0) for x in ["Text", "Image", "Form", "Frame", "Unknown"]])
        type_data = ws3.max_row
        ws3.append([])

        # Time cohorts
        if all_years:
            ws3.append(["First Seen 年份"] + all_years)
            time_header = ws3.max_row
            style_header(ws3, time_header, fill=COLOR_GRAY, font_color="000000")
            domain_new = [comp_time[comp]["domain_first"].get(y, 0) for y in all_years]
            domain_live = [comp_time[comp]["live_domain_first"].get(y, 0) for y in all_years]
            backlink_new = [comp_time[comp]["backlink_first"].get(y, 0) for y in all_years]
            backlink_live = [comp_time[comp]["live_backlink_first"].get(y, 0) for y in all_years]

            ws3.append(["当年首次获得 Ref Domains"] + domain_new)
            domain_new_row = ws3.max_row
            ws3.append(["累计历史 Ref Domains"] + cumulative(domain_new))
            domain_cum_row = ws3.max_row
            ws3.append(["当前仍存活的 Ref Domain Cohort"] + domain_live)
            live_domain_row = ws3.max_row
            ws3.append(["当年首次发现 Backlinks"] + backlink_new)
            backlink_new_row = ws3.max_row
            ws3.append(["当前仍 Live 的 Backlink Cohort"] + backlink_live)
            live_backlink_row = ws3.max_row
            ws3.append([])
        else:
            time_header = domain_new_row = domain_cum_row = live_domain_row = backlink_new_row = live_backlink_row = None

        # Charts: use helper data immediately above, and leave visual space.
        chart_row = ws3.max_row + 1

        as_chart = BarChart()
        as_chart.type = "col"
        as_chart.style = 2
        as_chart.title = f"[{comp}] 来源页 Page AS 结构｜Source Page AS Distribution"
        as_chart.width = 16
        as_chart.height = 8
        as_chart.add_data(Reference(ws3, min_col=1, min_row=as_hist_row, max_col=6, max_row=as_live_row), from_rows=True, titles_from_data=True)
        as_chart.set_categories(Reference(ws3, min_col=2, min_row=as_header, max_col=6, max_row=as_header))
        ws3.add_chart(as_chart, f"A{chart_row}")

        type_chart = BarChart()
        type_chart.type = "col"
        type_chart.style = 2
        type_chart.title = f"[{comp}] 现存链接类型｜Live Link Type"
        type_chart.width = 16
        type_chart.height = 8
        type_chart.add_data(Reference(ws3, min_col=2, min_row=type_header, max_col=6, max_row=type_data), titles_from_data=True)
        type_chart.set_categories(Reference(ws3, min_col=2, min_row=type_header, max_col=6, max_row=type_header))
        ws3.add_chart(type_chart, f"G{chart_row}")

        if all_years and time_header:
            line = LineChart()
            line.style = 2
            line.title = f"[{comp}] 引荐域名获取趋势｜Ref Domain Acquisition Cohorts"
            line.width = 16
            line.height = 8
            line.add_data(Reference(ws3, min_col=1, min_row=domain_new_row, max_col=len(all_years)+1, max_row=live_domain_row), from_rows=True, titles_from_data=True)
            line.set_categories(Reference(ws3, min_col=2, min_row=time_header, max_col=len(all_years)+1, max_row=time_header))
            ws3.add_chart(line, f"A{chart_row + 16}")

            line_b = LineChart()
            line_b.style = 2
            line_b.title = f"[{comp}] 外链首次发现趋势｜Backlink First-Seen Cohorts"
            line_b.width = 16
            line_b.height = 8
            line_b.add_data(Reference(ws3, min_col=1, min_row=backlink_new_row, max_col=len(all_years)+1, max_row=live_backlink_row), from_rows=True, titles_from_data=True)
            line_b.set_categories(Reference(ws3, min_col=2, min_row=time_header, max_col=len(all_years)+1, max_row=time_header))
            ws3.add_chart(line_b, f"G{chart_row + 16}")

        for _ in range(34):
            ws3.append([])

    # --------------------------------------------------
    # Sheet 4: Global competitor dashboard
    # --------------------------------------------------
    ws4 = wb.create_sheet("全局竞品大盘")
    ws4.sheet_view.showGridLines = False
    add_title(
        ws4,
        "🌐 Semrush 外链竞品全局分析｜Backlink Competitive Dashboard",
        17,
        fill=COLOR_DARK
    )
    for _ in range(62):
        ws4.append([])

    data_start = ws4.max_row + 1
    ws4.append([
    "竞品（Competitor）",
    "历史引荐域名（Ref Domains）",
    "历史外链（Backlinks）",
    "现存引荐域名（Live Ref Domains）",
    "现存外链（Live Backlinks）",

    "现存 Follow 占比（Live Follow %）",
    "现存 Sitewide 占比（Live Sitewide %）",
    "现存 Sponsored 占比（Live Sponsored %）",
    "现存 UGC 占比（Live UGC %）",
    f"现存高 Page AS 占比（Page AS ≥ {HIGH_PAGE_AS_THRESHOLD}）",

    "现存平均 Page AS（Live Avg Page AS）",
    "历史外链集中度（Backlinks / Ref Domain）",

    "外链流失率（Lost Backlink %）",
    "引荐域名流失率（Lost Domain %）",
    "现存外链集中度（Live Backlinks / Live Ref Domain）",
    "获链目标页面数（Target Pages）",
    "Semrush 新增链接标记（New Link Flags）",
    ])
    main_header = ws4.max_row
    style_header(ws4, main_header, fill=COLOR_BLUE)
    for comp in comp_list:
        st = comp_stats[comp]

        lost_bkl_pct = (
            st["lost_backlinks"] / st["total_backlinks"]
            if st["total_backlinks"] else 0
        )

        lost_dom_pct = (
            st["lost_domains"] / st["total_domains"]
            if st["total_domains"] else 0
        )

        ws4.append([
            comp,
            st["total_domains"],
            st["total_backlinks"],
            st["live_domains"],
            st["live_backlinks"],

            st["follow_rate_live"],
            st["sitewide_rate_live"],
            st["sponsored_rate_live"],
            st["ugc_rate_live"],
            st["high_as_rate_live"],

            round(st["avg_page_as_live"], 1),
            round(st["link_concentration"], 2),

            lost_bkl_pct,
            lost_dom_pct,
            round(st["live_link_concentration"], 2),
            st["unique_target_pages"],
            st["new_flagged_backlinks"],
        ])

        percent_fmt(
            ws4,
            ws4.max_row,
            [6, 7, 8, 9, 10, 13, 14]
        )
    main_end = ws4.max_row

    ws4.append([])
    ws4.append(["Live Page AS Structure"] + [b[0] for b in AS_BANDS])
    as_global_header = ws4.max_row
    style_header(ws4, as_global_header, fill=COLOR_GRAY, font_color="000000")
    for comp in comp_list:
        st = comp_stats[comp]
        total_live = st["live_backlinks"] or 1
        ws4.append([comp] + [st["live_as_dist"].get(b[0], 0) / total_live for b in AS_BANDS])
        percent_fmt(ws4, ws4.max_row, [2, 3, 4, 5, 6])
    as_global_end = ws4.max_row

    ws4.append([])
    if all_years:
        ws4.append(["Ref Domain First-Seen Trend"] + all_years)
        trend_header = ws4.max_row
        style_header(ws4, trend_header, fill=COLOR_GRAY, font_color="000000")
        for comp in comp_list:
            vals = [comp_time[comp]["domain_first"].get(y, 0) for y in all_years]
            ws4.append([comp] + cumulative(vals))
        trend_end = ws4.max_row
    else:
        trend_header = trend_end = None

    set_widths(ws4, {
        "A": 28,
        "B": 18,
        "C": 16,
        "D": 20,
        "E": 18,
        "F": 18,
        "G": 20,
        "H": 20,
        "I": 17,
        "J": 22,
        "K": 20,
        "L": 22,
        "M": 18,
        "N": 20,
        "O": 24,
        "P": 20,
        "Q": 22,
    })

    # ==================================================
    # 图表区：两排布局，中间留出明显间距
    # ==================================================

    # ==================================================
    # 全局大盘图表 —— 只保留这一套，不要再追加旧图
    # ==================================================

    c1 = BarChart()
    c1.type = "col"
    c1.style = 2
    c1.title = "外链规模对比｜Live Ref Domains vs Live Backlinks"
    c1.y_axis.title = "数量｜Count"
    c1.width = 22
    c1.height = 10
    c1.add_data(
        Reference(ws4, min_col=4, min_row=main_header,
                max_col=5, max_row=main_end),
        titles_from_data=True
    )
    c1.set_categories(
        Reference(ws4, min_col=1, min_row=main_header + 1,
                max_col=1, max_row=main_end)
    )
    ws4.add_chart(c1, "A3")


    c2 = BarChart()
    c2.type = "col"
    c2.style = 2
    c2.grouping = "clustered"
    c2.title = "现存链接属性结构｜Live Link Attribute Profile"
    c2.y_axis.title = "占比｜Share"
    c2.y_axis.number_format = "0%"
    c2.width = 22
    c2.height = 10
    c2.add_data(
        Reference(ws4, min_col=6, min_row=main_header,
                max_col=10, max_row=main_end),
        titles_from_data=True
    )
    c2.set_categories(
        Reference(ws4, min_col=1, min_row=main_header + 1,
                max_col=1, max_row=main_end)
    )
    ws4.add_chart(c2, "M3")


    c3 = BarChart()
    c3.type = "col"
    c3.style = 2
    c3.grouping = "stacked"
    c3.overlap = 100
    c3.title = "现存来源页 Page AS 结构｜Live Source Page AS Structure"
    c3.y_axis.title = "占比｜Share"
    c3.y_axis.number_format = "0%"
    c3.width = 22
    c3.height = 10
    c3.add_data(
        Reference(ws4, min_col=2, min_row=as_global_header,
                max_col=6, max_row=as_global_end),
        titles_from_data=True
    )
    c3.set_categories(
        Reference(ws4, min_col=1, min_row=as_global_header + 1,
                max_col=1, max_row=as_global_end)
    )
    ws4.add_chart(c3, "A25")


    if trend_header:
        c4 = LineChart()
        c4.style = 2
        c4.title = "累计引荐域名增长趋势｜Cumulative Ref Domain Acquisition"
        c4.y_axis.title = "累计引荐域名｜Ref Domains"
        c4.x_axis.title = "首次发现年份｜First Seen"
        c4.width = 22
        c4.height = 10

        c4.add_data(
            Reference(
                ws4,
                min_col=1,
                min_row=trend_header + 1,
                max_col=len(all_years) + 1,
                max_row=trend_end
            ),
            from_rows=True,
            titles_from_data=True
        )

        c4.set_categories(
            Reference(
                ws4,
                min_col=2,
                min_row=trend_header,
                max_col=len(all_years) + 1,
                max_row=trend_header
            )
        )

        ws4.add_chart(c4, "M25")





    # --------------------------------------------------
    # Sheet 5: Import audit + data dictionary
    # --------------------------------------------------
    ws5 = wb.create_sheet("数据口径与导入日志")
    ws5.sheet_view.showGridLines = False
    add_title(ws5, "数据口径 / Import Audit", 7, fill=COLOR_DARK)
    ws5.append([])
    ws5.append(["文件", "识别竞品", "原始行数", "成功写入", "说明"])
    style_header(ws5, ws5.max_row, fill=COLOR_BLUE)
    for item in import_log:
        ws5.append(list(item))

    ws5.append([])
    ws5.append(["字段/指标", "定义"])
    style_header(ws5, ws5.max_row, fill=COLOR_GRAY, font_color="000000")
    definitions = [
        ("Page AS", "Semrush 来源页面级 Authority Score；不是 Ahrefs DR，也不是 referring-domain Authority Score。"),
        ("Backlink 行数", "按 Semrush 导出的一行 backlink 计数；导入时仅去除完全相同的 backlink identity。"),
        ("Ref Domain", "从 Source URL 提取的可注册域名；若环境未安装 tldextract，则退化为 hostname。"),
        ("Live", "Lost link = false。"),
        ("Follow", "Nofollow / Sponsored / UGC 均为 false 的链接。"),
        ("Lost Domain", "该 Ref Domain 在当前导出中没有任何 Live backlink。"),
        ("Alpha", "同一来源域名被多少个不同竞品获取。"),
        ("Sitewide %", "Semrush Sitewide=true 的 backlink 行占比；仅作为结构信号，不直接标记为 Spam。"),
        ("High Page AS %", f"来源页面 Page AS >= {HIGH_PAGE_AS_THRESHOLD} 的 Live backlink 行占比。阈值仅用于内部分层。"),
        ("First-Seen Cohort", "按 Semrush 首次发现日期归因；Ref Domain 趋势按该域名在某竞品数据中的最早 First Seen 计算，避免同域多链接重复计数。"),
        ("Sponsored/UGC/Sitewide rates", "属性可以相互重叠，因此全局图使用 clustered bars，不把它们强行堆叠到 100%。"),
    ]
    for k, v in definitions:
        ws5.append([k, v])
        ws5.cell(ws5.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    set_widths(ws5, {"A": 28, "B": 100, "C": 15, "D": 15, "E": 80})

    # Generic workbook polish
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 90
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.alignment == Alignment():
                    cell.alignment = Alignment(vertical="top")

    wb.save(excel_filename)
    print(f"\n[✓] Semrush 外链情报报告已生成：{excel_filename}")
    print(f"[✓] Opportunity JSON 已生成：{json_filename}")
    print(f"[✓] 有效 backlink 事实行：{len(rows):,}")
    print(f"[✓] 竞品数量：{len(comp_list)}；跨竞品来源域名：{len(domain_intel):,}")
    return excel_filename, json_filename


def run_matrix_engine():
    print("=== Semrush Backlink Intelligence Engine ===")
    print(f"数据目录: {DATA_DIR}")
    conn = init_db()
    try:
        total_files, global_headers, import_log = process_semrush_exports(conn)
        if total_files > 0:
            generate_reports(conn, global_headers, import_log)
        else:
            print("[结束] 没有可处理的 Semrush Backlinks 文件。")
    finally:
        conn.close()


if __name__ == "__main__":
    run_matrix_engine()
