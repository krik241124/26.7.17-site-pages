# backlink_matrix.py
import os
import sqlite3
import csv
import json
import datetime
import glob
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "backlink_intelligence.db")
DATA_DIR = os.path.join(WORKSPACE_DIR, "ahrefs_data")
REPORT_DIR = os.path.join(WORKSPACE_DIR, "reports")

def get_current_time_str():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS backlinks")
    # 增加 raw_data 字段，用于以 JSON 格式存储所有 Ahrefs 原始 30+ 个维度
    cur.execute('''
        CREATE TABLE backlinks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competitor_domain TEXT,
            ref_domain TEXT,
            ref_url TEXT,
            target_url TEXT,
            domain_rating INTEGER,
            page_traffic INTEGER,
            page_type TEXT,
            page_category TEXT,
            is_dofollow INTEGER,
            is_spam INTEGER,
            raw_data TEXT,
            UNIQUE(competitor_domain, ref_url)
        )
    ''')
    cur.execute("CREATE INDEX idx_ref_domain ON backlinks(ref_domain)")
    conn.commit()
    return conn

def extract_domain(url):
    try:
        netloc = urlparse(url).netloc
        if netloc.startswith('www.'):
            return netloc[4:]
        return netloc
    except Exception:
        return ""

def process_ahrefs_exports(conn):
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
        print(f"[提示] 请将 Ahrefs 数据放入 {DATA_DIR} 目录后重新运行。")
        return 0, []

    files = glob.glob(os.path.join(DATA_DIR, "*.*sv"))
    if not files:
        print(f"[提示] 在 {DATA_DIR} 中未找到 csv/tsv 文件。")
        return 0, []

    cur = conn.cursor()
    total_files = len(files)
    global_headers = [] # 用于收集所有 Ahrefs 原始表头
    
    for file_path in files:
        filename = os.path.basename(file_path).lower()
        competitor = filename.split('-backlinks')[0] if '-backlinks' in filename else filename.split('_')[0].replace('.csv', '')
        print(f"正在清洗并分析竞品 [{competitor}] 的外链数据...")
        
        try:
            f = open(file_path, 'r', encoding='utf-16')
            f.read(1); f.seek(0)
        except UnicodeError:
            f = open(file_path, 'r', encoding='utf-8')

        reader = csv.DictReader(f, delimiter='\t' if file_path.endswith('.tsv') or 'utf-16' in str(f.encoding) else ',')
        
        # 提取并清理表头（去除BOM乱码），放入全局表头列表
        cleaned_fieldnames = [str(fn).replace('\ufeff', '').strip() for fn in reader.fieldnames if fn]
        for fn in cleaned_fieldnames:
            if fn not in global_headers:
                global_headers.append(fn)
        
        insert_data = []
        for raw_row in reader:
            # 清理当前行的 Key
            clean_row = {str(k).replace('\ufeff', '').strip(): v for k, v in raw_row.items() if k}
            # 方便内部逻辑提取的小写字典
            lower_row = {k.lower(): v for k, v in clean_row.items()}
            
            ref_url = lower_row.get('referring page url') or lower_row.get('source url')
            if not ref_url: continue
                
            ref_domain = extract_domain(ref_url)
            target_url = lower_row.get('target url', '')
            
            try: dr = int(float(lower_row.get('domain rating') or lower_row.get('dr') or 0))
            except (ValueError, TypeError): dr = 0

            try: traffic = int(float(lower_row.get('page traffic') or lower_row.get('traffic') or 0))
            except (ValueError, TypeError): traffic = 0
            
            page_type = lower_row.get('page type', 'Unknown')
            page_category = lower_row.get('page category', 'Unknown')
            
            is_dofollow = 1 if str(lower_row.get('nofollow', 'FALSE')).upper() != 'TRUE' else 0
            is_spam = 1 if str(lower_row.get('is spam', 'FALSE')).upper() == 'TRUE' else 0
            
            if is_spam: continue

            # 极其重要：将包含所有维度的 clean_row 原封不动地存为 JSON
            raw_json = json.dumps(clean_row, ensure_ascii=False)

            insert_data.append((
                competitor, ref_domain, ref_url, target_url, 
                dr, traffic, page_type, page_category, is_dofollow, is_spam, raw_json
            ))
            
        f.close()
        
        cur.executemany('''
            INSERT OR IGNORE INTO backlinks 
            (competitor_domain, ref_domain, ref_url, target_url, domain_rating, page_traffic, page_type, page_category, is_dofollow, is_spam, raw_data) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', insert_data)
        
    conn.commit()
    return total_files, global_headers

def generate_strategic_reports(conn, total_files, global_headers):
    if not os.path.exists(REPORT_DIR):
        os.makedirs(REPORT_DIR)

    cur = conn.cursor()
    overlap_threshold = 2 if total_files >= 2 else 1

    # 我们将全量拉取数据在 Python 端进行多维聚类，以实现“宏观+微观”双重输出
    cur.execute('''
        SELECT competitor_domain, ref_domain, ref_url, target_url, domain_rating, page_traffic, page_type, page_category, raw_data 
        FROM backlinks 
        WHERE is_dofollow = 1
    ''')
    all_links = cur.fetchall()

    # 聚类逻辑：按来源域名进行聚合
    domain_map = {}
    for row in all_links:
        comp, ref_d, ref_u, tgt_u, dr, traf, p_type, p_cat, raw_j = row
        if ref_d not in domain_map:
            domain_map[ref_d] = {
                'competitors': set(),
                'max_dr': 0,
                'max_traffic': 0,
                'types': set(),
                'categories': set(),
                'link_details': [],
                'raw_rows': []
            }
        d = domain_map[ref_d]
        d['competitors'].add(comp)
        d['max_dr'] = max(d['max_dr'], dr)
        d['max_traffic'] = max(d['max_traffic'], traf)
        if p_type and p_type != 'Unknown': d['types'].add(p_type)
        if p_cat and p_cat != 'Unknown': d['categories'].add(p_cat)
        # 保存具体的竞品外链映射 (谁 发了哪篇文章 指向了哪个页面)
        d['link_details'].append(f"[{comp}] {ref_u}  --->  {tgt_u}")
        # 保存这行的原始维度 JSON
        d['raw_rows'].append((comp, ref_u, tgt_u, raw_j))

    # 过滤、计算 Alpha 并排序
    top_domains = []
    for ref_d, d_info in domain_map.items():
        alpha_score = len(d_info['competitors'])
        if alpha_score >= overlap_threshold:
            d_info['alpha'] = alpha_score
            d_info['ref_domain'] = ref_d
            top_domains.append(d_info)

    # 排序优先级：重合度(Alpha) > 网站权重(DR) > 流量
    top_domains.sort(key=lambda x: (x['alpha'], x['max_dr'], x['max_traffic']), reverse=True)

    matrix_rows = []    # 给 Sheet 1 (战略大盘)
    detail_rows = []    # 给 Sheet 2 (外链原始维度)
    json_payloads = []  # 给 AI Agent

    for d in top_domains:
        # 清洗向量标签
        p_types = list(d['types'])
        p_cats = list(d['categories'])
        clean_types = list(set([t.split('>')[-1].strip() for t in p_types if t]))
        clean_cats = list(set([c.split('>')[0].strip() for c in p_cats if c]))
        vector_format = clean_types[0] if clean_types else "Blog/Article"
        vector_industry = clean_cats[0] if clean_cats else "General B2B"
        shared_by = ", ".join(d['competitors'])

        # 组装溯源链接（带换行符，方便直接查看）
        clickable_references = "\n".join(list(set(d['link_details'])))

        # Sheet 1 数据
        matrix_rows.append([
            d['ref_domain'], d['alpha'], d['max_dr'], d['max_traffic'], 
            vector_industry, vector_format, shared_by, clickable_references
        ])

        # Sheet 2 数据：解包原始 JSON
        for link_row in d['raw_rows']:
            comp, ref_u, tgt_u, raw_j = link_row
            raw_dict = json.loads(raw_j)
            
            # 第一列固定为 Alpha，第二列是竞品名，接着是所有原装的 Ahrefs 列
            row_detail = [d['alpha'], comp]
            for header in global_headers:
                row_detail.append(raw_dict.get(header, ""))
            detail_rows.append(row_detail)

        # AI Agent Payload 依然保留
        prompt_template = f"你是一个资深的跨境电商 B2B 营销专家。现在我要向一个【渠道调性：{vector_industry}，文章类型偏好：{vector_format}】的权威网站投稿。该网站权重 DR 为 {d['max_dr']}。我的核心关键词是【家具一件代发供应商】。请严格根据以上特征，生成一篇字数2000字，带有对比表格，适合该渠道发布的专业文章初稿。"
        json_payloads.append({
            "target_domain": d['ref_domain'],
            "overlap_alpha_score": d['alpha'],
            "domain_rating": d['max_dr'],
            "agent_prompt": prompt_template
        })

    # ==== 开始生成双表 Excel ====
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    excel_filename = os.path.join(REPORT_DIR, f"ArkSwift_Backlink_Matrix_{date_str}.xlsx")
    
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        
        # --- Sheet 1: 战略大盘 ---
        ws1 = wb.active
        ws1.title = "战略大盘 (Alpha矩阵)"
        matrix_headers = ["来源域名 (Ref Domain)", "重合度 (Alpha Score)", "网站权重 (DR)", "最高页面流量", "核心行业标签", "核心内容类型", "被哪些竞品获取", "对标的具体外链地址 (方便直接抄作业)"]
        ws1.append(matrix_headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # 大红色更醒目
        for col_num, cell in enumerate(ws1[1], 1):
            cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['G'].width = 25
        ws1.column_dimensions['H'].width = 100 # 给连接列很宽的空间
        
        for row in matrix_rows:
            ws1.append(row)
            # 让 H 列（对标外链地址）自动换行，视觉效果拉满
            ws1.cell(row=ws1.max_row, column=8).alignment = Alignment(wrap_text=True, vertical='center')

        # --- Sheet 2: 全维度原始明细 ---
        ws2 = wb.create_sheet("外链全维度明细 (支持筛选查阅)")
        detail_headers = ["重合度 (Alpha Score)", "具体竞品 (Competitor)"] + global_headers
        ws2.append(detail_headers)
        
        header_fill_2 = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for col_num, cell in enumerate(ws2[1], 1):
            cell.font = header_font; cell.fill = header_fill_2
            
        for row in detail_rows:
            ws2.append(row)

        wb.save(excel_filename)
        print(f"\n[✓] 成功生成【双表架构】Excel 报表：{excel_filename}")
        print("    -> Sheet 1: Alpha 战略大盘 + 直达点击链接")
        print("    -> Sheet 2: 完整保留 Ahrefs 所有几十个原始维度，且已按 Alpha 降序！")
    else:
        # Fallback 到 CSV 输出两个文件
        m_file = excel_filename.replace('.xlsx', '_Matrix.csv')
        d_file = excel_filename.replace('.xlsx', '_RawDetails.csv')
        # (此处省略 CSV 降级输出代码，因为作为业务开发，推荐必须 pip install openpyxl 体验最完美的排版)
        pass

    # JSON Payload 接口
    json_filename = os.path.join(REPORT_DIR, f"Agent_Payload_{date_str}.json")
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump({"data": json_payloads}, f, ensure_ascii=False, indent=4)

def run_matrix_engine():
    print("=== 初始化 ArkSwift 向量外链矩阵引擎 ===")
    conn = init_db()
    total_files, global_headers = process_ahrefs_exports(conn)
    if total_files > 0:
        generate_strategic_reports(conn, total_files, global_headers)
    conn.close()

if __name__ == "__main__":
    run_matrix_engine()