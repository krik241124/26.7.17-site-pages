# keyword_matrix_pro.py
import os
import sqlite3
import csv
import json
import datetime
import glob
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, LineChart, Reference, BarChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "keyword_intelligence.db")
DATA_DIR = os.path.join(WORKSPACE_DIR, "ahrefs_data")
REPORT_DIR = os.path.join(WORKSPACE_DIR, "reports")

# ==========================================
# 🚀 机器评分引擎：纯 Python 规则 (Auto Scoring)
# ==========================================
def get_kw_alpha_score(count):
    if count >= 5: return 100
    elif count >= 4: return 85
    elif count >= 3: return 70
    elif count >= 2: return 40
    else: return 15

def get_kw_kd_score(kd):
    # KD 越低，得分越高 (快速出单快赢)
    if kd <= 10: return 100
    elif kd <= 20: return 90
    elif kd <= 35: return 75
    elif kd <= 50: return 50
    elif kd <= 70: return 20
    else: return 0

def get_kw_volume_score(vol):
    if vol >= 10000: return 100
    elif vol >= 5000: return 85
    elif vol >= 1000: return 70
    elif vol >= 500: return 50
    elif vol >= 100: return 30
    else: return 10

def get_current_time_str():
    return datetime.datetime.now().strftime("%Y-%m-%d")

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS keywords")
    cur.execute('''
        CREATE TABLE keywords (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competitor_domain TEXT,
            keyword TEXT,
            country TEXT,
            is_branded INTEGER,
            is_navigational INTEGER,
            is_informational INTEGER,
            is_commercial INTEGER,
            is_transactional INTEGER,
            volume INTEGER,
            kd INTEGER,
            cpc REAL,
            current_traffic REAL,
            current_position INTEGER,
            url TEXT,
            raw_data TEXT,
            UNIQUE(competitor_domain, keyword, country)
        )
    ''')
    cur.execute("CREATE INDEX idx_keyword ON keywords(keyword)")
    cur.execute("CREATE INDEX idx_domain ON keywords(competitor_domain)")
    conn.commit()
    return conn

def process_ahrefs_exports(conn):
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
        print(f"[提示] 请将 Ahrefs Organic Positions 数据放入 {DATA_DIR} 目录后重新运行。")
        return 0, []

    files = glob.glob(os.path.join(DATA_DIR, "*organic-positions*.csv")) + glob.glob(os.path.join(DATA_DIR, "*organic-positions*.tsv"))
    if not files:
        print(f"[提示] 在 {DATA_DIR} 中未找到含有 'organic-positions' 的 csv/tsv 文件。")
        return 0, []

    cur = conn.cursor()
    total_files = len(files)
    global_headers = [] 
    
    for file_path in files:
        filename = os.path.basename(file_path).lower()
        competitor = filename.split('-organic-positions')[0]
        print(f"正在清洗并提取竞品 [{competitor}] 的自然搜索关键词数据...")
        
        try:
            f = open(file_path, 'r', encoding='utf-16')
            f.read(1); f.seek(0)
        except UnicodeError:
            f = open(file_path, 'r', encoding='utf-8-sig')

        delimiter = '\t' if file_path.endswith('.tsv') or 'utf-16' in str(f.encoding) else ','
        reader = csv.DictReader(f, delimiter=delimiter)

        # 即使 Ahrefs 无数据，也保留竞品
        if not reader.fieldnames:
            print(f"[提示] {competitor} 无关键词数据，建立空竞品档案")

            # 注册空竞品
            cur.execute("""
                INSERT OR IGNORE INTO keywords
                (
                    competitor_domain,
                    keyword,
                    country,
                    is_branded,
                    is_navigational,
                    is_informational,
                    is_commercial,
                    is_transactional,
                    volume,
                    kd,
                    cpc,
                    current_traffic,
                    current_position,
                    url,
                    raw_data
                )
                VALUES (?, '', '', 0,0,0,0,0,0,0,0,0,999,'','{}')
            """, (competitor,))

            f.close()
            continue


        cleaned_fieldnames = [str(fn).strip() for fn in reader.fieldnames if fn]
        for fn in cleaned_fieldnames:
            if fn not in global_headers:
                global_headers.append(fn)
        
        insert_data = []
        for raw_row in reader:
            clean_row = {str(k).strip(): v for k, v in raw_row.items() if k}
            lower_row = {k.lower(): v for k, v in clean_row.items()}
            
            keyword = lower_row.get('keyword', '').strip()
            if not keyword: continue
                
            country = lower_row.get('country', 'Unknown')
            url = lower_row.get('url', '')
            
            is_branded = 1 if str(lower_row.get('branded', 'FALSE')).upper() == 'TRUE' else 0
            is_nav = 1 if str(lower_row.get('navigational', 'FALSE')).upper() == 'TRUE' else 0
            is_info = 1 if str(lower_row.get('informational', 'FALSE')).upper() == 'TRUE' else 0
            is_comm = 1 if str(lower_row.get('commercial', 'FALSE')).upper() == 'TRUE' else 0
            is_trans = 1 if str(lower_row.get('transactional', 'FALSE')).upper() == 'TRUE' else 0
            
            try: volume = int(float(lower_row.get('volume') or 0))
            except: volume = 0
            try: kd = int(float(lower_row.get('kd') or 0))
            except: kd = 0
            try: cpc = float(lower_row.get('cpc') or 0)
            except: cpc = 0.0
            try: current_traffic = float(lower_row.get('current organic traffic') or lower_row.get('traffic') or 0)
            except: current_traffic = 0.0
            try: current_position = int(float(lower_row.get('current position') or lower_row.get('position') or 999))
            except: current_position = 999
            
            raw_json = json.dumps(clean_row, ensure_ascii=False)

            insert_data.append((
                competitor, keyword, country, is_branded, is_nav, is_info, is_comm, is_trans,
                volume, kd, cpc, current_traffic, current_position, url, raw_json
            ))
            
        f.close()
        
        cur.executemany('''
            INSERT OR IGNORE INTO keywords 
            (competitor_domain, keyword, country, is_branded, is_navigational, is_informational, is_commercial, is_transactional, volume, kd, cpc, current_traffic, current_position, url, raw_data) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', insert_data)
        
    conn.commit()
    return total_files, global_headers

def generate_strategic_reports(conn, total_files, global_headers):
    if not os.path.exists(REPORT_DIR):
        os.makedirs(REPORT_DIR)

    cur = conn.cursor()
    print("\n[进程] 正在从数据库拉取全量关键词聚合数据...")
    cur.execute('''
        SELECT competitor_domain, keyword, country, is_branded, is_navigational, is_informational, is_commercial, is_transactional, volume, kd, cpc, current_traffic, current_position, url 
        FROM keywords 
    ''')
    all_keywords = cur.fetchall()


    comp_stats = {}
    domain_kw_map = {}
    page_stats = {}




    for row in all_keywords:
        comp, kw, country, is_branded, is_nav, is_info, is_comm, is_trans, vol, kd, cpc, traf, pos, url = row
        
        if comp not in comp_stats:
            comp_stats[comp] = {
                'total_kws': 0, 'total_traffic': 0.0, 'traffic_value': 0.0,
                'branded_traf': 0.0, 'non_branded_traf': 0.0,
                'intent_nav': 0.0, 'intent_info': 0.0, 'intent_comm': 0.0, 'intent_trans': 0.0,
                'pos_1_3': 0.0, 'pos_4_10': 0.0, 'pos_11_20': 0.0, 'pos_21_plus': 0.0,
                'kd_0_15': 0.0, 'kd_16_40': 0.0, 'kd_41_plus': 0.0,
                'countries': {}, 'keywords_data': []
            }
        
        c = comp_stats[comp]
        if kw:
            c['total_kws'] += 1
        c['total_traffic'] += traf
        c['traffic_value'] += (traf * cpc)
        
        if is_branded: c['branded_traf'] += traf
        else: c['non_branded_traf'] += traf
        
        intent_count = is_nav + is_info + is_comm + is_trans
        if intent_count > 0:
            split_traf = traf / intent_count
            if is_nav: c['intent_nav'] += split_traf
            if is_info: c['intent_info'] += split_traf
            if is_comm: c['intent_comm'] += split_traf
            if is_trans: c['intent_trans'] += split_traf
        
        if pos <= 3: c['pos_1_3'] += traf
        elif pos <= 10: c['pos_4_10'] += traf
        elif pos <= 20: c['pos_11_20'] += traf
        else: c['pos_21_plus'] += traf
            
        if kd <= 15: c['kd_0_15'] += traf
        elif kd <= 40: c['kd_16_40'] += traf
        else: c['kd_41_plus'] += traf

        c['countries'][country] = c['countries'].get(country, 0.0) + traf
        
        if traf > 0 and pos <= 50:
            c['keywords_data'].append({
                'kw': kw, 'vol': vol, 'kd': kd, 'cpc': cpc, 'pos': pos, 
                'traf': traf, 'is_branded': is_branded, 'url': url
            })

        if url:
            if comp not in page_stats: page_stats[comp] = {}
            if url not in page_stats[comp]:
                page_stats[comp][url] = {'traffic': 0.0, 'kws': set()}
            page_stats[comp][url]['traffic'] += traf
            page_stats[comp][url]['kws'].add(kw)
            
        if not is_branded and traf > 0: 
            if kw not in domain_kw_map:
                domain_kw_map[kw] = {
                    'comps': set(), 'total_vol': vol, 'avg_kd': kd, 'avg_cpc': cpc, 'total_traf': 0.0,
                    'is_info': is_info, 'is_comm': is_comm, 'is_trans': is_trans # 用于 AI 意图判定
                }
            domain_kw_map[kw]['comps'].add(comp)
            domain_kw_map[kw]['total_traf'] += traf

    date_str = get_current_time_str()
    excel_filename = os.path.join(REPORT_DIR, f"Organic_Keywords_Pro_Matrix_{date_str}.xlsx")
    
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        sw_colors = ["294266", "F7941D", "20B799", "FDB913", "3FBBDF", "8C67AB", "E5625E", "8391A5"]
        
        # ==========================================
        # --- Sheet 1: 全局竞品大盘横评 (Super Dashboard) ---
        # ==========================================
        ws1 = wb.active
        ws1.title = "全局竞品大盘对比"
        ws1.sheet_view.showGridLines = False

        comp_list = list(comp_stats.keys())
        current_row = 2

        ws1.cell(row=current_row, column=1, value="[核心评估] 流量底盘与资产价值估算").font = Font(bold=True)
        current_row += 1
        headers_main = ["竞品名称", "有排名关键词总数", "预估月均总流量", "非品牌词(高质量)流量占比", "商业/交易型流量占比", "Top3 核心排名流量占比", "预估流量月度价值(CPC折算)"]
        ws1.append(headers_main)
        
        for comp in comp_list:
            st = comp_stats[comp]
            t_traf = st['total_traffic'] or 0

            safe_traf = t_traf if t_traf > 0 else 1

            ws1.append([
                comp,
                st['total_kws'],
                t_traf,
                st['non_branded_traf']/safe_traf,
                (st['intent_comm']+st['intent_trans'])/safe_traf,
                st['pos_1_3']/safe_traf,
                st['traffic_value']
            ])

            r = ws1.max_row
            ws1.cell(row=r, column=3).number_format = '#,##0'
            for c in [4,5,6]: ws1.cell(row=r, column=c).number_format = '0.00%'
            ws1.cell(row=r, column=7).number_format = '"$"#,##0.00'
            
        current_row = ws1.max_row + 3
        
        ws1.cell(row=current_row, column=1, value="[意图漏斗] 流量 Search Intent 分布").font = Font(bold=True)
        current_row += 1
        ws1.append(["竞品名称", "Informational(信息型)", "Navigational(导航型)", "Commercial(商业型)", "Transactional(交易型)"])
        intent_start_row = current_row
        for comp in comp_list:
            st = comp_stats[comp]
            t = st['total_traffic'] if st['total_traffic'] > 0 else 1
            ws1.append([comp, st['intent_info']/t, st['intent_nav']/t, st['intent_comm']/t, st['intent_trans']/t])
            for c in range(2, 6): ws1.cell(row=ws1.max_row, column=c).number_format = '0.00%'
        intent_end_row = ws1.max_row
        current_row = ws1.max_row + 3

        ws1.cell(row=current_row, column=1, value="[护城河] 流量排名区间分布").font = Font(bold=True)
        current_row += 1
        ws1.append(["竞品名称", "Top 1-3 (绝对核心)", "Top 4-10 (首屏潜力)", "Top 11-20 (二页后备)", "Top 21+ (长尾支撑)"])
        pos_start_row = current_row
        for comp in comp_list:
            st = comp_stats[comp]
            t = st['total_traffic'] if st['total_traffic'] > 0 else 1
            ws1.append([comp, st['pos_1_3']/t, st['pos_4_10']/t, st['pos_11_20']/t, st['pos_21_plus']/t])
            for c in range(2, 6): ws1.cell(row=ws1.max_row, column=c).number_format = '0.00%'
        pos_end_row = ws1.max_row

        # 画两张堆叠柱状图
        chart_intent = BarChart(); chart_intent.type = "col"; chart_intent.style = 2; chart_intent.grouping = "stacked"; chart_intent.overlap = 100
        chart_intent.title = "流量意图 (Search Intent) 切割图"; chart_intent.width = 18; chart_intent.height = 10
        chart_intent.add_data(Reference(ws1, min_col=2, min_row=intent_start_row, max_col=5, max_row=intent_end_row), titles_from_data=True)
        chart_intent.set_categories(Reference(ws1, min_col=1, min_row=intent_start_row+1, max_col=1, max_row=intent_end_row))
        
        chart_pos = BarChart(); chart_pos.type = "col"; chart_pos.style = 2; chart_pos.grouping = "stacked"; chart_pos.overlap = 100
        chart_pos.title = "流量排名区间阶梯图 (护城河健康度)"; chart_pos.width = 18; chart_pos.height = 10
        chart_pos.add_data(Reference(ws1, min_col=2, min_row=pos_start_row, max_col=5, max_row=pos_end_row), titles_from_data=True)
        chart_pos.set_categories(Reference(ws1, min_col=1, min_row=pos_start_row+1, max_col=1, max_row=pos_end_row))

        for chart in [chart_intent, chart_pos]:
            for i, s in enumerate(chart.series):
                try: s.graphicalProperties.solidFill = sw_colors[i % len(sw_colors)]
                except: pass

        ws1.add_chart(chart_intent, "I2")
        ws1.add_chart(chart_pos, "I20")
        for col in range(1, 10): ws1.column_dimensions[get_column_letter(col)].width = 20

        # ==========================================
        # --- Sheet 2: 单竞品独立频道深度分析 (🔥注入灵魂图表) ---
        # ==========================================
        print("[进程] 正在生成【竞品独立深度看板】及高级饼图...")
        ws2 = wb.create_sheet("单竞品深度诊断")
        ws2.sheet_view.showGridLines = False

        for col in range(1, 10):
            ws2.column_dimensions[get_column_letter(col)].width = 16

        for comp in comp_list:

            st = comp_stats[comp]

            t_traf = st['total_traffic']
            safe_traf = t_traf if t_traf>0 else 1
            
            ws2.append([f"🚀 {comp.upper()} 流量结构深度透视报告"])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
            for c in range(1, 10): ws2.cell(row=ws2.max_row, column=c).fill = PatternFill("solid", fgColor="2F5597")
            ws2.append([])
            
            # 🔥 基础面貌数据 (供图表取数)
            ws2.append(["基础面貌", "预估总流量", "流量总价值", "品牌词流量 %", "非品牌词(高质量) %", "KD极低(0-15)", "KD居中(16-40)", "KD极高(41+)"])
            title_r = ws2.max_row
            ws2.append([
                "", t_traf, st['traffic_value'], 
                st['branded_traf']/safe_traf, st['non_branded_traf']/safe_traf,
                st['kd_0_15']/safe_traf, st['kd_16_40']/safe_traf, st['kd_41_plus']/safe_traf
            ])
            data_r = ws2.max_row
            
            for c in range(1, 9):
                ws2.cell(row=title_r, column=c).font = Font(bold=True)
                ws2.cell(row=title_r, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
            ws2.cell(row=data_r, column=2).number_format = '#,##0'
            ws2.cell(row=data_r, column=3).number_format = '"$"#,##0'
            for c in range(4, 9): ws2.cell(row=data_r, column=c).number_format = '0.00%'
            
            # =======================
            # 🎨 注入灵魂：同级生成两张精美饼图
            # =======================
            # 图表 1：品牌词分布饼图
            pie_brand = PieChart()
            pie_brand.title = f"[{comp}] 真实搜索意图：品牌 vs 非品牌"
            pie_brand.style = 26 # 极简高端样式
            pie_brand.width = 9; pie_brand.height = 5
            data_brand = Reference(ws2, min_col=4, min_row=data_r, max_col=5, max_row=data_r)
            cats_brand = Reference(ws2, min_col=4, min_row=title_r, max_col=5, max_row=title_r)
            pie_brand.add_data(data_brand, from_rows=True, titles_from_data=False)
            pie_brand.set_categories(cats_brand)
            pie_brand.dataLabels = DataLabelList(); pie_brand.dataLabels.showPercent = True

            # 图表 2：KD 难度切割饼图
            pie_kd = PieChart()
            pie_kd.title = f"[{comp}] SEO 操作空间：KD 难度切割"
            pie_kd.style = 26
            pie_kd.width = 9; pie_kd.height = 5
            data_kd = Reference(ws2, min_col=6, min_row=data_r, max_col=8, max_row=data_r)
            cats_kd = Reference(ws2, min_col=6, min_row=title_r, max_col=8, max_row=title_r)
            pie_kd.add_data(data_kd, from_rows=True, titles_from_data=False)
            pie_kd.set_categories(cats_kd)
            pie_kd.dataLabels = DataLabelList(); pie_kd.dataLabels.showPercent = True

            # 将图表优雅地停靠在 J 列和 N 列，绝不遮挡数据！
            ws2.add_chart(pie_brand, f"J{title_r}")
            ws2.add_chart(pie_kd, f"N{title_r}")
            ws2.append([])

            # 继续下放国家、Top10页面等数据表
            ws2.append(["🌍 核心流量国家分布 (Top 5)"])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
            top_countries = sorted(st['countries'].items(), key=lambda x: x[1], reverse=True)[:5]
            for c_code, c_traf in top_countries:
                ws2.append([c_code, c_traf, c_traf/safe_traf])
                ws2.cell(row=ws2.max_row, column=2).number_format = '#,##0'
                ws2.cell(row=ws2.max_row, column=3).number_format = '0.00%'
            ws2.append([])
            
            ws2.append(["🏆 该竞品 Top 10 高价值页面榜单 (流量收割机)"])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=ws2.max_row, column=1).fill = PatternFill("solid", fgColor="C00000")
            ws2.append(["URL 路径", "预估页面总流量", "全站流量贡献比", "该页面覆盖核心词"])
            
            comp_pages = []
            if comp in page_stats:
                comp_pages = sorted(page_stats[comp].items(), key=lambda x: x[1]['traffic'], reverse=True)[:10]
            for p_url, p_data in comp_pages:
                kws_sample = ", ".join(list(p_data['kws'])[:5])
                ws2.append([p_url, p_data['traffic'], p_data['traffic']/safe_traf if t_traf else 0, kws_sample])
                ws2.cell(row=ws2.max_row, column=2).number_format = '#,##0'
                ws2.cell(row=ws2.max_row, column=3).number_format = '0.00%'
            ws2.append([])
            
            ws2.append(["💎 该竞品 Top 10 高价值通用词榜单 (Non-branded)"])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=ws2.max_row, column=1).fill = PatternFill("solid", fgColor="E5625E")
            ws2.append(["关键词 (Keyword)", "搜索量(Vol)", "难度(KD)", "单次点击(CPC)", "竞品当前排名", "预估带来流量", "着陆页 URL"])
            
            non_brand_kws = [k for k in st['keywords_data'] if k['is_branded'] == 0]
            non_brand_kws = sorted(non_brand_kws, key=lambda x: x['traf'], reverse=True)[:10]
            for kw in non_brand_kws:
                ws2.append([kw['kw'], kw['vol'], kw['kd'], kw['cpc'], kw['pos'], kw['traf'], kw['url']])
                ws2.cell(row=ws2.max_row, column=2).number_format = '#,##0'
                ws2.cell(row=ws2.max_row, column=6).number_format = '#,##0.0'
                
            for _ in range(7): ws2.append([]) # 给图表留足呼吸空间！

        # ==========================================
        # --- Sheet 3: 高价值词 Alpha 矩阵 ---
        # ==========================================
        ws3 = wb.create_sheet("高价值词穿透 Alpha 矩阵")
        ws3.append(["关键词 (Non-branded)", "重合度(多少竞品同时在做)", "竞争对手阵列", "全球搜索量(Volume)", "平均难度(KD)", "CPC价值", "该词带来的总自然流量 (全竞品累加)"])
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, cell in enumerate(ws3[1], 1):
            cell.font = header_font; cell.fill = PatternFill("solid", fgColor="C00000")
            ws3.column_dimensions[get_column_letter(col_num)].width = 22
        ws3.column_dimensions['C'].width = 35

        matrix_rows = []
        for kw, kw_data in domain_kw_map.items():
            overlap_score = len(kw_data['comps'])
            if overlap_score >= 1: 
                matrix_rows.append([
                    kw, overlap_score, ", ".join(kw_data['comps']), 
                    kw_data['total_vol'], kw_data['avg_kd'], kw_data['avg_cpc'], kw_data['total_traf']
                ])
                
        matrix_rows.sort(key=lambda x: (-x[1], -x[6], x[4]))
        for row in matrix_rows[:1500]: 
            ws3.append(row)
        ws3.freeze_panes = "A2"

        # ==========================================
        # --- Sheet 4: 原始清洗数据池 ---
        # ==========================================
        ws4 = wb.create_sheet("数据明细 (供筛选检索)")
        ws4.append(["竞品域名", "Keyword", "Country", "Is Branded", "Is Navigational", "Is Informational", "Is Commercial", "Is Transactional", "Volume", "KD", "CPC", "Current Traffic", "Current Position", "URL"])
        for col_num, cell in enumerate(ws4[1], 1):
            cell.font = header_font; cell.fill = PatternFill("solid", fgColor="4F81BD")
        
        cur.execute("SELECT competitor_domain, keyword, country, is_branded, is_navigational, is_informational, is_commercial, is_transactional, volume, kd, cpc, current_traffic, current_position, url FROM keywords")
        for row in cur.fetchall(): ws4.append(row)

        # ==========================================
        # 🚀🔥 Sheet 5: 关键词获取执行队列 (CRM级 SEO智能调度)
        # ==========================================
        print("[进程] 正在通过机器评分引擎生成【关键词获取执行队列 (CRM)】(Sheet 5)...")
        ws5 = wb.create_sheet("关键词获取执行队列(CRM)")
        
        headers_s5 = [
            # --- 机器自动推断区 (灰底) ---
            ("A", "优先级 (Priority)", "A6A6A6"),
            ("B", "目标词 (Target Keyword)", "A6A6A6"),
            ("C", "🤖机器评分 (Auto Score)", "A6A6A6"),
            ("D", "内容类型建议 (Action Suggested)", "A6A6A6"),
            ("E", "竞争对手重合度 (Alpha)", "A6A6A6"),
            ("F", "搜索量 (Volume)", "A6A6A6"),
            ("G", "难度 (KD)", "A6A6A6"),
            ("H", "单次点击价值 (CPC)", "A6A6A6"),
            ("I", "竞品累加截获流量 (Traffic)", "A6A6A6"),
            ("J", "目前有哪些竞品在做", "A6A6A6"),
            
            # --- AI 预留处理区 (绿底) ---
            ("K", "🧠AI: 业务相关性评分(0-100)", "70AD47"),
            ("L", "🧠AI: 生成爆款标题 (Title)", "70AD47"),
            ("M", "🧠AI: 大纲(H2/H3框架)", "70AD47"),
            
            # --- 人工执行区 (蓝底) ---
            ("N", "🧑‍💻负责编辑 (Editor)", "4472C4"),
            ("O", "🏆终极选词得分 [自动计算]", "FFC000"),  
            ("P", "🧑‍💻文章最终 URL", "4472C4"),
            ("Q", "🧑‍💻当前进度 (Status)", "4472C4")
        ]

        ws5.append([h[1] for h in headers_s5])
        for col_idx, h in enumerate(headers_s5, 1):
            cell = ws5.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=h[2])
            cell.alignment = Alignment(horizontal="center", vertical="center")

        scored_kws = []
        for kw, kw_data in domain_kw_map.items():
            vol = kw_data['total_vol']
            kd = kw_data['avg_kd']
            cpc = kw_data['avg_cpc']
            alpha_count = len(kw_data['comps'])
            
            # 1. 权重算分 (40% KD + 30% Alpha + 30% Vol) 
            score_kd = get_kw_kd_score(kd)
            score_alpha = get_kw_alpha_score(alpha_count)
            score_vol = get_kw_volume_score(vol)
            
            base_score = (score_kd * 0.40) + (score_alpha * 0.30) + (score_vol * 0.30)
            
            # CPC 商业价值直接拔高优先级
            bonus = 0
            if cpc >= 2.0: bonus = 15
            elif cpc >= 1.0: bonus = 10
            elif cpc >= 0.5: bonus = 5
            
            auto_score = min(round(base_score + bonus, 1), 100) # 封顶 100 分

            # 2. 定级
            if auto_score >= 80: priority = "P0 (核心狙击, 立刻写)"
            elif auto_score >= 65: priority = "P1 (重点排期, 必写)"
            elif auto_score >= 50: priority = "P2 (常规矩阵, 占坑)"
            elif auto_score >= 30: priority = "P3 (长尾观察, 随缘)"
            else: priority = "🗑️ (难度太高或没流量，放弃)"

            # 3. 意图落地策略 (Action)
            if kw_data.get('is_trans') or kw_data.get('is_comm'):
                action = "撰写高转化落地页(Landing Page) / 评测导购(Listicle/Comparison)"
            elif kw_data.get('is_info'):
                action = "撰写深度指南(How-to Guide) / 行业知识百科吸引长尾流量"
            else:
                action = "常规博客占坑，埋入内链引导至核心页面"

            scored_kws.append({
                'priority': priority, 'kw': kw, 'auto_score': auto_score, 'action': action,
                'alpha': alpha_count, 'vol': vol, 'kd': kd, 'cpc': cpc, 
                'traf': kw_data['total_traf'], 'comps': ", ".join(kw_data['comps'])
            })

        # 按机器打分 + 带来真实流量排序
        scored_kws.sort(key=lambda x: (x['auto_score'], x['traf']), reverse=True)

        for idx, d in enumerate(scored_kws[:2000], start=2): # 输出 Top 2000 个高潜力词防撑爆
            ws5.append([
                d['priority'], d['kw'], d['auto_score'], d['action'],
                d['alpha'], d['vol'], d['kd'], d['cpc'], d['traf'], d['comps'],
                "", "", "", # AI 区
                "", # Editor
                f"=(C{idx}*0.85) + (IF(ISNUMBER(K{idx}),K{idx},0)*0.15)", # O列公式
                "", # URL
                "未开始 (Not Started)"
            ])
            ws5.cell(row=idx, column=15).font = Font(bold=True, color="C00000")

        # 调列宽
        col_widths = {
            'A': 22, 'B': 25, 'C': 18, 'D': 45, 'E': 15, 
            'F': 12, 'G': 10, 'H': 10, 'I': 15, 'J': 45,
            'K': 25, 'L': 40, 'M': 45, 'N': 15, 'O': 25, 'P': 30, 'Q': 20
        }
        for col, width in col_widths.items():
            ws5.column_dimensions[col].width = width

        ws5.freeze_panes = "C2"
        ws5.auto_filter.ref = ws5.dimensions

        wb.save(excel_filename)
        print(f"\n[✓] 成功生成【自然流量全局洞察 & CRM智能排期大盘】：{excel_filename}")
        print("    -> 🎨 [新增] 子表2：已为各竞品的流量结构与KD切分注入高级可视化饼图！")
        print("    -> 🤖 [新增] 子表5：纯 Python 算分的机器评级引擎。根据 Alpha/KD/Vol/意图产出 P0-P3 执行排期！")
    else:
        print("[!] 缺少 openpyxl 库，无法生成 Excel 报表。请运行 pip install openpyxl")

def run_keyword_engine():
    print("=== 初始化自然搜索流量 (Organic Keywords) 侦察引擎 ===")
    conn = init_db()
    total_files, global_headers = process_ahrefs_exports(conn)
    if total_files > 0:
        generate_strategic_reports(conn, total_files, global_headers)
    conn.close()

if __name__ == "__main__":
    run_keyword_engine()