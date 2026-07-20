# backlink_matrix.py
import os
import sqlite3
import csv
import json
import datetime
import glob
from urllib.parse import urlparse
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, LineChart, Reference, BarChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "backlink_intelligence.db")
DATA_DIR = os.path.join(WORKSPACE_DIR, "ahrefs_data")
REPORT_DIR = os.path.join(WORKSPACE_DIR, "reports")

def get_current_time_str():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS backlinks")
    # 增加 links_in_group 字段，不耗费额度即可还原外链总盘
    cur.execute('''
        CREATE TABLE backlinks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competitor_domain TEXT,
            ref_domain TEXT,
            ref_url TEXT,
            target_url TEXT,
            domain_rating INTEGER,
            page_traffic INTEGER,
            page_type TEXT,
            page_category TEXT,
            is_dofollow INTEGER,
            is_spam INTEGER,
            links_in_group INTEGER DEFAULT 1,
            is_lost INTEGER DEFAULT 0,
            raw_data TEXT,
            UNIQUE(competitor_domain, ref_domain)
        )
    ''')
    cur.execute("CREATE INDEX idx_ref_domain ON backlinks(ref_domain)")
    conn.commit()
    return conn

def extract_domain(url):
    try:
        netloc = urlparse(url).netloc
        if netloc.startswith('www.'):
            return netloc[4:]
        return netloc
    except Exception:
        return ""

def process_ahrefs_exports(conn):
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
        print(f"[提示] 请将 Ahrefs 数据放入 {DATA_DIR} 目录后重新运行。")
        return 0, []

    files = glob.glob(os.path.join(DATA_DIR, "*.*sv"))
    if not files:
        print(f"[提示] 在 {DATA_DIR} 中未找到 csv/tsv 文件。")
        return 0, []

    cur = conn.cursor()
    total_files = len(files)
    global_headers = [] 
    
    for file_path in files:
        if "overview-competitors" in file_path.lower():
            continue
        filename = os.path.basename(file_path).lower()
        competitor = filename.split('-backlinks')[0] if '-backlinks' in filename else filename.split('_')[0].replace('.csv', '')
        print(f"正在清洗并提取竞品 [{competitor}] 的引荐域名(Ref Domain)级数据...")
        
        try:
            f = open(file_path, 'r', encoding='utf-16')
            f.read(1); f.seek(0)
        except UnicodeError:
            f = open(file_path, 'r', encoding='utf-8')

        reader = csv.DictReader(f, delimiter='\t' if file_path.endswith('.tsv') or 'utf-16' in str(f.encoding) else ',')
        
        cleaned_fieldnames = [str(fn).replace('\ufeff', '').strip() for fn in reader.fieldnames if fn]
        for fn in cleaned_fieldnames:
            if fn not in global_headers:
                global_headers.append(fn)
        
        insert_data = []
        for raw_row in reader:
            clean_row = {str(k).replace('\ufeff', '').strip(): v for k, v in raw_row.items() if k}
            lower_row = {k.lower(): v for k, v in clean_row.items()}
            
            ref_url = lower_row.get('referring page url') or lower_row.get('source url')
            if not ref_url: continue
                
            ref_domain = extract_domain(ref_url)
            target_url = lower_row.get('target url', '')
            
            try: dr = int(float(lower_row.get('domain rating') or lower_row.get('dr') or 0))
            except (ValueError, TypeError): dr = 0

            try: traffic = int(float(lower_row.get('page traffic') or lower_row.get('traffic') or 0))
            except (ValueError, TypeError): traffic = 0
            
            # 💡 核心升级：提取 Links in group 计算真实外链量
            try: links_in_group = int(float(lower_row.get('links in group', 1)))
            except (ValueError, TypeError): links_in_group = 1

            # 🔴 新增：如果 lost 字段有内容，视为该渠道代表外链已丢失
            is_lost = 1 if lower_row.get('lost') and str(lower_row.get('lost')).strip() != '' else 0


            page_type = lower_row.get('page type', 'Unknown')
            page_category = lower_row.get('page category', 'Unknown')
            
            is_dofollow = 1 if str(lower_row.get('nofollow', 'FALSE')).upper() != 'TRUE' else 0
            is_spam = 1 if str(lower_row.get('is spam', 'FALSE')).upper() == 'TRUE' else 0
            
            raw_json = json.dumps(clean_row, ensure_ascii=False)

            insert_data.append((
                competitor, ref_domain, ref_url, target_url, 
                dr, traffic, page_type, page_category, is_dofollow, is_spam, links_in_group, is_lost, raw_json
            ))
            
        f.close()
        
        cur.executemany('''
            INSERT OR IGNORE INTO backlinks 
            (competitor_domain, ref_domain, ref_url, target_url, domain_rating, page_traffic, page_type, page_category, is_dofollow, is_spam, links_in_group, is_lost, raw_data) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', insert_data)
        
    conn.commit()
    return total_files, global_headers

def generate_strategic_reports(conn, total_files, global_headers):
    if not os.path.exists(REPORT_DIR):
        os.makedirs(REPORT_DIR)

    cur = conn.cursor()
    overlap_threshold = 1

    print("\n[进程] 正在从数据库拉取独立域名级(Domain-level)数据...")
    cur.execute('''
        SELECT competitor_domain, ref_domain, ref_url, target_url, domain_rating, page_traffic, page_type, page_category, raw_data, is_dofollow, is_spam, links_in_group, is_lost 
        FROM backlinks 
    ''')
    all_links = cur.fetchall()

    domain_map = {}
    comp_stats = {}  
    comp_time_stats = {}  
    comp_time_stats_live = {}  # 🔴 增加 Live 时间字典

    print(f"[进程] 成功提纯 {len(all_links)} 个独立引荐域名，正在进行多维聚类计算与图表数据准备...")

    for row in all_links:
        comp, ref_d, ref_u, tgt_u, dr, traf, p_type, p_cat, raw_j, is_dofollow, is_spam, links_in_group, is_lost = row
        raw_dict = json.loads(raw_j)

        if comp not in comp_stats:
            comp_stats[comp] = {
                'total_domains': 0, 'total_backlinks': 0, '0-20': 0, '20-40': 0, '40-60': 0, '60-80': 0, '80-100': 0, 'dofollow': 0, 'spam': 0,
                'live_domains': 0, 'live_backlinks': 0, 'live_0-20': 0, 'live_20-40': 0, 'live_40-60': 0, 'live_60-80': 0, 'live_80-100': 0, 'live_dofollow': 0, 'live_spam': 0
            }
        
        comp_stats[comp]['total_domains'] += 1
        comp_stats[comp]['total_backlinks'] += links_in_group # 计算总外链数
        
        if dr <= 20: comp_stats[comp]['0-20'] += 1
        elif dr <= 40: comp_stats[comp]['20-40'] += 1
        elif dr <= 60: comp_stats[comp]['40-60'] += 1
        elif dr <= 80: comp_stats[comp]['60-80'] += 1
        else: comp_stats[comp]['80-100'] += 1

        if is_dofollow == 1: comp_stats[comp]['dofollow'] += 1
        if is_spam == 1: comp_stats[comp]['spam'] += 1

        # 🔴 如果这条没丢，追加到 Live 资产里
        if is_lost == 0:
            comp_stats[comp]['live_domains'] += 1
            comp_stats[comp]['live_backlinks'] += links_in_group
            if dr <= 20: comp_stats[comp]['live_0-20'] += 1
            elif dr <= 40: comp_stats[comp]['live_20-40'] += 1
            elif dr <= 60: comp_stats[comp]['live_40-60'] += 1
            elif dr <= 80: comp_stats[comp]['live_60-80'] += 1
            else: comp_stats[comp]['live_80-100'] += 1
            if is_dofollow == 1: comp_stats[comp]['live_dofollow'] += 1
            if is_spam == 1: comp_stats[comp]['live_spam'] += 1

        first_seen = str(raw_dict.get('first seen', raw_dict.get('First seen', '')))
        year = first_seen[:4] if len(first_seen) >= 4 and first_seen[:4].isdigit() else 'Unknown'
        
        if comp not in comp_time_stats:
            comp_time_stats[comp] = {}
        if comp not in comp_time_stats_live:
            comp_time_stats_live[comp] = {}
            
        if year != 'Unknown':
            comp_time_stats[comp][year] = comp_time_stats[comp].get(year, 0) + 1
            if is_lost == 0:
                comp_time_stats_live[comp][year] = comp_time_stats_live[comp].get(year, 0) + 1

        if ref_d not in domain_map:
            domain_map[ref_d] = {
                'competitors': set(), 'max_dr': 0, 'max_traffic': 0, 'types': set(), 'categories': set(),
                'link_details': [], 'raw_rows': [], 'is_dofollow': 0, 'is_spam': 1, 'domain_total_links': 0
            }
        d = domain_map[ref_d]
        d['competitors'].add(comp)
        d['max_dr'] = max(d['max_dr'], dr)
        d['max_traffic'] = max(d['max_traffic'], traf)
        d['domain_total_links'] += links_in_group
        if is_dofollow == 1: d['is_dofollow'] = 1 
        if is_spam == 0: d['is_spam'] = 0         
        if p_type and p_type != 'Unknown': d['types'].add(p_type)
        if p_cat and p_cat != 'Unknown': d['categories'].add(p_cat)
        
        # 🔥 新增：提取锚文本 (Anchor) 拼接入范例
        anchor = str(raw_dict.get('Anchor', raw_dict.get('anchor', ''))).strip()
        if not anchor: anchor = "无锚文本"
        d['link_details'].append(f"[{comp}] {ref_u}  --[{anchor}]-->  {tgt_u}")
        
        d['raw_rows'].append((comp, ref_u, tgt_u, raw_j))

    print("[进程] 正在按策略权重进行高质量域名智能排序...")

    top_domains = []
    for ref_d, d_info in domain_map.items():
        alpha_score = len(d_info['competitors'])
        if alpha_score >= overlap_threshold:
            d_info['alpha'] = alpha_score
            d_info['ref_domain'] = ref_d
            top_domains.append(d_info)

    top_domains.sort(key=lambda x: (-x['is_spam'], x['is_dofollow'], x['alpha'], x['max_dr'], x['max_traffic']), reverse=True)

    matrix_rows = []    
    detail_rows = []    
    json_payloads = []  

    for d in top_domains:
        p_types = list(d['types'])
        p_cats = list(d['categories'])
        clean_types = list(set([t.split('>')[-1].strip() for t in p_types if t]))
        clean_cats = list(set([c.split('>')[0].strip() for c in p_cats if c]))
        vector_format = clean_types[0] if clean_types else "Blog/Article"
        vector_industry = clean_cats[0] if clean_cats else "General B2B"
        shared_by = ", ".join(d['competitors'])
        clickable_references = "\n".join(list(set(d['link_details'])))

        spam_str = "是(Spam沉底)" if d.get('is_spam') == 1 else "否"
        dof_str = "是" if d.get('is_dofollow') == 1 else "否(Nofollow)"

        matrix_rows.append([
            d['ref_domain'], d['alpha'], d['domain_total_links'], d['max_dr'], d['max_traffic'], 
            vector_industry, vector_format, shared_by, spam_str, dof_str, clickable_references
        ])

        for link_row in d['raw_rows']:
            comp, ref_u, tgt_u, raw_j = link_row
            raw_dict = json.loads(raw_j)
            row_detail = [d['alpha'], comp]
            for header in global_headers:
                row_detail.append(raw_dict.get(header, ""))
            detail_rows.append(row_detail)

        prompt_template = f"你是一个资深的跨境电商 B2B 营销专家。现在我要向一个【渠道调性：{vector_industry}，文章类型偏好：{vector_format}】的权威网站投稿。该网站权重 DR 为 {d['max_dr']}。我的核心关键词是【家具一件代发供应商】。请严格根据以上特征，生成一篇字数2000字，带有对比表格，适合该渠道发布的专业文章初稿。"
        json_payloads.append({
            "target_domain": d['ref_domain'],
            "overlap_alpha_score": d['alpha'],
            "domain_rating": d['max_dr'],
            "agent_prompt": prompt_template
        })

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    excel_filename = os.path.join(REPORT_DIR, f"Backlink_Matrix_Pro_{date_str}.xlsx")
    
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        
        # ==========================================
        # --- Sheet 1 & Sheet 2 ---
        # ==========================================
        ws1 = wb.active
        ws1.title = "战略大盘 (Alpha矩阵)"
        # 增加一列该域名下的总外链数
        matrix_headers = ["来源域名 (Ref Domain)", "重合度 (Alpha Score)", "该域名下总外链簇数", "网站权重 (DR)", "最高页面流量", "核心行业标签", "核心内容类型", "被哪些竞品获取", "是否Spam", "是否Dofollow", "对标域名下的具体外链范例"]
        ws1.append(matrix_headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        for col_num, cell in enumerate(ws1[1], 1):
            cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        
        ws1.column_dimensions['A'].width = 30; ws1.column_dimensions['H'].width = 25; ws1.column_dimensions['K'].width = 100 
        for idx, row in enumerate(matrix_rows, start=2):
            ws1.append(row)
            ws1.cell(row=idx, column=11).alignment = Alignment(wrap_text=True, vertical='center')

        ws2 = wb.create_sheet("外链全维度明细 (支持筛选查阅)")
        ws2.append(["重合度 (Alpha Score)", "具体竞品 (Competitor)"] + global_headers)
        for col_num, cell in enumerate(ws2[1], 1):
            cell.font = header_font; cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for row in detail_rows:
            ws2.append(row)

        # ==========================================
        # --- Sheet 3: 单竞品看版 (紧凑高颜值版) ---
        # ==========================================
        print("[进程] 正在生成【竞品独立域名战略看板】...")
        ws3 = wb.create_sheet("单竞品深度分析")
        ws3.sheet_view.showGridLines = False  
        
        ws3.column_dimensions['A'].width = 18
        for i in range(2, 20):
            ws3.column_dimensions[get_column_letter(i)].width = 15

        comp_list = list(comp_stats.keys())
        all_years = sorted(list(set(year for ts in comp_time_stats.values() for year in ts.keys())))
        
        for comp in comp_list:
            stats = comp_stats[comp]
            total_domains = stats['total_domains']
            total_backlinks = stats['total_backlinks']
            if total_domains == 0: continue
            
            # 1. 标题区
            ws3.append([f"📊 {comp.upper()} 渠道质量深度诊断看板"])
            header_row = ws3.max_row
            ws3.cell(row=header_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
            for col in range(1, 15):
                ws3.cell(row=header_row, column=col).fill = PatternFill("solid", fgColor="2F5597") 
            ws3.append([]) 
            
            # 2. 核心高价值指标区 (融合 Live 和 流失率)
            live_domains = stats['live_domains']
            live_backlinks = stats['live_backlinks']
            lost_rate = (total_domains - live_domains) / total_domains if total_domains else 0
            live_dof_pct = stats['live_dofollow'] / live_domains if live_domains else 0
            live_high_dr_pct = (stats['live_40-60'] + stats['live_60-80'] + stats['live_80-100']) / live_domains if live_domains else 0
            
            # 🔥 文字修改为 (Last 5 years)
            ws3.append(["核心数据资产", "历史引荐域名总数(Last 5 years)", "历史预估外链数", "流失率(Lost %)", "现存引荐域名(Live)", "现存预估外链数(Live)", "现存Dofollow占比", "现存高权重占比"])
            metric_title_row = ws3.max_row
            ws3.append(["", total_domains, total_backlinks, f"{lost_rate:.1%}", live_domains, live_backlinks, f"{live_dof_pct:.1%}", f"{live_high_dr_pct:.1%}"])
            metric_data_row = ws3.max_row
            
            for c in range(1, 9):
                ws3.cell(row=metric_title_row, column=c).font = Font(bold=True)
                ws3.cell(row=metric_title_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
                ws3.cell(row=metric_title_row, column=c).alignment = Alignment(horizontal="center")
                ws3.cell(row=metric_data_row, column=c).alignment = Alignment(horizontal="center")
            ws3.append([]) 
            
            # 3. DR 分布数据表 (🔥加入 Live 统计行)
            ws3.append(["DR 权重分布", "0-20 (极差)", "20-40 (普通)", "40-60 (优质)", "60-80 (权威)", "80-100 (顶配)"])
            dr_title_row = ws3.max_row
            ws3.append(["历史域名数量", stats['0-20'], stats['20-40'], stats['40-60'], stats['60-80'], stats['80-100']])
            dr_data_row = ws3.max_row
            ws3.append(["历史占盘比", f"{stats['0-20']/total_domains:.1%}" if total_domains else "0%", f"{stats['20-40']/total_domains:.1%}" if total_domains else "0%", f"{stats['40-60']/total_domains:.1%}" if total_domains else "0%", f"{stats['60-80']/total_domains:.1%}" if total_domains else "0%", f"{stats['80-100']/total_domains:.1%}" if total_domains else "0%"])
            ws3.append(["Live 域名数量", stats['live_0-20'], stats['live_20-40'], stats['live_40-60'], stats['live_60-80'], stats['live_80-100']])
            dr_live_data_row = ws3.max_row
            
            for c in range(1, 7):
                ws3.cell(row=dr_title_row, column=c).font = Font(bold=True)
                ws3.cell(row=dr_title_row, column=c).fill = PatternFill("solid", fgColor="F2F2F2")
                ws3.cell(row=dr_title_row, column=c).alignment = Alignment(horizontal="center")
                ws3.cell(row=dr_data_row, column=c).alignment = Alignment(horizontal="center")
                ws3.cell(row=dr_data_row+1, column=c).alignment = Alignment(horizontal="center")
                ws3.cell(row=dr_live_data_row, column=c).alignment = Alignment(horizontal="center")
            ws3.append([]) 
            
            # 4. 历年增长趋势表 (分离 历史 与 Live)
            ws3.append(["年份趋势"] + all_years)
            time_title_row = ws3.max_row
            
            new_counts = [comp_time_stats[comp].get(y, 0) for y in all_years]
            live_new_counts = [comp_time_stats_live[comp].get(y, 0) for y in all_years]
            
            cum_counts = []; cumulative = 0
            live_cum_counts = []; cumulative_live = 0
            for i in range(len(all_years)):
                cumulative += new_counts[i]
                cum_counts.append(cumulative)
                cumulative_live += live_new_counts[i]
                live_cum_counts.append(cumulative_live)
                
            ws3.append(["(历史)当年新增"] + new_counts)
            time_new_row = ws3.max_row
            ws3.append(["(历史)累加总盘"] + cum_counts)
            time_cum_row = ws3.max_row
            
            ws3.append(["(Live)当年新增"] + live_new_counts)
            time_new_live_row = ws3.max_row
            ws3.append(["(Live)累加现存"] + live_cum_counts)
            time_cum_live_row = ws3.max_row
            
            for r in [time_title_row, time_new_row, time_cum_row, time_new_live_row, time_cum_live_row]:
                for c in range(1, len(all_years)+2):
                    if r == time_title_row:
                        ws3.cell(row=r, column=c).font = Font(bold=True)
                        ws3.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F2F2F2")
                    ws3.cell(row=r, column=c).alignment = Alignment(horizontal="center")
            ws3.append([]) 
            
            # --- 🔥 生成 4 张精美图表 (历史 2 张，Live 2 张) ---
            chart_anchor_row = ws3.max_row + 1
            sw_colors = ["294266", "F7941D", "20B799", "FDB913", "3FBBDF", "8C67AB", "E5625E", "8391A5"]
            
            # (1) 历史折线图
            line_hist = LineChart()
            line_hist.title = f"[{comp}] - 历史增长趋势 (努力痕迹)"
            line_hist.style = 13
            line_hist.width = 15
            line_hist.height = 7.5
            line_hist.add_data(Reference(ws3, min_col=1, min_row=time_new_row, max_col=len(all_years)+1, max_row=time_cum_row), from_rows=True, titles_from_data=True)
            line_hist.set_categories(Reference(ws3, min_col=2, min_row=time_title_row, max_col=len(all_years)+1, max_row=time_title_row))
            for i, s in enumerate(line_hist.series):
                try: 
                    s.graphicalProperties.line.solidFill = sw_colors[i % len(sw_colors)]
                    s.graphicalProperties.line.width = 30000 
                except: 
                    pass
                
            # (2) 历史饼图
            pie_hist = PieChart()
            pie_hist.title = f"[{comp}] - 历史 DR 资产分布"
            pie_hist.style = 26
            pie_hist.width = 12
            pie_hist.height = 7.5
            pie_hist.add_data(Reference(ws3, min_col=2, min_row=dr_data_row, max_col=6, max_row=dr_data_row), from_rows=True, titles_from_data=False)
            pie_hist.set_categories(Reference(ws3, min_col=2, min_row=dr_title_row, max_col=6, max_row=dr_title_row))
            pie_hist.dataLabels = DataLabelList()
            pie_hist.dataLabels.showPercent = True
            pie_hist.dataLabels.showVal = False
            pie_hist.dataLabels.showSerName = False
            
            # (3) Live 折线图
            line_live = LineChart()
            line_live.title = f"[{comp}] - 现存外链趋势 (Live实际资产)"
            line_live.style = 13
            line_live.width = 15
            line_live.height = 7.5
            line_live.add_data(Reference(ws3, min_col=1, min_row=time_new_live_row, max_col=len(all_years)+1, max_row=time_cum_live_row), from_rows=True, titles_from_data=True)
            line_live.set_categories(Reference(ws3, min_col=2, min_row=time_title_row, max_col=len(all_years)+1, max_row=time_title_row))
            for i, s in enumerate(line_live.series):
                try: 
                    s.graphicalProperties.line.solidFill = sw_colors[(i+2) % len(sw_colors)]
                    s.graphicalProperties.line.width = 30000 
                except: 
                    pass
                
            # (4) Live 饼图
            pie_live = PieChart()
            pie_live.title = f"[{comp}] - 现存 DR 资产分布 (Live)"
            pie_live.style = 26
            pie_live.width = 12
            pie_live.height = 7.5
            pie_live.add_data(Reference(ws3, min_col=2, min_row=dr_live_data_row, max_col=6, max_row=dr_live_data_row), from_rows=True, titles_from_data=False)
            pie_live.set_categories(Reference(ws3, min_col=2, min_row=dr_title_row, max_col=6, max_row=dr_title_row))
            pie_live.dataLabels = DataLabelList()
            pie_live.dataLabels.showPercent = True
            pie_live.dataLabels.showVal = False
            pie_live.dataLabels.showSerName = False

            # 将 4 个图表并排叠加布置！
            ws3.add_chart(line_hist, f"A{chart_anchor_row}")
            ws3.add_chart(pie_hist, f"I{chart_anchor_row}")
            ws3.add_chart(line_live, f"A{chart_anchor_row + 15}")
            ws3.add_chart(pie_live, f"I{chart_anchor_row + 15}")
            
            # 将底部空行设为 32 行，给上下两排图表预留足够的高度！
            for _ in range(32):
                ws3.append([])

        # ==========================================
        # --- Sheet 4: 全局竞品大盘横向对比 (Super Dashboard) ---
        # ==========================================
        if all_years:
            print("[进程] 正在生成核心模块：【全局竞品大盘横向对比】(Super Dashboard)...")
            ws4 = wb.create_sheet("全局竞品大盘对比")
            ws4.sheet_view.showGridLines = False

            data_start_row = 100
            current_row = data_start_row
            
            # ------ 1. 新增域名数表 ------
            ws4.cell(row=current_row, column=1, value="[底层数据源] 全竞品当年新增引荐域名").font = Font(bold=True)
            current_row += 1
            headers_time = ["竞品名称"] + all_years
            ws4.append(headers_time)
            new_links_start_row = current_row
            
            comp_new_data = {}
            comp_cum_data = {}
            for comp in comp_list:
                new_counts = [comp_time_stats[comp].get(y, 0) for y in all_years]
                comp_new_data[comp] = new_counts
                cum_counts = []; cumulative = 0
                for nc in new_counts:
                    cumulative += nc
                    cum_counts.append(cumulative)
                comp_cum_data[comp] = cum_counts
                ws4.append([comp] + new_counts)
                
            new_links_end_row = ws4.max_row
            current_row = ws4.max_row + 2
            
            # ------ 2. 累加域名数表 ------
            ws4.cell(row=current_row, column=1, value="[底层数据源] 全竞品累加引荐域名总规模").font = Font(bold=True)
            current_row += 1
            ws4.append(headers_time)
            cum_links_start_row = current_row
            
            for comp in comp_list:
                ws4.append([comp] + comp_cum_data[comp])
                
            cum_links_end_row = ws4.max_row
            current_row = ws4.max_row + 2
            
            # ------ 3. 【全新增】全站链接作弊侦测 (Sitewide Spam Index) ------
            ws4.cell(row=current_row, column=1, value="[底层数据源] 外链广度 vs 作弊虚高指数").font = Font(bold=True)
            current_row += 1
            headers_sitewide = ["竞品名称", "引荐域名(Ref Domains)", "预估外链总数(Total Backlinks)", "单域名外链比 (Sitewide指标)"]
            ws4.append(headers_sitewide)
            sitewide_start_row = current_row
            
            for comp in comp_list:
                stats = comp_stats[comp]
                t_dom = stats['total_domains']
                t_bkl = stats['total_backlinks']
                ratio = t_bkl / t_dom if t_dom else 0
                ws4.append([comp, t_dom, t_bkl, ratio])
                
            sitewide_end_row = ws4.max_row
            current_row = ws4.max_row + 2

            # ------ 4. 核心健康度对比表 ------
            ws4.cell(row=current_row, column=1, value="[底层数据源] 核心引荐域名健康度对比").font = Font(bold=True)
            current_row += 1
            headers_health = ["竞品名称", "Dofollow占比", "Spam垃圾占比", "高权重(DR>40)占比"]
            ws4.append(headers_health)
            health_start_row = current_row
            
            for comp in comp_list:
                stats = comp_stats[comp]
                t_dom = stats['total_domains']
                if t_dom == 0:
                    ws4.append([comp, 0, 0, 0])
                    continue
                dof_pct = stats['dofollow'] / t_dom
                spam_pct = stats['spam'] / t_dom
                high_dr_pct = (stats['40-60'] + stats['60-80'] + stats['80-100']) / t_dom
                ws4.append([comp, dof_pct, spam_pct, high_dr_pct])
                for c in range(2, 5):
                    ws4.cell(row=ws4.max_row, column=c).number_format = '0.00%'
                    
            health_end_row = ws4.max_row
            current_row = ws4.max_row + 2
            
            # ------ 5. DR 分布结构表 ------
            ws4.cell(row=current_row, column=1, value="[底层数据源] 引荐域名 DR 权重结构").font = Font(bold=True)
            current_row += 1
            headers_dr = ["竞品名称", "0-20(极差)", "20-40(普通)", "40-60(优质)", "60-80(权威)", "80-100(顶配)"]
            ws4.append(headers_dr)
            dr_start_row = current_row
            
            for comp in comp_list:
                stats = comp_stats[comp]
                t_dom = stats['total_domains']
                if t_dom == 0:
                    ws4.append([comp, 0, 0, 0, 0, 0])
                    continue
                ws4.append([
                    comp,
                    stats['0-20'] / t_dom, stats['20-40'] / t_dom,
                    stats['40-60'] / t_dom, stats['60-80'] / t_dom,
                    stats['80-100'] / t_dom
                ])
                for c in range(2, 7):
                    ws4.cell(row=ws4.max_row, column=c).number_format = '0.00%'
                    
            dr_end_row = ws4.max_row

            # ==========================
            # 💎 5 张 Super Dashboard 高端图表 (矩阵排布)
            # ==========================
            cats_time = Reference(ws4, min_col=2, min_row=cum_links_start_row, max_col=len(all_years)+1, max_row=cum_links_start_row)

            # --- 图 1: 累计域名 (A2) ---
            chart_cum = LineChart()
            chart_cum.title = "全竞品引荐域名走势 (谁的护城河最广)"
            chart_cum.style = 2
            chart_cum.y_axis.title = "总域名规模 (个)"
            chart_cum.width = 23; chart_cum.height = 12
            data_cum = Reference(ws4, min_col=1, min_row=cum_links_start_row+1, max_col=len(all_years)+1, max_row=cum_links_end_row)
            chart_cum.add_data(data_cum, from_rows=True, titles_from_data=True)
            chart_cum.set_categories(cats_time)

            # --- 图 2: 健康度拆解 (M2) ---
            chart_health = BarChart()
            chart_health.type = "col"
            chart_health.style = 2 
            chart_health.grouping = "clustered"
            chart_health.title = "域名资产健康度拆解 (Dofollow / Spam / 高权重)"
            chart_health.y_axis.title = "占其引荐域名总数的百分比"
            chart_health.y_axis.number_format = '0%'
            chart_health.width = 23; chart_health.height = 12
            data_health = Reference(ws4, min_col=2, min_row=health_start_row, max_col=4, max_row=health_end_row)
            cats_health = Reference(ws4, min_col=1, min_row=health_start_row+1, max_col=1, max_row=health_end_row)
            chart_health.add_data(data_health, titles_from_data=True)
            chart_health.set_categories(cats_health)

            # --- 图 3: DR 结构 100% 堆叠 (A24) ---
            chart_dr = BarChart()
            chart_dr.type = "col"
            chart_dr.style = 2
            chart_dr.grouping = "stacked"
            chart_dr.overlap = 100
            chart_dr.title = "DR 权重结构横向切割 (谁掌握顶层高权重渠道)"
            chart_dr.y_axis.title = "各 DR 梯队占比 (拉平至100%)"
            chart_dr.y_axis.number_format = '0%'
            chart_dr.width = 23; chart_dr.height = 12
            data_dr = Reference(ws4, min_col=2, min_row=dr_start_row, max_col=6, max_row=dr_end_row)
            cats_dr = Reference(ws4, min_col=1, min_row=dr_start_row+1, max_col=1, max_row=dr_end_row)
            chart_dr.add_data(data_dr, titles_from_data=True)
            chart_dr.set_categories(cats_dr)

            # --- 图 4: 🔥 全新 - Sitewide 作弊虚高预警图 (M24) ---
            chart_sitewide = BarChart()
            chart_sitewide.type = "col"
            chart_sitewide.style = 2
            chart_sitewide.grouping = "clustered"
            chart_sitewide.title = "单域名外链比 (预警: 比例超高说明重度依赖全站底栏垃圾链接)"
            chart_sitewide.y_axis.title = "平均单域名外链数 (Ratio)"
            chart_sitewide.width = 23; chart_sitewide.height = 12
            data_sw = Reference(ws4, min_col=4, min_row=sitewide_start_row, max_col=4, max_row=sitewide_end_row)
            cats_sw = Reference(ws4, min_col=1, min_row=sitewide_start_row+1, max_col=1, max_row=sitewide_end_row)
            chart_sitewide.add_data(data_sw, titles_from_data=True)
            chart_sitewide.set_categories(cats_sw)

            # 注入 Similarweb 高级定制色系
            sw_colors = ["294266", "F7941D", "20B799", "FDB913", "3FBBDF", "8C67AB", "E5625E", "8391A5"]
            
            for i, s in enumerate(chart_cum.series):
                try: s.graphicalProperties.line.solidFill = sw_colors[i % len(sw_colors)]; s.graphicalProperties.line.width = 30000  
                except: pass
            
            for chart in [chart_health, chart_dr, chart_sitewide]:
                for i, s in enumerate(chart.series):
                    try: s.graphicalProperties.solidFill = sw_colors[i % len(sw_colors)]
                    except: pass

            ws4.add_chart(chart_cum, "A2")
            ws4.add_chart(chart_health, "M2")
            ws4.add_chart(chart_dr, "A24")
            ws4.add_chart(chart_sitewide, "M24") # 第四张图换成极具价值的作弊侦测图！



            # ==========================
            # 🔴 🔥 无缝追加：现存外链(Live) 专属大盘数据及 4 张图表
            # ==========================
            current_row = ws4.max_row + 5 
            
            ws4.cell(row=current_row, column=1, value="[🔴 LIVE 底层数据源] 现存累加引荐域名总规模").font = Font(bold=True)
            current_row += 1
            ws4.append(headers_time)
            live_cum_start_row = current_row
            for comp in comp_list:
                live_new = [comp_time_stats_live[comp].get(y, 0) for y in all_years]
                live_cum = []; cum = 0
                for n in live_new: cum += n; live_cum.append(cum)
                ws4.append([comp] + live_cum)
            live_cum_end_row = ws4.max_row
            current_row = ws4.max_row + 2
            
            ws4.cell(row=current_row, column=1, value="[🔴 LIVE 底层数据源] 现存引荐域名健康度对比").font = Font(bold=True)
            current_row += 1
            ws4.append(["竞品名称", "Live Dofollow占比", "Live Spam垃圾占比", "Live 高权重(DR>40)占比"])
            live_health_start_row = current_row
            for comp in comp_list:
                stats = comp_stats[comp]
                ld = stats['live_domains']
                if ld == 0: ws4.append([comp, 0, 0, 0])
                else:
                    ws4.append([comp, stats['live_dofollow']/ld, stats['live_spam']/ld, (stats['live_40-60']+stats['live_60-80']+stats['live_80-100'])/ld])
                    for c in range(2, 5): ws4.cell(row=ws4.max_row, column=c).number_format = '0.00%'
            live_health_end_row = ws4.max_row
            current_row = ws4.max_row + 2
            
            ws4.cell(row=current_row, column=1, value="[🔴 LIVE 底层数据源] 现存引荐域名 DR 权重结构").font = Font(bold=True)
            current_row += 1
            ws4.append(["竞品名称", "0-20(极差)", "20-40(普通)", "40-60(优质)", "60-80(权威)", "80-100(顶配)"])
            live_dr_start_row = current_row
            for comp in comp_list:
                stats = comp_stats[comp]
                ld = stats['live_domains']
                if ld == 0: ws4.append([comp, 0, 0, 0, 0, 0])
                else:
                    ws4.append([comp, stats['live_0-20']/ld, stats['live_20-40']/ld, stats['live_40-60']/ld, stats['live_60-80']/ld, stats['live_80-100']/ld])
                    for c in range(2, 7): ws4.cell(row=ws4.max_row, column=c).number_format = '0.00%'
            live_dr_end_row = ws4.max_row
            current_row = ws4.max_row + 2
            
            ws4.cell(row=current_row, column=1, value="[🔴 LIVE 底层数据源] 现存单域名外链比 (Sitewide)").font = Font(bold=True)
            current_row += 1
            ws4.append(["竞品名称", "现存引荐域名(Live)", "现存总外链(Live)", "单域名外链比"])
            live_sw_start_row = current_row
            for comp in comp_list:
                stats = comp_stats[comp]
                ld = stats['live_domains']
                lb = stats['live_backlinks']
                ws4.append([comp, ld, lb, lb/ld if ld else 0])
            live_sw_end_row = ws4.max_row

            # --- 画 4 个专属的 LIVE 仪表盘图表 ---
            chart_cum_live = LineChart(); chart_cum_live.title = "[🔴Live 现存] 全竞品真实护城河走势"; chart_cum_live.style = 2; chart_cum_live.width = 23; chart_cum_live.height = 12
            chart_cum_live.add_data(Reference(ws4, min_col=1, min_row=live_cum_start_row+1, max_col=len(all_years)+1, max_row=live_cum_end_row), from_rows=True, titles_from_data=True)
            chart_cum_live.set_categories(Reference(ws4, min_col=2, min_row=live_cum_start_row, max_col=len(all_years)+1, max_row=live_cum_start_row))

            chart_health_live = BarChart(); chart_health_live.type = "col"; chart_health_live.style = 2; chart_health_live.grouping = "clustered"; chart_health_live.title = "[🔴Live 现存] 域名资产健康度"
            chart_health_live.width = 23; chart_health_live.height = 12
            chart_health_live.add_data(Reference(ws4, min_col=2, min_row=live_health_start_row, max_col=4, max_row=live_health_end_row), titles_from_data=True)
            chart_health_live.set_categories(Reference(ws4, min_col=1, min_row=live_health_start_row+1, max_col=1, max_row=live_health_end_row))

            chart_dr_live = BarChart(); chart_dr_live.type = "col"; chart_dr_live.style = 2; chart_dr_live.grouping = "stacked"; chart_dr_live.overlap = 100; chart_dr_live.title = "[🔴Live 现存] DR 权重真实结构"
            chart_dr_live.width = 23; chart_dr_live.height = 12
            chart_dr_live.add_data(Reference(ws4, min_col=2, min_row=live_dr_start_row, max_col=6, max_row=live_dr_end_row), titles_from_data=True)
            chart_dr_live.set_categories(Reference(ws4, min_col=1, min_row=live_dr_start_row+1, max_col=1, max_row=live_dr_end_row))

            chart_sw_live = BarChart(); chart_sw_live.type = "col"; chart_sw_live.style = 2; chart_sw_live.grouping = "clustered"; chart_sw_live.title = "[🔴Live 现存] 单域名外链比 (预警)"
            chart_sw_live.width = 23; chart_sw_live.height = 12
            chart_sw_live.add_data(Reference(ws4, min_col=4, min_row=live_sw_start_row, max_col=4, max_row=live_sw_end_row), titles_from_data=True)
            chart_sw_live.set_categories(Reference(ws4, min_col=1, min_row=live_sw_start_row+1, max_col=1, max_row=live_sw_end_row))

            for i, s in enumerate(chart_cum_live.series):
                try: 
                    s.graphicalProperties.line.solidFill = sw_colors[i % len(sw_colors)]
                    s.graphicalProperties.line.width = 30000
                except: 
                    pass
            for chart in [chart_health_live, chart_dr_live, chart_sw_live]:
                for i, s in enumerate(chart.series):
                    try: 
                        s.graphicalProperties.solidFill = sw_colors[i % len(sw_colors)]
                    except: 
                        pass

            ws4.add_chart(chart_cum_live, "A46")
            ws4.add_chart(chart_health_live, "M46")
            ws4.add_chart(chart_dr_live, "A68")
            ws4.add_chart(chart_sw_live, "M68")

        # ==========================================
        # --- Sheet 5: 流量对比 (大图置顶 + 高端色系 + 完美布局) ---
        # ==========================================
        print("[进程] 正在扫描并生成【流量趋势对比】(Sheet 5)...")
        overview_files = glob.glob(os.path.join(DATA_DIR, "*overview-competitors*.csv"))
        
        if overview_files:
            target_csv = overview_files[0]
            ws5 = wb.create_sheet("流量趋势横向对比")
            ws5.sheet_view.showGridLines = False
            
            with open(target_csv, 'r', encoding='utf-8-sig') as f:
                csv_reader = list(csv.reader(f, delimiter='\t' if target_csv.endswith('.tsv') else ','))
            
            if len(csv_reader) > 5:
                domain_row = csv_reader[0]
                metric_row = csv_reader[1]
                
                traffic_cols = {} 
                for idx, (dom, metric) in enumerate(zip(domain_row, metric_row)):
                    if dom and "subdomains" in dom and metric.strip() == "Organic traffic":
                        clean_dom = dom.replace("/ (subdomains)", "").replace(" (subdomains)", "").strip()
                        if clean_dom not in traffic_cols:
                            traffic_cols[clean_dom] = idx
                
                domains = list(traffic_cols.keys())
                
                # 👑 创意 1：添加高端大标题与深色顶栏
                ws5.append(["竞品大盘自然流量深度追踪 (Organic Traffic)"])
                ws5.cell(row=1, column=1).font = Font(bold=True, size=15, color="FFFFFF")
                for col in range(1, len(domains) + 2):
                    ws5.cell(row=1, column=col).fill = PatternFill("solid", fgColor="1F4E78")
                
                # 👑 创意 2：留出极其阔绰的空间给图表，彻底告别遮挡 (预留 35 行)
                for _ in range(35):
                    ws5.append([])
                
                # 第 37 行开始写数据表头
                ws5.append(["Date"] + domains)
                header_row_ws5 = ws5.max_row
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="294266", end_color="294266", fill_type="solid")
                for col_num, cell in enumerate(ws5[header_row_ws5], 1):
                    cell.font = header_font; cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    # 放宽列距，让数字有呼吸感
                    ws5.column_dimensions[get_column_letter(col_num)].width = 22 
                
                date_data_start = 4 
                for row in csv_reader[date_data_start:]:
                    if not row or not row[0]: continue
                    row_data = [row[0]]
                    for dom in domains:
                        col_idx = traffic_cols.get(dom)
                        try: val = int(row[col_idx])
                        except: val = 0
                        row_data.append(val)
                    ws5.append(row_data)
                    
                    # 👑 创意 3：数据千分位格式化，告别密密麻麻的 0，提升专业度
                    current_row = ws5.max_row
                    for c_idx in range(2, len(domains) + 2):
                        ws5.cell(row=current_row, column=c_idx).number_format = '#,##0'
                    
                # 渲染图表
                max_row_ws5 = ws5.max_row
                chart_traffic = LineChart()
                chart_traffic.title = "全盘竞品自然流量走势"
                chart_traffic.style = 2 
                chart_traffic.y_axis.title = "预估自然流量"
                chart_traffic.x_axis.title = "时间节点"
                
                # 👑 创意 4：超宽屏影院级比例，霸气侧漏
                chart_traffic.width = 35; chart_traffic.height = 17.5 
                
                data_traffic = Reference(ws5, min_col=2, min_row=header_row_ws5, max_col=len(domains)+1, max_row=max_row_ws5)
                cats_traffic = Reference(ws5, min_col=1, min_row=header_row_ws5+1, max_col=1, max_row=max_row_ws5)
                
                chart_traffic.add_data(data_traffic, titles_from_data=True)
                chart_traffic.set_categories(cats_traffic)
                
                sw_colors = ["294266", "F7941D", "20B799", "FDB913", "3FBBDF", "8C67AB", "E5625E", "8391A5"]
                for i, s in enumerate(chart_traffic.series):
                    try:
                        s.graphicalProperties.line.solidFill = sw_colors[i % len(sw_colors)]
                        s.graphicalProperties.line.width = 30000 
                        # 👑 创意 5：开启平滑曲线，线条极其丝滑高级
                        s.smooth = True 
                    except: pass
                
                # 图表定海神针：放置在 B3，绝不会遮挡下方 37 行开始的数据表
                ws5.add_chart(chart_traffic, "B3")
                
                # 👑 创意 6：冻结表头！无论你怎么往下滚动看明细，表头永远吸附在最上方！
                ws5.freeze_panes = ws5[f'A{header_row_ws5 + 1}']
                
                print("[✓] 流量趋势数据已成功导入并生成置顶趋势大图。")


        wb.save(excel_filename)
        print(f"\n[✓] 成功生成【独立域名级 + 隐藏外链穿透】战略报表：{excel_filename}")
        print("    -> 利用 Links in group 字段不耗费额度还原了【总外链量】")
        print("    -> 新增【单域名外链比(Sitewide Index)】侦测虚假繁荣的外链作弊")
    else:
        pass

    json_filename = os.path.join(REPORT_DIR, f"Agent_Payload_{date_str}.json")
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump({"data": json_payloads}, f, ensure_ascii=False, indent=4)

def run_matrix_engine():
    print("=== 初始化引荐域名穿透 (Domain-Level) 矩阵引擎 ===")
    conn = init_db()
    total_files, global_headers = process_ahrefs_exports(conn)
    if total_files > 0:
        generate_strategic_reports(conn, total_files, global_headers)
    conn.close()

if __name__ == "__main__":
    run_matrix_engine()
