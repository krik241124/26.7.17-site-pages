# page-read.py
# Page Intelligence System (PIS/GIS Module)
import os
import sqlite3
import csv
import json
import datetime
import glob
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "page_intelligence.db") # 独立的内页数据库
DATA_DIR = os.path.join(WORKSPACE_DIR, "ahrefs_data")       # 存放 Top Pages 数据的专属目录
REPORT_DIR = os.path.join(WORKSPACE_DIR, "reports")

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS top_pages")
    # 创建适合 Top Pages 的数据表
    cur.execute('''
        CREATE TABLE top_pages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competitor_domain TEXT,
            page_url TEXT,
            url_path TEXT,
            page_category TEXT,
            page_type TEXT,
            ur INTEGER,
            current_traffic INTEGER,
            current_traffic_value REAL,
            current_ref_domains INTEGER,
            current_keywords INTEGER,
            top_keyword TEXT,
            top_keyword_volume INTEGER,
            top_keyword_position INTEGER,
            raw_data TEXT,
            UNIQUE(competitor_domain, page_url)
        )
    ''')
    cur.execute("CREATE INDEX idx_competitor ON top_pages(competitor_domain)")
    conn.commit()
    return conn

def extract_competitor_from_filename(filename):
    # 例如：appscenic-top-pages.csv -> appscenic
    name = os.path.basename(filename).lower()
    if '-top-pages' in name: return name.split('-top-pages')[0]
    if '_top_pages' in name: return name.split('_top_pages')[0]
    return name.split('.')[0]

def infer_page_category(url):
    """通过 URL 路径智能推测页面所属的主类别 (GIS 策略分析用)"""
    try:
        path = urlparse(url).path.strip('/').lower()
        if not path: return "Homepage" # 首页
        parts = path.split('/')
        first_dir = parts[0]
        
        # 常见分类字典 (你可以根据行业自行丰富)
        if first_dir in ['blog', 'article', 'news', 'guides', 'learn']: return "Blog & Content"
        if first_dir in ['features', 'product', 'tour']: return "Product Features"
        if first_dir in ['tools', 'calculators', 'generator']: return "Free Tools (Engine)"
        if first_dir in ['pricing', 'plans']: return "Pricing"
        if first_dir in ['glossary', 'dictionary', 'wiki']: return "Glossary (Programmatic)"
        if first_dir in ['integrations', 'apps', 'plugins']: return "Integrations Ecosystem"
        if first_dir in ['vs', 'compare', 'alternatives']: return "Comparison Pages"
        
        return f"Other (/{first_dir}/)"
    except Exception:
        return "Unknown"

def process_top_pages_exports(conn):
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
        print(f"[提示] 请将 Ahrefs 的 Top Pages 数据放入 {DATA_DIR} 目录后重新运行。")
        return 0, []

    files = glob.glob(os.path.join(DATA_DIR, "*.*sv"))
    if not files:
        print(f"[提示] 在 {DATA_DIR} 中未找到 csv/tsv 文件。")
        return 0, []

    cur = conn.cursor()
    total_files = len(files)
    global_headers = [] 
    
    for file_path in files:
        competitor = extract_competitor_from_filename(file_path)
        print(f"正在清洗并提取竞品 [{competitor}] 的核心内页(Top Pages)数据...")
        
        try:
            f = open(file_path, 'r', encoding='utf-16')
            f.read(1); f.seek(0)
        except UnicodeError:
            f = open(file_path, 'r', encoding='utf-8-sig')

        reader = csv.DictReader(f, delimiter='\t' if file_path.endswith('.tsv') or 'utf-16' in str(f.encoding) else ',')
        
        # 🛡️ 新增：防御性代码，如果遇到空文件或无表头文件，自动跳过，不让程序崩溃
        if reader.fieldnames is None:
            print(f"  -> [警告] 竞品 {competitor} 的文件为空或无有效表头，已自动跳过。")
            f.close()
            continue
            
        cleaned_fieldnames = [str(fn).replace('\ufeff', '').strip() for fn in reader.fieldnames if fn]
        for fn in cleaned_fieldnames:
            if fn not in global_headers: global_headers.append(fn)
        
        insert_data = []
        for raw_row in reader:
            clean_row = {str(k).replace('\ufeff', '').strip(): v for k, v in raw_row.items() if k}
            lower_row = {k.lower(): v for k, v in clean_row.items()}
            
            url = lower_row.get('url', lower_row.get('page url', ''))
            if not url: continue
                
            try: ur = int(float(lower_row.get('ur', 0) or 0))
            except: ur = 0
            
            try: current_traffic = int(float(lower_row.get('current traffic', lower_row.get('traffic', 0)) or 0))
            except: current_traffic = 0
            
            try: current_traffic_value = float(lower_row.get('current traffic value', lower_row.get('traffic value', 0)) or 0)
            except: current_traffic_value = 0.0
            
            try: current_ref_domains = int(float(lower_row.get('current referring domains', lower_row.get('referring domains', 0)) or 0))
            except: current_ref_domains = 0
            
            try: current_keywords = int(float(lower_row.get('current # of keywords', lower_row.get('keywords', 0)) or 0))
            except: current_keywords = 0
            
            top_keyword = str(lower_row.get('current top keyword', lower_row.get('top keyword', ''))).strip()
            try: top_keyword_volume = int(float(lower_row.get('current top keyword: volume', lower_row.get('volume', 0)) or 0))
            except: top_keyword_volume = 0
            
            try: top_keyword_position = int(float(lower_row.get('current top keyword: position', lower_row.get('position', 0)) or 0))
            except: top_keyword_position = 0
            
            page_type_ahrefs = str(lower_row.get('page type', '')).strip()
            page_category = infer_page_category(url)
            url_path = urlparse(url).path
            
            raw_json = json.dumps(clean_row, ensure_ascii=False)

            insert_data.append((
                competitor, url, url_path, page_category, page_type_ahrefs,
                ur, current_traffic, current_traffic_value, current_ref_domains,
                current_keywords, top_keyword, top_keyword_volume, top_keyword_position, raw_json
            ))
            
        f.close()
        
        cur.executemany('''
            INSERT OR IGNORE INTO top_pages 
            (competitor_domain, page_url, url_path, page_category, page_type, ur, current_traffic, current_traffic_value, current_ref_domains, current_keywords, top_keyword, top_keyword_volume, top_keyword_position, raw_data) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', insert_data)
        
    conn.commit()
    return total_files, global_headers

def generate_page_reports(conn, global_headers):
    if not os.path.exists(REPORT_DIR): os.makedirs(REPORT_DIR)

    cur = conn.cursor()
    print("\n[进程] 正在拉取内页情报并进行多维战略聚合...")
    
    # 提取所有页面
    cur.execute('''SELECT competitor_domain, page_url, page_category, current_traffic, current_traffic_value, current_ref_domains, current_keywords, top_keyword, top_keyword_volume, top_keyword_position, raw_data FROM top_pages ORDER BY current_traffic DESC''')
    all_pages = cur.fetchall()

    comp_stats = {}
    for row in all_pages:
        comp, url, cat, traf, traf_val, rd, kwds, top_kw, vol, pos, raw_j = row
        
        if comp not in comp_stats:
            comp_stats[comp] = {
                'total_pages': 0, 'total_traffic': 0, 'total_traffic_value': 0, 'total_rd': 0,
                'category_traffic': {}, 'category_pages': {}, 'top_pages_list': []
            }
            
        comp_stats[comp]['total_pages'] += 1
        comp_stats[comp]['total_traffic'] += traf
        comp_stats[comp]['total_traffic_value'] += traf_val
        comp_stats[comp]['total_rd'] += rd
        
        comp_stats[comp]['category_traffic'][cat] = comp_stats[comp]['category_traffic'].get(cat, 0) + traf
        comp_stats[comp]['category_pages'][cat] = comp_stats[comp]['category_pages'].get(cat, 0) + 1
        
        # 保存竞品的前50大流量页面用于展示
        if len(comp_stats[comp]['top_pages_list']) < 50 and traf > 0:
            comp_stats[comp]['top_pages_list'].append({
                'url': url, 'cat': cat, 'traf': traf, 'kwds': kwds, 'top_kw': top_kw, 'vol': vol, 'pos': pos
            })

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    excel_filename = os.path.join(REPORT_DIR, f"Page_Intelligence_Matrix_{date_str}.xlsx")
    
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        sw_colors = ["294266", "F7941D", "20B799", "FDB913", "3FBBDF", "8C67AB", "E5625E", "8391A5"]
        
        # ==========================================
        # Sheet 1: 全局优质页面内容机会池 (GIS Opportunity Queue)
        # ==========================================
        ws1 = wb.active
        ws1.title = "内容机会池 (Opportunity Queue)"
        ws1.sheet_view.showGridLines = False
        
        headers = ["优先级/流量", "竞品来源", "页面大类", "具体URL", "核心关键词 (Top Keyword)", "该词月搜量", "竞品当前排名", "该页总引荐域名", "该页包含词数", "预估带入流量", "流量预估价值 ($)"]
        ws1.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        for col_num, cell in enumerate(ws1[1], 1):
            cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        
        ws1.column_dimensions['D'].width = 60
        ws1.column_dimensions['E'].width = 30
        for c in ['A','B','C','F','G','H','I','J','K']: ws1.column_dimensions[c].width = 15
        
        # 写入前 500 个最高流量页面作为核心参考
        for idx, row in enumerate(all_pages[:500], start=2):
            comp, url, cat, traf, traf_val, rd, kwds, top_kw, vol, pos, raw_j = row
            ws1.append([traf, comp, cat, url, top_kw, vol, pos, rd, kwds, traf, traf_val])
            if idx % 2 == 0:
                for col in range(1, 12):
                    ws1.cell(row=idx, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # ==========================================
        # Sheet 2: 竞品流量结构与战略分析看板
        # ==========================================
        ws2 = wb.create_sheet("单竞品流量结构诊断")
        ws2.sheet_view.showGridLines = False
        
        ws2.column_dimensions['A'].width = 35
        for i in range(2, 10): ws2.column_dimensions[get_column_letter(i)].width = 18
        
        for comp, stats in comp_stats.items():
            if stats['total_traffic'] == 0: continue
            
            # 1. 抬头
            ws2.append([f"🚀 {comp.upper()} 页面流量结构与增长策略侦测"])
            header_row = ws2.max_row
            ws2.cell(row=header_row, column=1).font = Font(bold=True, size=15, color="FFFFFF")
            for col in range(1, 10): ws2.cell(row=header_row, column=col).fill = PatternFill("solid", fgColor="294266")
            ws2.append([])
            
            # 2. 宏观数据
            ws2.append(["核心数据", "有流量的页面总数", "全站月预估流量", "全站流量价值 ($)", "全站平均页面外链比 (RD/Page)"])
            m_title_row = ws2.max_row
            ws2.append(["", stats['total_pages'], stats['total_traffic'], round(stats['total_traffic_value'],2), round(stats['total_rd']/stats['total_pages'] if stats['total_pages'] else 0, 1)])
            m_data_row = ws2.max_row
            
            for c in range(1, 6):
                ws2.cell(row=m_title_row, column=c).font = Font(bold=True); ws2.cell(row=m_title_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
                ws2.cell(row=m_title_row, column=c).alignment = Alignment(horizontal="center"); ws2.cell(row=m_data_row, column=c).alignment = Alignment(horizontal="center")
            ws2.append([])

            # 3. 目录/类目流量拆解
            ws2.append(["页面结构类型 (Category)", "页面数量 (投入成本)", "获取流量总计 (产出)", "占全站流量比重", "篇均流量 (ROI评估)"])
            cat_title_row = ws2.max_row
            
            sorted_cats = sorted(stats['category_traffic'].items(), key=lambda x: x[1], reverse=True)
            for cat, traf in sorted_cats:
                pages = stats['category_pages'][cat]
                pct = traf / stats['total_traffic'] if stats['total_traffic'] else 0
                avg_traf = traf / pages if pages else 0
                ws2.append([cat, pages, traf, pct, round(avg_traf, 1)])
                ws2.cell(row=ws2.max_row, column=4).number_format = '0.00%'
                
            cat_end_row = ws2.max_row
            
            # 生成图表：页面类别流量饼图
            chart_pie = PieChart()
            chart_pie.title = f"[{comp}] - 流量结构护城河 (Traffic Share)"
            chart_pie.style = 26
            chart_pie.width = 30; chart_pie.height = 20
            labels = Reference(ws2, min_col=1, min_row=cat_title_row+1, max_col=1, max_row=cat_end_row)
            data = Reference(ws2, min_col=3, min_row=cat_title_row+1, max_col=3, max_row=cat_end_row)
            chart_pie.add_data(data, titles_from_data=False)
            chart_pie.set_categories(labels)
            chart_pie.dataLabels = DataLabelList()
            chart_pie.dataLabels.showPercent = True; chart_pie.dataLabels.showVal = False
            
            # 图表放置位置
            chart_anchor = f"G{m_title_row+2}"
            ws2.add_chart(chart_pie, chart_anchor)
            
            ws2.append([])
            
            # 4. Top 10 引擎级页面 (The Traffic Engines)
            ws2.append(["🔥 Top 10 核心流量引擎 (建议交给 PIS 直接复刻或超越)"])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=ws2.max_row, column=1).fill = PatternFill("solid", fgColor="E5625E")
            
            ws2.append(["URL", "类型", "流量", "核心词", "搜索量", "排名"])
            for t_idx, t_page in enumerate(stats['top_pages_list'][:10]):
                ws2.append([t_page['url'], t_page['cat'], t_page['traf'], t_page['top_kw'], t_page['vol'], t_page['pos']])
            
            # 预留空间供下一个竞品
            for _ in range(5): ws2.append([])

        # ==========================================
        # Sheet 3: 明细宽表
        # ==========================================
        ws3 = wb.create_sheet("内页底层全量明细")
        headers3 = ["具体竞品", "URL", "自定义分类", "Ahrefs原始类型"] + global_headers
        ws3.append(headers3)
        for col_num, cell in enumerate(ws3[1], 1):
            cell.font = header_font; cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
        for row in all_pages:
            comp, url, cat, traf, traf_val, rd, kwds, top_kw, vol, pos, raw_j = row
            raw_dict = json.loads(raw_j)
            row_data = [comp, url, cat, raw_dict.get('page type', '')]
            for header in global_headers:
                row_data.append(raw_dict.get(header, ""))
            ws3.append(row_data)

        wb.save(excel_filename)
        print(f"\n[✓] 成功生成【内页情报穿透】战略报表：{excel_filename}")
        print("    -> 已自动推导页面结构归属 (Blog / Feature / Tools 等)")
        print("    -> 已为你锁定竞品的【Top 10 核心流量引擎页面】供 PIS 使用")

def run_page_intelligence_engine():
    print("=== 初始化内页穿透 (Page-Level) 矩阵引擎 ===")
    conn = init_db()
    total_files, global_headers = process_top_pages_exports(conn)
    if total_files > 0:
        generate_page_reports(conn, global_headers)
    conn.close()

if __name__ == "__main__":
    run_page_intelligence_engine()