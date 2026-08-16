import os
import sqlite3
import csv
import json
import datetime
import glob
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("请先安装 openpyxl: pip install openpyxl")

# ==========================================
# ⚙️ 核心全局配置
# ==========================================
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "keyword_intelligence.db")
DATA_DIR = os.path.join(WORKSPACE_DIR, "ahrefs_data")
OUTPUT_DIR = os.path.join(WORKSPACE_DIR, "reports")

# ==========================================
# 🧠 Arkswift 智能评分与路由引擎 (核心逻辑)
# ==========================================
def calculate_arkswift_opportunity_score(vol, kd, cpc, intent_commercial, intent_transactional):
    """
    真正适合 Arkswift (0-1新站) 的 KGR 模型：
    核心逻辑：绝对的一票否决制。KD > 35 的词直接不及格（0-1阶段没法做）。
    优先奖励 KD 在 0-15 的蓝海词。
    """
    if vol < 30: return 0 # 搜索量太小的长尾词不要
    
    # 1. 核心权重：KD 难度分层 (0-1 新站生命线)
    if kd > 40:
        base_score = 15   # 纯纯绞肉机，直接放弃
    elif kd > 25:
        base_score = 45   # 很难，随缘
    elif kd > 15:
        base_score = 65   # 可以竞争
    elif kd >= 5:
        base_score = 85   # 绝佳机会
    else:
        base_score = 95   # KD 0-4，必写金矿！
    
    # 2. 搜索量加分 (在 KD 合格的前提下，搜索量越大越好)
    if base_score >= 60:
        if vol >= 5000: base_score += 15
        elif vol >= 1000: base_score += 10
        elif vol >= 300: base_score += 5
        
    # 3. 商业意图与 CPC 附加值 (值钱的词加分)
    if intent_commercial or intent_transactional:
        base_score += 10
    if cpc >= 1.0:
        base_score += 10
    elif cpc >= 0.5:
        base_score += 5
        
    # 封顶 100，并惩罚高难度大词
    return min(100, round(base_score, 1))

def determine_page_strategy(keyword, intent_commercial, intent_transactional, intent_navigational):
    """
    通过意图推断我们应该建什么类型的页面，以及指派给哪个部门
    """
    kw_lower = keyword.lower()
    
    if intent_navigational or "login" in kw_lower or "app" in kw_lower:
        return "平台集成页 (Integrations)", "技术+内容", "必须强调 API 对接、一键抓取订单等技术硬实力，底本使用 SSR 渲染。"
    
    elif intent_commercial or intent_transactional or "dropshipping suppliers" in kw_lower:
        return "垂直品类聚合页 (Niche Hubs)", "技术+内容", "必须采用 SSR (服务端渲染) 在列表底部增加 800 字类目描述与 FAQ。中间调用 Arkswift 前 10 名热销 SKU，附带价格与利润率对比。"
    
    elif "calculator" in kw_lower or "converter" in kw_lower or "generator" in kw_lower:
        return "免费工具页 (Free Tools)", "技术部", "前端开发 JS 轻量级小工具单页，用于吸引自然外链 (Backlinks)。页面保持极简，侧边加注册按钮。"
        
    else:
        return "痛点答疑博客 (Educational Blog)", "内容部", "撰写 1500 字以上深度长文，分点阐述。页面滚动至 40% 处触发悬浮注册 CTA，文章需埋设 2-3 个内链。"

def generate_url_slug(keyword):
    """将关键词转换为标准的 URL slug"""
    slug = re.sub(r'[^a-z0-9\s]', '', keyword.lower())
    slug = re.sub(r'\s+', '-', slug.strip())
    return slug

# ==========================================
# 📊 第一阶段：清洗与入库
# ==========================================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS onpage_keywords")
    cur.execute('''
        CREATE TABLE onpage_keywords (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competitor_domain TEXT,
            keyword TEXT,
            is_informational INTEGER,
            is_commercial INTEGER,
            is_transactional INTEGER,
            is_navigational INTEGER,
            volume INTEGER,
            kd INTEGER,
            cpc REAL,
            current_traffic REAL,
            url TEXT,
            UNIQUE(competitor_domain, keyword)
        )
    ''')
    conn.commit()
    return conn

def parse_ahrefs_data(conn):
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
        print(f"[提示] 请将 Ahrefs 数据放入 {DATA_DIR}。")
        return False

    files = glob.glob(os.path.join(DATA_DIR, "*.csv")) + glob.glob(os.path.join(DATA_DIR, "*.tsv"))
    if not files:
        print(f"[提示] 未在 {DATA_DIR} 找到文件。")
        return False

    cur = conn.cursor()
    
    for file_path in files:
        filename = os.path.basename(file_path).lower()
        competitor = filename.split('-organic-positions')[0]
        if "appscenic" in filename: competitor = "appscenic"
        elif "cjdropshipping" in filename: competitor = "cjdropshipping"
        elif "doba" in filename: competitor = "doba"
        elif "syncee" in filename: competitor = "syncee"
        
        print(f"[数据入库] 正在解析竞品: {competitor.upper()} ...")
        
        try:
            f = open(file_path, 'r', encoding='utf-16')
            f.read(1); f.seek(0)
        except UnicodeError:
            f = open(file_path, 'r', encoding='utf-8-sig')

        delimiter = '\t' if file_path.endswith('.tsv') or 'utf-16' in str(f.encoding) else ','
        reader = csv.DictReader(f, delimiter=delimiter)

        insert_data = []
        for row in reader:
            lower_row = {str(k).strip().lower(): v for k, v in row.items() if k}
            
            kw = lower_row.get('keyword', '').strip()
            if not kw: continue
            
            # 过滤掉带有竞品品牌词的无用数据
            if competitor in kw.replace(" ", ""):
                continue

            try: vol = int(float(lower_row.get('volume', 0) or 0))
            except: vol = 0
            try: kd = int(float(lower_row.get('kd', 0) or 0))
            except: kd = 0
            try: cpc = float(lower_row.get('cpc', 0) or 0)
            except: cpc = 0.0
            
            # 兼容多个版本的表头
            traf_str = lower_row.get('current organic traffic') or lower_row.get('traffic') or '0'
            try: traf = float(traf_str)
            except: traf = 0.0
            
            url = lower_row.get('url', '')
            
            is_info = 1 if str(lower_row.get('informational', 'FALSE')).upper() == 'TRUE' else 0
            is_comm = 1 if str(lower_row.get('commercial', 'FALSE')).upper() == 'TRUE' else 0
            is_trans = 1 if str(lower_row.get('transactional', 'FALSE')).upper() == 'TRUE' else 0
            is_nav = 1 if str(lower_row.get('navigational', 'FALSE')).upper() == 'TRUE' else 0

            insert_data.append((
                competitor, kw, is_info, is_comm, is_trans, is_nav, 
                vol, kd, cpc, traf, url
            ))
            
        cur.executemany('''
            INSERT OR IGNORE INTO onpage_keywords 
            (competitor_domain, keyword, is_informational, is_commercial, is_transactional, is_navigational, volume, kd, cpc, current_traffic, url) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', insert_data)
        f.close()
    
    conn.commit()
    return True

# ==========================================
# 🚀 第二阶段：合并竞品，生成飞书多维表格结构
# ==========================================
def generate_feishu_crm(conn):
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    cur = conn.cursor()
    # 直接读取上一个脚本清洗好的数据，剔除品牌词
    cur.execute('''
        SELECT keyword, competitor_domain, is_informational, is_commercial, is_transactional, is_navigational, volume, kd, cpc, current_traffic, url 
        FROM keywords
        WHERE is_branded = 0
    ''')
    rows = cur.fetchall()

# 1. 关键词聚合 (Alpha 重合度计算 & URL 全球流量合并去重)
    keyword_map = {}
    for r in rows:
        kw, comp, is_info, is_comm, is_trans, is_nav, vol, kd, cpc, traf, url = r
        
        if kw not in keyword_map:
            keyword_map[kw] = {
                'comps': set(),
                'benchmarks_dict': {}, # 改用字典，按 URL 去重
                'vol': 0, 'kd': kd, 'cpc': cpc, 
                'total_traf': 0.0,
                'is_comm': is_comm, 'is_trans': is_trans, 'is_nav': is_nav, 'is_info': is_info
            }
        
        k_node = keyword_map[kw]
        k_node['comps'].add(comp.capitalize())
        k_node['total_traf'] += traf
        
        # 提取各个国家记录中最大的 Volume 和对应的 KD (代表最大核心市场数据)
        if vol > k_node['vol']:
            k_node['vol'] = vol
            k_node['kd'] = kd
            k_node['cpc'] = cpc

        # URL 全球流量合并去重逻辑
        if traf > 0 and url:
            if url not in k_node['benchmarks_dict']:
                k_node['benchmarks_dict'][url] = {'comp': comp.capitalize(), 'traf': 0.0}
            k_node['benchmarks_dict'][url]['traf'] += traf # 累加该 URL 在不同国家的流量

    # 2. 算分与定级
    task_list = []
    task_counter = 1
    today_str = datetime.datetime.now().strftime("%Y-%m-%d")

    # 2. 算分与定级
    task_list = []
    task_counter = 1
    today_str = datetime.datetime.now().strftime("%Y-%m-%d")

    for kw, d in keyword_map.items():
        # KD 太高或没有流量潜力的，直接不生成工单，保持 CRM 干净
        score = calculate_arkswift_opportunity_score(d['vol'], d['kd'], d['cpc'], d['is_comm'], d['is_trans'])
        if score < 50: 
            continue 
        
        page_type, dept, action_sop = determine_page_strategy(kw, d['is_comm'], d['is_trans'], d['is_nav'])
        slug = generate_url_slug(kw)
        
        if "聚合" in page_type: target_url = f"/category/{slug}"
        elif "博客" in page_type: target_url = f"/blog/{slug}"
        elif "工具" in page_type: target_url = f"/tools/{slug}"
        else: target_url = f"/integrations/{slug}"
        
        title_placeholder = f"{kw.title()} (2026 Guide & Top Suppliers) | Arkswift"
        
        # 处理去重后的 Benchmark 列表 (按流量降序排序)
        sorted_benchmarks = sorted(d['benchmarks_dict'].items(), key=lambda x: x[1]['traf'], reverse=True)
        benchmark_strs = [f"[{data['comp']}] {b_url} (Traffic: {round(data['traf'], 1)})" for b_url, data in sorted_benchmarks]
        
        task_list.append({
            'score': score,
            'task_id': f"SEO-PAGE-{str(task_counter).zfill(4)}",
            'kw': kw,
            'vol': d['vol'], 'kd': d['kd'], 'cpc': d['cpc'],
            'traf': d['total_traf'],
            'comps': ", ".join(list(d['comps'])),
            'benchmarks': "\n".join(benchmark_strs), # 有多少放多少，已经去重过了
            'page_type': page_type,
            'dept': dept,
            'sop': action_sop,
            'target_url': target_url,
            'title': title_placeholder,
            'discovered_date': today_str
        })
        task_counter += 1

    # 按机器分降序，相同分数按能抢到的竞品流量降序
    task_list.sort(key=lambda x: (x['score'], x['traf']), reverse=True)

    # 3. 导出到 Excel (精美样式)
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Arkswift内页SEO执行排期表"

        headers = [
            # Zone 1 & 2: 核心情报 (Blue/Gray)
            ("A", "核心锁定词 (Target Keyword)", "4472C4"),
            ("B", "Task ID", "A6A6A6"),
            ("C", "页面类型 (Page Type)", "A6A6A6"),
            ("D", "协作部门 (Dept)", "A6A6A6"),
            ("E", "🤖 蓝海机器分 (KGR Score)", "4472C4"),
            ("F", "搜索量 (Volume)", "4472C4"),
            ("G", "难度 (KD)", "4472C4"),
            ("H", "CPC价值 ($)", "4472C4"),
            ("I", "截获竞品总流量 (Est. Traffic)", "4472C4"),
            ("J", "目前布局该词的竞品", "4472C4"),
            
            # Zone 3: 落地执行指令 (Orange)
            ("K", "目标 URL 规范", "ED7D31"),
            ("L", "建议 Title 标签", "ED7D31"),
            ("M", "技术与内容 SOP 指令", "ED7D31"),
            ("N", "🥇 高流量对标案例 (Benchmarks)", "ED7D31"),
            
            # Zone 4: 生命周期与状态追踪 (Green/Yellow)
            ("O", "挖掘发现日期", "70AD47"),
            ("P", "了结日期 (Completed)", "70AD47"),
            ("Q", "当前状态 (Status)", "FFC000") # 放到最后，且高亮
        ]

        ws.append([h[1] for h in headers])
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=h[2])
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, d in enumerate(task_list, start=2):
            # 按全新的 Header 顺序注入数据，最后一个字段固定为 "Pending"
            row = [
                d['kw'], d['task_id'], d['page_type'], d['dept'], 
                d['score'], d['vol'], d['kd'], d['cpc'], round(d['traf'], 1), d['comps'],
                d['target_url'], d['title'], d['sop'], d['benchmarks'],
                d['discovered_date'], "", "Pending"
            ]
            ws.append(row)
            
            # 直接使用 idx 指定行号，速度提升上千倍！
            ws.cell(row=idx, column=14).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=idx, column=13).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=idx, column=5).font = Font(bold=True, color="C00000") # 机器分数标红加粗

        # 调整列宽 (匹配新顺序)
        col_widths = {
            'A': 30, 'B': 15, 'C': 25, 'D': 15, 'E': 22, 
            'F': 12, 'G': 10, 'H': 12, 'I': 20, 'J': 25,
            'K': 35, 'L': 45, 'M': 50, 'N': 70,
            'O': 15, 'P': 15, 'Q': 20
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 冻结第一列(A列)，方便往右滑时始终能看到关键词
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        out_file = os.path.join(OUTPUT_DIR, f"Arkswift_OnPage_SEO_To-Do-List_{today_str}.xlsx")
        wb.save(out_file)
        
        print(f"\n[🚀 任务完成] 成功生成 Arkswift 站内 SEO 多维执行表单！")
        print(f"👉 文件保存至: {out_file}")
        print("💡 亮点说明：")
        print("1. 完美分离为四大区域（工单基础、数据情报、SOP落地、生命周期）。")
        print("2. 根据意图，自动分配【垂直类目 / 博客 / 工具】等建站路由，并分配给【技术部 / 内容部】。")
        print("3. 直接将该 Excel 拖入【飞书多维表格】创建数据表即可无缝对接！")
    else:
        print("未安装 openpyxl，无法生成报表。")

# ==========================================
# 主程序入口
# ==========================================
if __name__ == "__main__":
    print("=== 初始化 Arkswift 站内任务调度引擎 (复用 keyword_intelligence.db) ===")
    conn = sqlite3.connect(DB_FILE)
    generate_feishu_crm(conn)
    conn.close()