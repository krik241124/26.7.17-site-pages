# backlink_opportunity_engine.py
import os
import sqlite3
import datetime
import json
import re
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("请先安装 openpyxl: pip install openpyxl")

WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "backlink_intelligence.db")
OUTPUT_DIR = os.path.join(WORKSPACE_DIR, "reports")

# ==========================================
# 🚀 引入企业级分类器 (BacklinkClassifier)
# ==========================================
class BacklinkClassifier:
    DOMAIN_PATTERNS = {
        "App Marketplace": ["apps.shopify.com", "appstore.com", "apps.apple.com", "play.google.com", "wordpress.org/plugins", "woocommerce.com/products", "marketplace.visualstudio.com", "chrome.google.com/webstore", "addons.mozilla.org", "apps.microsoft.com", "zapier.com/apps", "make.com/apps", "slack.com/apps", "salesforce.com/appexchange", "hubspot.com/marketplace", "storeleads.app", "appstacked.io"],
        "Software Review Directory": ["g2.com", "capterra.com", "getapp.com", "softwareadvice.com", "trustradius.com", "saashub.com", "sourceforge.net", "producthunt.com", "alternativeto.net", "serchen.com", "financesonline.com"],
        "Ecommerce Resource": ["ecommerce-platforms.com", "dropshipping.com", "dodropshipping.com", "dsers.com", "hypersku.com", "minea.com", "alidrop.co", "zikanalytics.com", "leelinesourcing.com"],
        "Marketplace": ["amazon.com", "ebay.com", "etsy.com", "walmart.com", "aliexpress.com", "alibaba.com", "temu.com", "rakuten.com", "shopify.com", "bigcommerce.com"],
        "Social Media": ["youtube.com", "youtu.be", "tiktok.com", "instagram.com", "facebook.com", "pinterest.com", "reddit.com", "linkedin.com", "twitter.com", "x.com", "threads.net", "medium.com", "dev.to"],
        "Business Directory": ["yelp.com", "yellowpages.com", "bbb.org", "hotfrog.com", "foursquare.com", "manta.com", "kompass.com", "clutch.co", "goodfirms.co", "designrush.com"],
        "Coupon Deal": ["coupon", "coupons", "discount", "deal", "deals", "promo", "voucher", "cashback"],
        "Media": ["forbes.com", "entrepreneur.com", "techcrunch.com", "venturebeat.com", "wired.com", "theverge.com", "businessinsider.com"],
        "SEO Link Network": ["backlink", "directory", "linkfarm", "guestpost", "writeforus", "submitarticle", "articlehub"]
    }

    URL_PATTERNS = {
        "Guest Post": ["write-for-us", "writeforus", "guest-post", "guestpost", "contribute", "submit-article", "submit-post", "author-guidelines"],
        "Listicle": ["best-", "best/", "top-", "top10", "top-10", "recommended", "ranking", "alternatives"],
        "Comparison": ["vs", "-vs-", "compare", "comparison", "alternative", "competitor"],
        "Review": ["review", "reviews", "rating", "testimonial", "case-study"],
        "Resource": ["/resources/", "/resource/", "/library/", "/guide/", "/guides/", "/academy/", "/knowledge-base/", "/tools/"],
        "App Detail": ["/app/", "/apps/", "/software/", "/product/"],
        "Directory Listing": ["/directory/", "/listing/", "/vendors/", "/suppliers/", "/companies/"],
        "Forum Thread": ["/forum/", "/thread/", "/topic/", "/discussion/", "/question/"],
        "Blog Article": ["/blog/", "/article/", "/post/", "/news/", "/insights/"]
    }

    ANCHOR_PATTERNS = {
        "Commercial Intent": ["buy", "price", "pricing", "supplier", "vendor", "software"],
        "Comparison Intent": ["vs", "alternative", "compare", "comparison"],
        "Review Intent": ["review", "rating", "feedback"],
        "SEO Intent": ["guest post", "write for us", "submit"]
    }

    def normalize_domain(self, domain):
        domain = domain.lower()
        domain = domain.replace("www.", "")
        return domain

    def classify_domain(self, domain):
        domain = self.normalize_domain(domain)
        for category, patterns in self.DOMAIN_PATTERNS.items():
            for p in patterns:
                if "." in p:
                    if domain == p or domain.endswith("." + p): return category, 1.0
                else:
                    if p in domain: return category, 0.9
        return "Unknown Domain", 0.3

    def classify_url(self, urls):
        if isinstance(urls, str): urls = [urls]
        for url in urls:
            url_lower = url.lower()
            for category, patterns in self.URL_PATTERNS.items():
                for p in patterns:
                    if p in url_lower: return category, 0.8
        return "Unknown URL", 0.3

    def classify_anchor(self, anchors):
        if isinstance(anchors, str): anchors = [anchors]
        for anchor in anchors:
            anchor_lower = str(anchor).lower()
            for category, patterns in self.ANCHOR_PATTERNS.items():
                for p in patterns:
                    if p in anchor_lower: return category
        return "Unknown"

    def classify(self, domain, urls, anchors=[]):
        d_type, d_score = self.classify_domain(domain)
        u_type, u_score = self.classify_url(urls)
        a_type = self.classify_anchor(anchors)

        if d_score >= 0.9: return d_type, d_score
        elif u_score >= 0.8: return u_type, u_score
        else: return "General Blog", 0.5

# ==========================================
# 第一阶段：纯 Python 规则评分引擎 (Auto Scoring)
# ==========================================
def get_alpha_score(count):
    if count >= 7: return 100
    elif count >= 6: return 90
    elif count >= 5: return 80
    elif count >= 4: return 70
    elif count >= 3: return 55
    elif count >= 2: return 35
    else: return 15

def get_dr_score(dr):
    if dr >= 80: return 100
    elif dr >= 60: return 80
    elif dr >= 40: return 60
    elif dr >= 20: return 40
    else: return 20

def get_traffic_score(traffic):
    # 【核心优化】从无到有(0到1) 赋予极其夸张的分数拉升，证明页面被收录且存活
    if traffic == 0: return 0       # 死亡/未收录页面，得 0 分
    elif traffic < 50: return 70    # 只要有1点流量，起步就是70分！产生质变
    elif traffic < 500: return 80   # 有基础流量
    elif traffic < 5000: return 90  # 流量优秀
    else: return 100                # 绝对头部流量

# ==========================================
# 第三阶段：处理数据并输出
# ==========================================
def generate_execution_queue():
    if not os.path.exists(DB_FILE):
        print(f"[错误] 未找到数据库 {DB_FILE}")
        return

    if not os.path.exists(OUTPUT_DIR): os.makedirs(OUTPUT_DIR)

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    
    print("[进程] 正在从情报数据库中抽取数据...")
    cur.execute('''
        SELECT competitor_domain, ref_domain, ref_url, domain_rating, page_traffic, is_dofollow, is_spam, links_in_group, target_url, raw_data 
        FROM backlinks 
    ''')
    rows = cur.fetchall()
    
    print("[进程] 正在通过 BacklinkClassifier 与漏斗引擎进行深度打分...")
    domain_stats = {}
    
    for row in rows:
        comp, ref_d, ref_u, dr, traf, is_dof, is_spam, links_in_group, tgt_u, raw_data = row
        
        if ref_d not in domain_stats:
            domain_stats[ref_d] = {
                'competitors': set(), 'urls': [], 'anchors': [], 'max_dr': 0, 'max_traffic': 0,
                'is_dofollow': 0, 'is_spam': 0, 'total_links': 0,
                'comp_link_counts': {}, 'link_details': set()
            }
            
        d = domain_stats[ref_d]
        d['competitors'].add(comp)
        d['urls'].append(ref_u)
        d['max_dr'] = max(d['max_dr'], dr)
        d['max_traffic'] = max(d['max_traffic'], traf)
        d['total_links'] += links_in_group 
        d['comp_link_counts'][comp] = d['comp_link_counts'].get(comp, 0) + links_in_group
        
        try:
            raw_dict = json.loads(raw_data)
            anchor = raw_dict.get('Link anchor') or raw_dict.get('Anchor') or raw_dict.get('anchor') or "无锚文本"
        except:
            anchor = "未知"
            
        d['anchors'].append(anchor)
        detail_str = f"[{comp}] {ref_u} ---> {tgt_u} (锚文本: {anchor})"
        d['link_details'].add(detail_str)
        if is_dof == 1: d['is_dofollow'] = 1
        if is_spam == 1: d['is_spam'] = 1

    scored_domains = []
    classifier = BacklinkClassifier()

    for ref_d, d in domain_stats.items():
        alpha_count = len(d['competitors'])
        
        # 1. 基础算分：Traffic 的权重发挥了巨大作用
        score_alpha = get_alpha_score(alpha_count)
        score_dr = get_dr_score(d['max_dr'])
        score_traffic = get_traffic_score(d['max_traffic'])
        
        base_score = (score_alpha * 0.15) + (score_dr * 0.40) + (score_traffic * 0.25)
        dofollow_bonus = 20 if d['is_dofollow'] == 1 else 0
        
        # 给“哪怕有一点自然流量”的网站，再增加一个硬核奖励分，与死站彻底拉开差距
        traffic_alive_bonus = 10 if d['max_traffic'] > 0 else 0
        
        auto_score = round(base_score + dofollow_bonus + traffic_alive_bonus, 1)
        if auto_score > 100: auto_score = 100.0 # 封顶 100
        
        if d['is_spam'] == 1: auto_score = -100
        
        # 2. 调用企业级分类器
        final_type, confidence = classifier.classify(ref_d, d['urls'], d['anchors'])
        
        # 3. 核心分层逻辑 (A/B/C) 集成在同一张表
        # -----------------------------
        # 全新漏斗：基于【获取难度 x 价值】的核心分流逻辑 (完全抛弃 Alpha 作为判断条件)
        # -----------------------------
        BATCHABLE_TYPES = ["Social Media", "Business Directory", "Coupon Deal", "SEO Link Network", "Directory Listing", "Forum Thread"]
        
        # 1. 垃圾桶 (Spam 或者 纯死站)
        is_trash = d['is_spam'] == 1 or (d['max_dr'] < 10 and d['max_traffic'] == 0)
        
        # 2. C类 (明确可以批量的水链，或者 价值偏低的长尾站)
        is_c_tier = final_type in BATCHABLE_TYPES or d['max_dr'] < 40

        if is_trash:
            tier = "🗑️ SPAM/废弃"
            tier_rank = 4
            action_suggest = "无价值死站/Spam，直接放弃"
            
        elif is_c_tier:
            tier = "🔵 C类 (全自动批量)"
            tier_rank = 3
            if final_type not in BATCHABLE_TYPES:
                # 哪怕是普通博客，因为 DR < 40，也被强行降级到 C类
                final_type += " (低权重降级)"
                action_suggest = "【全自动/VA】长尾小站，尝试用工具群发评论或低成本触达"
            else:
                action_suggest = "【全自动/VA】天生批量型站点，交由兼职按SOP批量填表发水帖"
                
        else:
            # 走到这里的，全都是 DR >= 40 且非黄页/论坛的优质站点！
            # 3. A类 与 B类的精细划分 (看获取门槛)
            if final_type in ["App Marketplace", "Ecommerce Resource", "Software Review Directory", "Media"]:
                tier = "🔴 A类 (核心狙击)"
                tier_rank = 1
                if final_type == "App Marketplace": action_suggest = "【产品部】开发相关插件并上架该生态"
                elif final_type == "Media": action_suggest = "【公关部】撰写高质量新闻稿寻求官方报道"
                else: action_suggest = "【商务部】联系官方入驻、谈深度商务合作或请求测评"
            else:
                tier = "🟡 B类 (半自动)"
                tier_rank = 2
                action_suggest = "【外展运营】人工审阅价值，邮件套用模板群发请求客座投稿或资源位"

        scored_domains.append({
            'tier_rank': tier_rank,
            'tier': tier,
            'domain': ref_d,
            'auto_score': auto_score,
            'link_type': f"{final_type} ({int(confidence*100)}%)",
            'action': action_suggest,
            'alpha': alpha_count,
            'total_links': f"总计:{d['total_links']} (" + ", ".join([f"{k}:{v}" for k, v in d['comp_link_counts'].items()]) + ")",
            'dr': d['max_dr'],
            'traffic': d['max_traffic'],
            'dofollow': "✅ Yes" if d['is_dofollow'] else "❌ No",
            'spam': "⚠️ Yes" if d['is_spam'] else "No",
            'example_url': "\n\n".join(list(d['link_details'])) 
        })

    # 【排序核心】首先按 A->B->C->Spam 阶级排序，阶级内按 机器分(Auto Score) 降序，再按 DR 降序
    scored_domains.sort(key=lambda x: (x['tier_rank'], -x['auto_score'], -x['dr']), reverse=False)

    # ==========================================
    # 第三阶段：输出带 AI/人工 交接带的 CRM Excel (单表保留原格式)
    # ==========================================
    print("[进程] 正在生成合并版 CRM 报表...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "外链获取执行队列 (CRM)"

    headers = [
        # --- Python 自动生成区 (灰底) ---
        ("A", "分层与优先级 (Tier)", "A6A6A6"),
        ("B", "目标域名 (Domain)", "A6A6A6"),
        ("C", "🤖机器评分 (Auto Score)", "A6A6A6"),
        ("D", "内容类型 (Link Type)", "A6A6A6"),
        ("E", "建议策略 (Action Suggested)", "A6A6A6"),
        ("F", "竞品重合度 (Alpha)", "A6A6A6"),
        ("G", "🔗该域名下总外链 (Total Links)", "A6A6A6"), 
        ("H", "DR权重", "A6A6A6"),
        ("I", "页面流量", "A6A6A6"),
        ("J", "Dofollow", "A6A6A6"),
        ("K", "Spam", "A6A6A6"),
        ("L", "对标链接样例", "A6A6A6"),
        
        # --- AI 预留处理区 (绿底) ---
        ("M", "🧠AI: 业务相关性评分(0-100)", "70AD47"),
        ("N", "🧠AI: 配套内容(如Pitch邮件/大纲)", "70AD47"),
        ("O", "🧠AI: 难度评级(Easy/Hard)", "70AD47"),
        
        # --- 人工执行区 (蓝底) ---
        ("P", "🧑‍💻联系邮箱 (Email)", "4472C4"),
        ("Q", "🧑‍💻获取入口URL", "4472C4"),
        ("R", "🏆总分 (Total Score) [自动计算]", "FFC000"),  
        ("S", "🧑‍💻人工: 审阅状态", "4472C4"),
        ("T", "🧑‍💻当前状态 (Status)", "4472C4")
    ]

    ws.append([h[1] for h in headers])
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h[2])
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, d in enumerate(scored_domains, start=2):
        row_data = [
            d['tier'], d['domain'], d['auto_score'], d['link_type'],
            d['action'], d['alpha'], d['total_links'], d['dr'],
            d['traffic'], d['dofollow'], d['spam'], d['example_url'],
            "", "", "", "", "", "", "Pending", "未开始 (Not Started)" 
        ]
        ws.append(row_data)
        
        ws.cell(row=idx, column=7).alignment = Alignment(wrap_text=False, vertical="center")
        ws.cell(row=idx, column=12).alignment = Alignment(wrap_text=False, vertical="center")
        
        # 依旧保留你的公式 R列(第18列)：85% Auto + 15% AI
        formula = f"=(C{idx}*0.85) + (IF(ISNUMBER(M{idx}),M{idx},0)*0.15)"
        ws.cell(row=idx, column=18, value=formula).font = Font(bold=True, color="C00000")
        
    column_widths = {
        'A': 20, 'B': 25, 'C': 18, 'D': 35, 'E': 45, 
        'F': 15, 'G': 40, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 90,
        'M': 25, 'N': 35, 'O': 20, 
        'P': 25, 'Q': 30, 'R': 25, 'S': 20, 'T': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    out_file = os.path.join(OUTPUT_DIR, f"Backlink_Execution_Queue_AllInOne_{date_str}.xlsx")
    wb.save(out_file)
    
    print(f"\n[✓] 成功生成外链执行队列 (CRM级): {out_file}")
    print("--------------------------------------------------")
    print("✅ 改进亮点已生效：")
    print("1. 【流量质变】：自然流量>0的页面获得了暴击加分，机器分大幅度提升！")
    print("2. 【企业级指纹】：完全采用 BacklinkClassifier，附带匹配置信度。")
    print("3. 【单表输出】：全部保留在单一 Sheet，并按 A/B/C 层级 > 机器评分降序 完美排序。")
    print("--------------------------------------------------")

if __name__ == "__main__":
    generate_execution_queue()