# bench_read.py
import os
import sqlite3
import gzip
import datetime
import xml.etree.ElementTree as ET
from urllib.parse import urlparse, urlunparse
import re
import json
import glob
import urllib.request
import urllib.error
import concurrent.futures

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "sitemap_monitor.db")
TARGETS_FILE = os.path.join(WORKSPACE_DIR, "targets.txt") 
REQUEST_TIMEOUT = 15 
USER_AGENT = "Mozilla/5.0 (compatible; bench-read-agent/1.0; +https://openclaw.ai/)"
MAX_WORKERS = 20 

def get_current_time_str():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def log_info(msg):
    print(f"[{get_current_time_str()}] {msg}")

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS sites (id INTEGER PRIMARY KEY AUTOINCREMENT, url TEXT UNIQUE)")
    cur.execute("CREATE TABLE IF NOT EXISTS pages (site_id INTEGER, page_url TEXT, first_seen TEXT, last_seen TEXT, PRIMARY KEY (site_id, page_url))")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pages_url ON pages(page_url)")
    conn.commit()
    return conn

def normalize_to_sitemap(raw_url):
    u = raw_url.strip()
    if not u.startswith('http'): u = 'https://' + u
    parsed = urlparse(u)
    if parsed.path.endswith('.xml'): return u
    return urlunparse((parsed.scheme, parsed.netloc, '/sitemap.xml', '', '', ''))

def sync_targets(conn):
    if not os.path.exists(TARGETS_FILE): 
        log_info(f"未找到目标文件 {TARGETS_FILE}")
        return
    with open(TARGETS_FILE, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines()
    cur = conn.cursor()
    added = 0
    for line in lines:
        if not line.strip(): continue
        try: 
            cur.execute("INSERT INTO sites (url) VALUES (?)", (normalize_to_sitemap(line),))
            added += 1
        except sqlite3.IntegrityError: 
            pass 
    conn.commit()
    log_info(f"同步目标列表完成，新增 {added} 个站点。")

def fetch_url_bytes(url):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as response:
            return response.read()
    except Exception as e:
        log_info(f"请求失败 [{url}]: {str(e)}")
        return None 

def parse_sitemap_bytes(content):
    try: 
        root = ET.fromstring(content)
    except Exception:
        try: 
            root = ET.fromstring(content.decode('utf-8', errors='ignore'))
        except Exception: 
            return [], False
    locs = []
    for elem in root.iter():
        if elem.tag.lower().endswith('loc') and elem.text:
            locs.append(elem.text.strip())
    return locs, root.tag.lower().endswith('sitemapindex')

def fetch_sitemap(url, depth=0, visited=None):
    if visited is None: visited = set()
    if depth > 3 or url in visited: return []
    visited.add(url)
    
    content = fetch_url_bytes(url)
    if not content: return []
    
    if url.endswith('.gz'):
        try: 
            content = gzip.decompress(content)
        except Exception: 
            pass
            
    locs, is_index = parse_sitemap_bytes(content)
    
    if is_index:
        log_info(f"发现索引文件: {url} (含 {len(locs)} 个子 Sitemap)")
        urls = []
        for s in locs: 
            urls += fetch_sitemap(s, depth + 1, visited)
        return urls
    else:
        if depth > 0:
            log_info(f"解析子文件: {url} (提取 {len(locs)} 个 URL)")
        return locs

def normalize_url(u):
    try:
        p = urlparse(u)
        path = p.path or '/'
        while '//' in path: path = path.replace('//', '/')
        if path != '/' and path.endswith('/'): path = path[:-1]
        return urlunparse((p.scheme, p.netloc, path, '', '', ''))
    except Exception: return u

def extract_keywords_from_url(u):
    try:
        p = urlparse(u)
        path = p.path
        if path != '/' and path.endswith('/'): path = path[:-1]
        if not path or path == '/': return []
        slug = path.split('/')[-1]
        slug_without_ext = re.sub(r'\.(html|htm|php|aspx|asp)$', '', slug, flags=re.IGNORECASE)
        if not slug_without_ext: return []
        keyword_phrase = re.sub(r'[-_]+', ' ', slug_without_ext)
        keyword_phrase = re.sub(r'\s+', ' ', keyword_phrase).strip()
        
        if len(keyword_phrase) > 3 and not keyword_phrase.isdigit():
            return [keyword_phrase.lower()]
        return []
    except Exception: return []

def process_single_site(site_data):
    site_id, site_url = site_data
    log_info(f"开始抓取站点: {site_url}")
    urls = fetch_sitemap(site_url)
    current_urls = {normalize_url(u) for u in urls if u}
    log_info(f"站点抓取完毕: {site_url} (共获取 {len(current_urls)} 个不重复 URL)")
    return site_id, site_url, current_urls

def get_domain_name(url):
    try: return urlparse(url).netloc.replace('www.', '')
    except: return url

def generate_reports(conn, new_pages_today, removed_pages_today, current_time_str):
    log_info("开始计算关键词数据与 Alpha 值...")
    cur = conn.cursor()
    
    cur.execute("SELECT id, url FROM sites")
    site_map = {row[0]: get_domain_name(row[1]) for row in cur.fetchall()}

    global_kw_matrix = {}
    
    cur.execute("SELECT page_url, site_id FROM pages")
    rows = cur.fetchall()
    log_info(f"正在从本地数据库提取关键词，总计数据行数: {len(rows)}")
    
    for page_url, site_id in rows:
        domain = site_map.get(site_id, "Unknown")
        for kw in extract_keywords_from_url(page_url):
            if kw not in global_kw_matrix:
                global_kw_matrix[kw] = {}
            if domain not in global_kw_matrix[kw]:
                global_kw_matrix[kw][domain] = []
            global_kw_matrix[kw][domain].append(page_url)
            
    keyword_alpha_list = []
    for kw, domains_data in global_kw_matrix.items():
        alpha = len(domains_data.keys())
        shared_by = ", ".join(domains_data.keys())
        
        clickable_links = []
        for d, urls in domains_data.items():
            for u in urls[:2]: 
                clickable_links.append(f"[{d}] {u}")
                
        keyword_alpha_list.append({
            'keyword': kw,
            'alpha': alpha,
            'shared_by': shared_by,
            'links': "\n".join(clickable_links)
        })

    keyword_alpha_list.sort(key=lambda x: x['alpha'], reverse=True)

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    excel_filename = os.path.join(WORKSPACE_DIR, f"PSEO_Keyword_Matrix_{date_str}.xlsx")
    
    if HAS_OPENPYXL:
        log_info(f"正在生成 Excel 文件: {excel_filename}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "竞品PSEO关键词大盘"
        
        headers = ["PSEO提取核心词 (Seed Keyword)", "竞品重合度 (Alpha Score)", "涵盖竞品 (Shared By)", "具体对标内页地址"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
            
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 100
        
        for item in keyword_alpha_list:
            if item['alpha'] < 2 and len(site_map) >= 2:
                continue
                
            row_data = [item['keyword'], item['alpha'], item['shared_by'], item['links']]
            ws.append(row_data)
            ws.cell(row=ws.max_row, column=4).alignment = Alignment(wrap_text=True, vertical='center')
            
        wb.save(excel_filename)
        log_info("Excel 报表生成完毕。")

    output_json = {
        "event_type": "pseo_reverse_engineering_complete",
        "timestamp": current_time_str,
        "data_summary": {
            "sites_analyzed": len(site_map),
            "total_keywords_extracted": len(keyword_alpha_list)
        }
    }
    
    with open(os.path.join(WORKSPACE_DIR, f"log_{date_str}.json"), 'w') as f:
        json.dump(output_json, f)
    log_info("运行任务结束。")

def run_monitor():
    log_info("程序启动。")
    conn = init_db()
    sync_targets(conn)
    current_time_str = get_current_time_str()
    
    cur = conn.cursor()
    cur.execute("SELECT id, url FROM sites")
    sites = cur.fetchall()
    
    if not sites:
        log_info("未检测到有效目标站点，任务终止。")
        return

    new_pages_today = {}
    removed_pages_today = {}

    log_info(f"开始并发抓取 {len(sites)} 个站点的 Sitemap，最大线程数 {MAX_WORKERS}...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        results = list(executor.map(process_single_site, sites))

    log_info("数据抓取完成，开始比对本地数据库差异...")
    for site_id, site_url, current_urls in results:
        cur.execute("SELECT page_url FROM pages WHERE site_id=?", (site_id,))
        history_urls = {row[0] for row in cur.fetchall()}
        
        added = current_urls - history_urls
        removed = history_urls - current_urls
        
        if added:
            insert_list = [(site_id, u, current_time_str, current_time_str) for u in added]
            for u in added: new_pages_today.setdefault(u, []).append(site_url)
            cur.executemany("INSERT OR REPLACE INTO pages (site_id, page_url, first_seen, last_seen) VALUES (?, ?, ?, ?)", insert_list)

        if removed:
            delete_list = [(site_id, u) for u in removed]
            for u in removed: removed_pages_today.setdefault(u, []).append(site_url)
            cur.executemany("DELETE FROM pages WHERE site_id=? AND page_url=?", delete_list)

    conn.commit()
    log_info("数据库更新完成。")
    
    generate_reports(conn, new_pages_today, removed_pages_today, current_time_str)

if __name__ == "__main__":
    run_monitor()