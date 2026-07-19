# backlink_opportunity_engine.py
import os
import sqlite3
import datetime
import json   # <--- 新增这行
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("请先安装 openpyxl: pip install openpyxl")

WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(WORKSPACE_DIR, "backlink_intelligence.db")
OUTPUT_DIR = os.path.join(WORKSPACE_DIR, "reports")

# ==========================================
# 第一阶段：纯 Python 规则评分引擎 (Auto Scoring)
# ==========================================
def get_alpha_score(count):
    if count >= 7: return 100
    elif count >= 6: return 90
    elif count >= 5: return 80
    elif count >= 4: return 70
    elif count >= 3: return 55
    elif count >= 2: return 35
    else: return 15

def get_dr_score(dr):
    if dr >= 80: return 100
    elif dr >= 60: return 80
    elif dr >= 40: return 60
    elif dr >= 20: return 40
    else: return 20

def get_traffic_score(traffic):
    if traffic >= 100000: return 100
    elif traffic >= 50000: return 90
    elif traffic >= 10000: return 70
    elif traffic >= 1000: return 50
    else: return 20

# ==========================================
# 第二阶段：URL 意图自动识别 (Link Type - 企业级规则引擎)
# ==========================================
def guess_link_type(domain, urls):
    domain_lower = str(domain).lower()
    
    # ---------------------------------------------------------
    # 1. 强指纹库 (Domain Fingerprint) - 权重最高，一锤定音
    # ---------------------------------------------------------
    DOMAIN_PATTERNS = {
        "App Market (应用市场/生态插件)": [
            "apps.shopify.com", "appstore.com", "apps.apple.com", "play.google.com", 
            "wordpress.org", "wordpress.com/plugins", "woocommerce.com/products", 
            "marketplace.visualstudio.com", "chrome.google.com/webstore", "addons.mozilla.org", 
            "apps.microsoft.com", "zapier.com/apps", "make.com/apps", "slack.com/apps", 
            "salesforce.com/appexchange", "hubspot.com/products/marketplace", "canva.com/apps", 
            "figma.com/community", "atlassian.com/software/marketplace"
        ],
        "Social / Video (视频社媒)": [
            "youtube.com", "youtu.be", "vimeo.com", "dailymotion.com", "tiktok.com", 
            "instagram.com", "pinterest.com", "facebook.com", "twitter.com", "x.com", 
            "threads.net", "snapchat.com", "reddit.com", "linkedin.com", "tumblr.com", 
            "flickr.com", "behance.net", "dribbble.com"
        ],
        "Directory / Review (企业目录评价)": [
            "g2.com", "capterra.com", "trustpilot.com", "saashub.com", "sourceforge.net", 
            "producthunt.com", "getapp.com", "softwareadvice.com", "alternativeto.net", 
            "slashdot.org", "crunchbase.com", "owler.com", "clutch.co", "goodfirms.co", 
            "designrush.com", "serchen.com", "financesonline.com", "trustradius.com", 
            "storeleads.app", "builtwith.com", "wappalyzer.com"
        ],
        "Business Directory (商业黄页)": [
            "yelp.com", "yellowpages.com", "bbb.org", "hotfrog.com", "foursquare.com", 
            "manta.com", "brownbook.net", "kompass.com", "business.com", "merchantcircle.com", 
            "citysearch.com", "mapquest.com"
        ],
        "Forum / Community (论坛社区)": [
            "stackexchange.com", "stackoverflow.com", "medium.com", "dev.to", 
            "hashnode.com", "indiehackers.com", "warriorforum.com", "blackhatworld.com", 
            "digitalpoint.com", "sitepoint.com", "moz.com/community", "community.shopify.com"
        ],
        "Ecommerce Marketplace (电商平台)": [
            "amazon.com", "ebay.com", "etsy.com", "walmart.com", "aliexpress.com", 
            "alibaba.com", "temu.com", "rakuten.com", "shopify.com", "woocommerce.com", 
            "bigcommerce.com", "magento.com"
        ],
        "Coupon Deal Site (优惠券)": [
            "coupon", "coupons", "deal", "deals", "discount", "promo", "voucher", 
            "cashback", "retailmenot", "slickdeals", "couponfollow", "dealspotr"
        ],
        "News Media (新闻媒体)": [
            "forbes.com", "entrepreneur.com", "businessinsider.com", "techcrunch.com", 
            "venturebeat.com", "mashable.com", "theverge.com", "wired.com", "huffpost.com", "substack.com"
        ],
        "Education Resource (教育资源)": [
            "edu", "coursera.org", "udemy.com", "edx.org", "skillshare.com", "khanacademy.org", "codecademy.com"
        ],
        "Affiliate Review (联盟测评)": [
            "review", "reviews", "best", "top", "compare", "versus", "alternative", "guide", "buyer"
        ]
    }

    for category, patterns in DOMAIN_PATTERNS.items():
        for p in patterns:
            # 如果指纹包含 "." (比如 x.com, apps.shopify.com)，说明是具体域名
            if "." in p:
                # 必须完全相等 (wix.com != x.com) 或者是它的子域名 (www.x.com)
                if domain_lower == p or domain_lower.endswith('.' + p):
                    return category
            # 如果指纹不包含 "." (比如 coupon, review)，说明是关键词，允许包含匹配
            else:
                if p in domain_lower:
                    return category

    # ---------------------------------------------------------
    # 2. URL 路径意图 (URL Path Intent) - 权重第二，判断具体页面
    # ---------------------------------------------------------
    URL_PATTERNS = {
        "Guest Post": ["write-for-us", "writeforus", "guest-post", "guestpost", "guest-blog", "guestblog", "contribute", "become-author", "submit-article", "submit-post", "author-guidelines", "editorial-guidelines"],
        "Resource Page": ["/resources/", "/resource/", "/library/", "/tools/", "/downloads/", "/free-tools/", "/knowledge-base/", "/learning/", "/academy/", "/guide/", "/guides/", "/references/"],
        "Listicle (合集列表)": ["best-", "top-", "top10", "top-10", "top20", "top-20", "best-tools", "best-software", "best-platforms", "best-companies", "recommended", "alternatives"],
        "Comparison": ["-vs-", "vs", "compare", "comparison", "alternative", "alternatives", "competitor", "competitors", "difference-between"],
        "Partner / Integration": ["/partner", "/partners", "/integration", "/integrations", "/apps/", "/marketplace/", "/solutions/", "/technology-partners"],
        "Profile Page": ["/author/", "/user/", "/profile/", "/member/", "/account/", "/contributors/", "/team/"],
        "Forum Thread": ["/forum/", "/thread/", "/topic/", "/discussion/", "/question/", "/answers/", "/community/"],
        "Blog Article": ["/blog/", "/news/", "/article/", "/post/", "/insights/", "/stories/", "/updates/"],
        "Review Page": ["/review/", "/reviews/", "/rating/", "/testimonial/", "/customer-story/", "/case-study/"],
        "Coupon": ["/coupon/", "/coupons/", "/discount/", "/promo/", "/deal/", "/offers/"],
        "Directory Listing": ["/directory/", "/listing/", "/companies/", "/vendors/", "/suppliers/", "/software/", "/products/"]
    }

    for url in urls:
        url_lower = str(url).lower()
        for category, patterns in URL_PATTERNS.items():
            if any(p in url_lower for p in patterns):
                return category

    # ---------------------------------------------------------
    # 3. 兜底策略
    # ---------------------------------------------------------
    return "General / Blog (常规博客/其他)"

def generate_execution_queue():
    if not os.path.exists(DB_FILE):
        print(f"[错误] 未找到数据库 {DB_FILE}，请先运行数据抓取分析脚本。")
        return

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    
    print("[进程] 正在从情报数据库中抽取数据...")
    cur.execute('''
        SELECT competitor_domain, ref_domain, ref_url, domain_rating, page_traffic, is_dofollow, is_spam, links_in_group, target_url, raw_data 
        FROM backlinks 
    ''')
    rows = cur.fetchall()
    
    print("[进程] 正在通过 Python 规则引擎进行聚类和打分...")
    domain_stats = {}
    
    for row in rows:
        comp, ref_d, ref_u, dr, traf, is_dof, is_spam, links_in_group, tgt_u, raw_data = row  # <--- 多解包出 tgt_u 和 raw_data
        
        if ref_d not in domain_stats:
            domain_stats[ref_d] = {
                'competitors': set(),
                'urls': [],
                'max_dr': 0,
                'max_traffic': 0,
                'is_dofollow': 0,
                'is_spam': 0,
                'total_links': 0,
                'comp_link_counts': {},  # <--- 新增：按竞品拆解外链数
                'link_details': set()    # <--- 新增：存储全部链接对和锚文本
            }
            
        d = domain_stats[ref_d]
        d['competitors'].add(comp)
        d['urls'].append(ref_u)
        d['max_dr'] = max(d['max_dr'], dr)
        d['max_traffic'] = max(d['max_traffic'], traf)
        d['total_links'] += links_in_group 
        
        # 新增1：记录各个竞品在这个域名下具体有多少条链接
        d['comp_link_counts'][comp] = d['comp_link_counts'].get(comp, 0) + links_in_group
        
        # 新增2：解析 JSON 抓取锚文本，拼接格式
        try:
            raw_dict = json.loads(raw_data)
            # Ahrefs 导出文件中的锚文本字段通常为 Link anchor 或 Anchor
            anchor = raw_dict.get('Link anchor') or raw_dict.get('Anchor') or raw_dict.get('anchor') or "无锚文本/图片链接"
        except:
            anchor = "未知"
            
        detail_str = f"[{comp}] {ref_u} ---> {tgt_u} (锚文本: {anchor})"
        d['link_details'].add(detail_str)
        if is_dof == 1: d['is_dofollow'] = 1
        if is_spam == 1: d['is_spam'] = 1

    # 计算分数
    scored_domains = []
    for ref_d, d in domain_stats.items():
        alpha_count = len(d['competitors'])
        
        # 1. 基础维度算分
        score_alpha = get_alpha_score(alpha_count)
        score_dr = get_dr_score(d['max_dr'])
        score_traffic = get_traffic_score(d['max_traffic'])
        
        # 2. 权重组合 (满分100 = Alpha 35% + DR 30% + Traffic 15% + Dofollow 20%)
        base_score = (score_alpha * 0.35) + (score_dr * 0.30) + (score_traffic * 0.15)
        dofollow_bonus = 20 if d['is_dofollow'] == 1 else 0
        auto_score = round(base_score + dofollow_bonus, 1)
        
        # 3. Spam 一票否决
        if d['is_spam'] == 1:
            auto_score = -100
            
        # 4. 判断优先级
        if auto_score >= 80: priority = "P0 (核心狙击)"
        elif auto_score >= 60: priority = "P1 (重点跟进)"
        elif auto_score >= 40: priority = "P2 (日常铺垫)"
        elif auto_score > 0: priority = "P3 (长尾观察)"
        else: priority = "🗑️ SPAM (直接放弃)"
        
        # 5. 意图识别 (传入了 ref_d 域名变量)
        link_type = guess_link_type(ref_d, d['urls'])
        
        # 6. 行动建议 (配合企业级分类库，提供极具针对性的落地 SOP)
        action_suggest = "获取邮箱并发送破冰 Pitch 邮件"  # 默认兜底
        
        if "App Market" in link_type: action_suggest = "技术/产品部：研发并上架对应平台的 App/插件"
        elif "Social / Video" in link_type: action_suggest = "社媒运营：注册官方账号留链接，或联系博主商单带货"
        elif "Directory / Review" in link_type: action_suggest = "基础运营：无需沟通，直接认领企业主页并刷几条好评"
        elif "Business Directory" in link_type: action_suggest = "基础运营：提交企业基础信息(NAP: 名字/地址/网址)"
        elif "Forum" in link_type or "Thread" in link_type: action_suggest = "社媒运营：注册账号，养号并参与问答(带外链)"
        elif "Ecommerce" in link_type: action_suggest = "商务部：注册卖家/供应商账号，或寻求分销合作"
        elif "Coupon" in link_type: action_suggest = "联盟营销：提交 Arkswift 独家折扣码给平台编辑"
        elif "News Media" in link_type: action_suggest = "PR/公关部：撰写高质量新闻稿，联系记者或通过PR平台分发"
        elif "Education" in link_type: action_suggest = "市场部：提供针对学生的专属方案，或提交教学案例获取 .edu 链接"
        elif "Affiliate" in link_type or "Comparison" in link_type or "Review" in link_type: 
            action_suggest = "联盟营销：联系站长提议加入 Arkswift 联盟计划(CPS)并请求测评"
        elif "Guest Post" in link_type: action_suggest = "内容SEO：按对方指南(Guidelines)撰写高质量行业文章投稿"
        elif "Resource Page" in link_type: action_suggest = "内容SEO：邮件联系编辑，请求将 Arkswift 加入该资源库"
        elif "Listicle" in link_type: action_suggest = "内容SEO：邮件联系作者，请求将 Arkswift 补充进入 Top 榜单"
        elif "Partner" in link_type: action_suggest = "BD/商务部：联系对方 Partner 团队，探讨 API 集成或联合营销"
        elif "Profile" in link_type: action_suggest = "基础运营：注册免费账户，在个人 Bio/主页留下外链"
        elif "Directory Listing" in link_type: action_suggest = "基础运营：提交工具/企业信息入驻该黄页目录"

        scored_domains.append({
            'priority': priority,
            'domain': ref_d,
            'auto_score': auto_score,
            'link_type': link_type,
            'action': action_suggest,
            'alpha': alpha_count,
            # 将总外链格式化为：总计:751 (doba.com:500, bigbuy.eu:251)
            'total_links': f"总计:{d['total_links']} (" + ", ".join([f"{k}:{v}" for k, v in d['comp_link_counts'].items()]) + ")",
            'dr': d['max_dr'],
            'traffic': d['max_traffic'],
            'dofollow': "✅ Yes" if d['is_dofollow'] else "❌ No",
            'spam': "⚠️ Yes" if d['is_spam'] else "No",
            # 用回车符拼出所有竞品的链接和锚文本
            'example_url': "\n\n".join(list(d['link_details'])) 
        })

    # 排序：优先度高 -> 分数高 -> DR高
    scored_domains.sort(key=lambda x: (x['auto_score'], x['dr'], x['traffic']), reverse=True)

    # ==========================================
    # 第三阶段：输出带 AI/人工 交接带的 CRM Excel
    # ==========================================
    print("[进程] 正在生成带有 AI/人工 工作流区域的最终 CRM 报表...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "外链获取执行队列 (CRM)"

    # 定义三种颜色的表头，用于物理分割不同阶段的工作流
    headers = [
        # --- Python 自动生成区 (灰底) ---
        ("A", "优先级 (Priority)", "A6A6A6"),
        ("B", "目标域名 (Domain)", "A6A6A6"),
        ("C", "🤖机器评分 (Auto Score)", "A6A6A6"),
        ("D", "内容类型 (Link Type)", "A6A6A6"),
        ("E", "建议策略 (Action Suggested)", "A6A6A6"),
        ("F", "竞品重合度 (Alpha)", "A6A6A6"),
        ("G", "🔗该域名下总外链 (Total Links)", "A6A6A6"), # <--- 新增列插在这里
        ("H", "DR权重", "A6A6A6"),
        ("I", "页面流量", "A6A6A6"),
        ("J", "Dofollow", "A6A6A6"),
        ("K", "Spam", "A6A6A6"),
        ("L", "对标链接样例", "A6A6A6"),
        
        # --- AI 预留处理区 (绿底) ---
        ("M", "🧠AI: 业务相关性评分(0-100)", "70AD47"),
        ("N", "🧠AI: 配套内容(如Pitch邮件/大纲)", "70AD47"),
        ("O", "🧠AI: 难度评级(Easy/Hard)", "70AD47"),
        
        # --- 人工执行区 (蓝底) ---
        ("P", "🧑‍💻人工: 最终获取难度评分(0-100)", "4472C4"),
        ("Q", "🏆总分 (Total Score) [自动计算]", "FFC000"),  
        ("R", "🧑‍💻联系邮箱 (Email)", "4472C4"),
        ("S", "🧑‍💻联系页面 (Contact URL)", "4472C4"),
        ("T", "🧑‍💻当前状态 (Status)", "4472C4")
    ]

    # 写入表头并上色
    ws.append([h[1] for h in headers])
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=h[2])
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for idx, d in enumerate(scored_domains, start=2):
        row_data = [
            d['priority'],
            d['domain'],
            d['auto_score'],
            d['link_type'],
            d['action'],
            d['alpha'],
            d['total_links'],  # <--- 新字段对应 G 列
            d['dr'],
            d['traffic'],
            d['dofollow'],
            d['spam'],
            d['example_url'],
            "", # M: AI 相关性
            "", # N: AI 角度
            "", # O: AI 难度
            "", # P: 人工 难度
        ]
        ws.append(row_data)
        # 强行让 G列(外链拆解) 和 L列(全部对标链接) 自动换行并居中，确保排版整洁
        ws.cell(row=idx, column=7).alignment = Alignment(wrap_text=False, vertical="center")
        ws.cell(row=idx, column=12).alignment = Alignment(wrap_text=False, vertical="center")
        
        # Q列(第17列) 注入Excel公式，列号已更新：M列是AI，P列是人工
        formula = f"=(C{idx}*0.8) + (IF(ISNUMBER(M{idx}),M{idx},0)*0.15) + (IF(ISNUMBER(P{idx}),P{idx},0)*0.05)"
        ws.cell(row=idx, column=17, value=formula).font = Font(bold=True, color="C00000")
        
        # 预留最后几个空位
        ws.cell(row=idx, column=18, value="") # Email
        ws.cell(row=idx, column=19, value="") # Contact URL
        ws.cell(row=idx, column=20, value="未开始 (Not Started)") # Status

    # 调整列宽 (加入了新列 G)
    column_widths = {
        'A': 15, 'B': 25, 'C': 18, 'D': 25, 'E': 30, 
        'F': 15, 'G': 40, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 90, # G改40，L改90
        'M': 25, 'N': 35, 'O': 20, 
        'P': 28, 'Q': 25, 'R': 25, 'S': 30, 'T': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    out_file = os.path.join(OUTPUT_DIR, f"Backlink_Execution_Queue_{date_str}.xlsx")
    wb.save(out_file)
    
    print(f"\n[✓] 成功生成外链执行队列 (CRM级): {out_file}")
    print("--------------------------------------------------")
    print("下一步工作流建议 (SOP)：")
    print("1. [已完成] 机器已过滤掉大量低质外链，计算出了基础分 (Auto Score) 和链路类型。")
    print("2. [待执行] 将 L、M、N 列提取为 JSON 交给 GPT，批量判断业务相关性，并将数据粘回表格。")
    print("3. [待执行] 人工根据 P 列的最终得分(降序)，锁定 Top 500 开始手工寻找邮箱(Q、R列)，并推动状态流转(S列)。")
    print("--------------------------------------------------")

if __name__ == "__main__":
    generate_execution_queue()
